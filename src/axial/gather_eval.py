"""Phase A v1 slice 08 (issue #478): score Gather's disagreement entries on
grounding, calibrated against the founder (D15/D17, DEC-54, the ruling in
#447).

Phase A v1 has no measure of whether a Gather disagreement (`axial names
gather`, `data/names/disagreements.jsonl`) is real. DEC-54 settled the unit
(the disagreement entry) and the reference (grounding, checked against the
provenance Gather already persists, not a hand-labeled key): an entry names
positions and attributes them to authors, and every entry already carries
the `chunk_ids` behind it, so "is this real?" decomposes into two checkable
questions --

  1. **Attribution** -- is each position the entry attributes to an author
     actually present in a note that was cited for it?
  2. **Conflict** -- do the attributed positions actually oppose each other,
     or is agreement being reported as disagreement?

**No Gather re-run.** This module rebuilds the exact packet Gather saw for
one entry from its own stored `chunk_ids`, reusing `axial.gather`'s own
`build_packets`/`render_packet`/`MemberPacket` (and, for the null re-ask
below, `GatherJob`/`split_into_batches`/`_gather_one` too) rather than
re-deriving any of it. The 447 disagreement entries are a fixed,
re-runnable set precisely because this reads what is already on disk.

**One record per judgment (D17).** Every judge call appends one record to
`DEFAULT_JUDGMENTS_PATH` (`data/names/gather_eval.jsonl`, alongside
`disagreements.jsonl`) carrying what was shown (the `name_key` and
`chunk_ids` that reconstruct it), what came back (`attribution_grounded`,
`conflict_grounded`, `explanation`), and the `name_key` that ties it back to
the entry. That record IS the resume checkpoint -- one file, not two --
content-keyed the way `axial.gather.GatherJob.key` is: `eval_key` hashes the
disagreement record's own `name_key` together with its `disagreement` text,
so an unchanged entry never re-asks and a re-run of Gather that draws a
different text (DEC-54: `France` flipped null<->finding on identical packets
at temperature 1) always does.

**Calibration before the full run.** `run_gather_eval_sheet` emits a
~30-entry sheet, stratified across `MEMBER_COUNT_BANDS`, in the shape
`axial.gold` already uses for founder marking (a workbook, one row per
entry, a blank dropdown column). `run_gather_eval_score` reads the founder's
returned marks (`<calibration_dir>/labels/label_sheet.xlsx`) and reports
judge-vs-founder agreement -- the same judge call and the same checkpoint,
so scoring the calibration 30 costs nothing extra once `score` reaches the
rest of the 447.

**The null re-ask bypasses the checkpoint on purpose.** `_recheck_nulls`
samples ~50 entries whose `disagreement` is `None` and calls
`axial.gather._gather_one` directly on a freshly rebuilt `GatherJob` --
`_gather_one` itself never touches disk (only `run_gather`'s own
result-collecting loop appends), so this reproduces exactly what a live
Gather re-run would do without ever consulting or writing
`disagreements.jsonl`. Routing this through the checkpoint-keyed judge path
instead would just return the cached null and measure nothing -- the
easiest way to ship a broken instrument, per the issue's own warning.

Out of scope (issue #478): setting the ship-blocking grounding-rate bar (a
follow-up once this run has a number), the 18,198 names below the
`gather.min_members` scope cut, and Phase B answer quality.
"""

from __future__ import annotations

import hashlib
import json
import random
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from axial.checkpoint import append_checkpoint_record, load_checkpoint_records
from axial.gather import (
    DEFAULT_DISAGREEMENTS_PATH,
    DEFAULT_WORKERS,
    DisagreementRecordsCorruptError,
    GatherJob,
    MemberPacket,
    _gather_one,
    _utc_now,
    build_packets,
    load_gather_context,
    render_packet,
    split_into_batches,
)
from axial.intake import SOURCE_META_DIR
from axial.interrogate import _default_answers_dir
from axial.llm import GATHER_EVAL_PASS_NAME, LLMClient, get_client
from axial.model_json import ModelJsonError, complete_json, parse_model_json
from axial.names import DEFAULT_NAMES_DATA_DIR
from axial.paths import DEFAULT_PIPELINE_CONFIG_PATH

# The judgment record AND its own resume checkpoint, one file, mirroring
# `disagreements.jsonl`'s own convention (module docstring, D17).
DEFAULT_JUDGMENTS_PATH = DEFAULT_NAMES_DATA_DIR / "gather_eval.jsonl"

# Where the calibration sheet is written and where the founder's returned,
# marked copy is read back from -- mirrors `axial.gold`'s
# `data/gold/label_sheet.xlsx` -> `data/gold/labels/label_sheet.xlsx` shape.
DEFAULT_GATHER_EVAL_DIR = Path("data/gather_eval")

# DEC-54's own stratification (issue #447 -> #478), not a tuned heuristic:
# the member-count bands the founder named for the calibration sample.
MEMBER_COUNT_BANDS: tuple[tuple[str, int, int | None], ...] = (
    ("10-20", 10, 20),
    ("20-50", 20, 50),
    ("50-100", 50, 100),
    ("100-300", 100, 300),
    ("300+", 300, None),
)

# Issue #478's own targets: "~30" for the calibration sheet, "~50" for the
# null re-ask sample.
CALIBRATION_SAMPLE_SIZE = 30
NULL_SAMPLE_SIZE = 50

DEFAULT_SEED = 0


class GatherEvalError(Exception):
    """Base class for all gather-eval errors."""


class EmptyDisagreementsError(GatherEvalError):
    """Raised when `disagreements_path` carries no records at all -- `axial
    names gather` must run first."""

    def __init__(self, path: Path):
        self.path = path
        super().__init__(f"no disagreement records found at {path}; run `axial names gather` first")


class GatherEvalResponseError(GatherEvalError):
    """Raised when a judge response is not usable as a grounding verdict --
    a shape failure, so it is re-askable within `complete_json`'s own
    bounded budget (mirrors `axial.gather.GatherResponseError`)."""


class JudgmentRecordsCorruptError(GatherEvalError):
    def __init__(self, path: Path, line_no: int, cause: json.JSONDecodeError):
        super().__init__(
            f"gather-eval judgment record {path} is corrupt at line {line_no}: {cause}"
        )


# ---------------------------------------------------------------------------
# Member-count bands and seeded, stratified sampling
# ---------------------------------------------------------------------------


def member_count_band(member_count: int) -> str | None:
    """Which of `MEMBER_COUNT_BANDS` `member_count` falls in, or `None` when
    it clears none of them (below the `gather.min_members` scope cut -- not
    expected among scored entries, since Gather never asks about a name that
    small)."""
    for label, low, high in MEMBER_COUNT_BANDS:
        if member_count >= low and (high is None or member_count < high):
            return label
    return None


def _member_count(record: dict[str, Any]) -> int:
    return len(record.get("chunk_ids") or [])


def stratified_sample(records: list[dict[str, Any]], size: int, seed: int) -> list[dict[str, Any]]:
    """A real seeded stratified sample of up to `size` records, round-robined
    across `MEMBER_COUNT_BANDS` in a fixed order so every band represented in
    `records` gets a share, sorted by `name_key` before shuffling so the same
    `seed` always replays the identical sample (DEC-53's own warning: `axial
    names gather --limit` looked like sampling and was not -- this is)."""
    ordered = sorted(records, key=lambda r: r["name_key"])
    bands: dict[str, list[dict[str, Any]]] = {
        label: [] for label, _low, _high in MEMBER_COUNT_BANDS
    }
    for record in ordered:
        label = member_count_band(_member_count(record))
        if label is not None:
            bands[label].append(record)

    rng = random.Random(seed)
    for members in bands.values():
        rng.shuffle(members)

    order = [label for label, _low, _high in MEMBER_COUNT_BANDS]
    selected: list[dict[str, Any]] = []
    index = 0
    while len(selected) < size and any(bands.values()):
        label = order[index % len(order)]
        if bands[label]:
            selected.append(bands[label].pop())
        index += 1

    selected.sort(key=lambda r: r["name_key"])
    return selected


def seeded_sample(records: list[dict[str, Any]], size: int, seed: int) -> list[dict[str, Any]]:
    """A plain seeded sample of up to `size` records (used for the null
    re-ask, which issue #478 does not ask to be band-stratified): sorted by
    `name_key` for a stable base order, then shuffled by `seed` so the same
    seed always replays the identical sample."""
    ordered = sorted(records, key=lambda r: r["name_key"])
    rng = random.Random(seed)
    rng.shuffle(ordered)
    return ordered[:size]


# ---------------------------------------------------------------------------
# The judge call: rebuild the packet Gather saw, ask the two questions
# ---------------------------------------------------------------------------

_JUDGE_PROMPT_TEMPLATE = """\
Below are passages Gather read about {name}, and the disagreement it wrote \
about what these authors think.

PASSAGES
{members}

DISAGREEMENT GATHER WROTE
{disagreement}

Judge two things about this disagreement, using ONLY the passages above.

1. Attribution: does the disagreement attribute a position to an author \
that is actually present in one of that author's OWN passages above? A \
position that is invented, or attributed to the wrong author, fails \
attribution.
2. Conflict: do the positions the disagreement attributes actually oppose \
each other, or is it reporting agreement (or unrelated claims) as if it \
were a disagreement?

RESPONSE. Reply with ONLY a JSON object, no prose and no markdown fences:
{{"attribution_grounded": <true or false>, "conflict_grounded": <true or \
false>, "explanation": "<one or two sentences, citing the passages where \
useful>"}}
"""


def compose_judge_prompt(canonical: str, packets: list[MemberPacket], disagreement: str) -> str:
    """One entry's judge prompt: the same member packets Gather itself was
    shown (reusing `render_packet` verbatim, never a note's full text) plus
    the disagreement text Gather wrote from them."""
    members = "\n".join(f"- {render_packet(packet)}" for packet in packets)
    return _JUDGE_PROMPT_TEMPLATE.format(
        name=repr(canonical), members=members, disagreement=disagreement
    )


def parse_judge_response(raw: str) -> dict[str, Any]:
    """Parse one judge response into `{attribution_grounded, conflict_grounded,
    explanation}`. Raises `GatherEvalResponseError` on a shape failure (not
    an object, a missing/non-boolean verdict key), which is re-asked by
    `complete_json` exactly like `axial.gather.parse_gather_response`'s own
    shape failures."""
    try:
        data = parse_model_json(raw)
    except ModelJsonError as exc:
        raise GatherEvalResponseError(f"judge response was not valid JSON: {exc}") from exc
    if not isinstance(data, dict):
        raise GatherEvalResponseError("judge response must be a JSON object")
    for key in ("attribution_grounded", "conflict_grounded"):
        if not isinstance(data.get(key), bool):
            raise GatherEvalResponseError(f"judge response carried no boolean {key!r}")
    explanation = data.get("explanation")
    if not isinstance(explanation, str):
        explanation = ""
    return {
        "attribution_grounded": data["attribution_grounded"],
        "conflict_grounded": data["conflict_grounded"],
        "explanation": explanation.strip(),
    }


def eval_key(record: dict[str, Any]) -> str:
    """This judgment's own content-addressed checkpoint key (D17): a hash of
    the disagreement record's `name_key` (itself a hash of the rendered
    packets Gather saw, `axial.gather.GatherJob.key`) together with the
    disagreement TEXT. The text alone can change across a Gather re-run over
    unchanged packets (DEC-54: `France` flipped null<->finding on identical
    packets at temperature 1), so `name_key` alone would under-invalidate;
    hashing both means unchanged input never re-asks and changed input
    always does, mirroring `GatherJob.key`'s own reasoning exactly."""
    payload = f"{record['name_key']}\n{record['disagreement']}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _judge_one(
    record: dict[str, Any], packets: list[MemberPacket], client: LLMClient
) -> tuple[dict[str, Any], dict[str, Any] | None, str | None]:
    """Judge one disagreement entry. Runs on a worker thread; never writes
    to disk (mirrors `axial.gather._gather_one`'s own shape) -- the caller's
    single result-collecting thread does the append."""
    if not packets:
        return record, None, "no member packets could be reconstructed from this record's chunk_ids"
    prompt = compose_judge_prompt(record["canonical"], packets, record["disagreement"])
    try:
        raw = complete_json(
            client, prompt, pass_name=GATHER_EVAL_PASS_NAME, validate=parse_judge_response
        )
    except (ModelJsonError, GatherEvalResponseError) as exc:
        return record, None, str(exc)
    judged = parse_judge_response(raw)
    member_count = _member_count(record)
    judgment = {
        "name_key": record["name_key"],
        "eval_key": eval_key(record),
        "canonical": record["canonical"],
        "member_count": member_count,
        "band": member_count_band(member_count),
        "chunk_ids": record["chunk_ids"],
        "disagreement": record["disagreement"],
        "attribution_grounded": judged["attribution_grounded"],
        "conflict_grounded": judged["conflict_grounded"],
        "grounded": bool(judged["attribution_grounded"] and judged["conflict_grounded"]),
        "explanation": judged["explanation"],
        "pass": GATHER_EVAL_PASS_NAME,
        "model": client.model_for_pass(GATHER_EVAL_PASS_NAME),
        "judged_at": _utc_now(),
    }
    return record, judgment, None


def load_judgments(path: Path) -> dict[str, dict[str, Any]]:
    """Every recorded judgment, keyed by `eval_key` -- `{}` before the first
    run (mirrors `axial.gather.load_disagreements`)."""
    records = load_checkpoint_records(path, JudgmentRecordsCorruptError)
    return {record["eval_key"]: record for record in records}


# ---------------------------------------------------------------------------
# The calibration sheet: emit, and read the founder's marks back
# ---------------------------------------------------------------------------

CALIBRATION_SHEET_COLUMNS = (
    "name_key",
    "canonical",
    "band",
    "member_count",
    "disagreement",
    "evidence",
    "grounded",
    "notes",
)

# Filled by the founder: a yes/no verdict and optional free text -- arrive
# blank, same "blind" convention `axial.gold.BLIND_COLUMNS` uses.
_CALIBRATION_BLIND_COLUMNS = ("grounded", "notes")

_NAME_KEY_COLUMN = CALIBRATION_SHEET_COLUMNS.index("name_key") + 1
_GROUNDED_COLUMN = CALIBRATION_SHEET_COLUMNS.index("grounded") + 1


def _build_calibration_workbook(rows: list[dict[str, Any]]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "calibration"

    for col, name in enumerate(CALIBRATION_SHEET_COLUMNS, start=1):
        sheet.cell(row=1, column=col, value=name)

    for row_index, row in enumerate(rows, start=2):
        for col, name in enumerate(CALIBRATION_SHEET_COLUMNS, start=1):
            if name in _CALIBRATION_BLIND_COLUMNS:
                continue
            sheet.cell(row=row_index, column=col, value=row.get(name))

    validation = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    sheet.add_data_validation(validation)
    last_row = 1 + len(rows)
    letter = get_column_letter(_GROUNDED_COLUMN)
    validation.add(f"{letter}2:{letter}{last_row}")

    return workbook


def run_gather_eval_sheet(
    *,
    disagreements_path: Path | None = None,
    calibration_dir: Path | None = None,
    answers_dir: Path | None = None,
    source_meta_dir: Path | None = None,
    config_path: Path = DEFAULT_PIPELINE_CONFIG_PATH,
    sample_size: int = CALIBRATION_SAMPLE_SIZE,
    seed: int = DEFAULT_SEED,
) -> Path:
    """Emit `<calibration_dir>/calibration_sheet.xlsx`: a real seeded
    stratified sample of `sample_size` disagreement entries across
    `MEMBER_COUNT_BANDS`, each row carrying the entry's own evidence (the
    member packets Gather saw, rebuilt from its `chunk_ids`) and disagreement
    text, plus a blank `grounded` dropdown for the founder to mark. Offline
    -- no LLM call. Returns the written path."""
    disagreements_path = (
        Path(disagreements_path) if disagreements_path is not None else DEFAULT_DISAGREEMENTS_PATH
    )
    calibration_dir = (
        Path(calibration_dir) if calibration_dir is not None else DEFAULT_GATHER_EVAL_DIR
    )
    answers_dir = (
        Path(answers_dir) if answers_dir is not None else _default_answers_dir(config_path)
    )
    source_meta_dir = Path(source_meta_dir) if source_meta_dir is not None else SOURCE_META_DIR

    records = load_checkpoint_records(disagreements_path, DisagreementRecordsCorruptError)
    scoreable = [record for record in records if record.get("disagreement") is not None]
    if not scoreable:
        raise EmptyDisagreementsError(disagreements_path)

    sample = stratified_sample(scoreable, sample_size, seed)
    answers_by_chunk_id, author_year = load_gather_context(answers_dir, source_meta_dir)

    rows = []
    for record in sample:
        packets = build_packets(record["chunk_ids"], answers_by_chunk_id, author_year)
        member_count = _member_count(record)
        rows.append(
            {
                "name_key": record["name_key"],
                "canonical": record["canonical"],
                "band": member_count_band(member_count),
                "member_count": member_count,
                "disagreement": record["disagreement"],
                "evidence": "\n".join(render_packet(packet) for packet in packets),
            }
        )

    workbook = _build_calibration_workbook(rows)
    calibration_dir.mkdir(parents=True, exist_ok=True)
    sheet_path = calibration_dir / "calibration_sheet.xlsx"
    workbook.save(sheet_path)
    return sheet_path


def _normalize_grounded_cell(value: Any) -> bool | None:
    """A returned sheet's `grounded` cell, read back: `yes`/`no`
    (case-insensitive, whitespace-tolerant) map to `True`/`False`; a blank
    or unrecognized cell is `None` -- not yet marked, excluded from
    agreement."""
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in ("yes", "y", "true", "grounded"):
        return True
    if text in ("no", "n", "false", "not grounded", "not-grounded", "ungrounded"):
        return False
    return None


def load_calibration_marks(sheet_path: Path) -> dict[str, bool | None]:
    """Read a returned, founder-marked calibration sheet into
    `{name_key: grounded_or_None}`."""
    worksheet = load_workbook(sheet_path, read_only=True).worksheets[0]
    marks: dict[str, bool | None] = {}
    for row in worksheet.iter_rows(min_row=2):
        raw_key = row[_NAME_KEY_COLUMN - 1].value
        name_key = str(raw_key).strip() if raw_key is not None else ""
        if not name_key:
            continue
        marks[name_key] = _normalize_grounded_cell(row[_GROUNDED_COLUMN - 1].value)
    return marks


def _score_calibration(
    calibration_dir: Path, judgments_by_key: dict[str, dict[str, Any]]
) -> dict[str, Any] | None:
    """Judge-vs-founder agreement over the calibration sheet, or `None` when
    no marked sheet has been returned yet -- `score` still runs (and still
    scores the calibration 30 through the same judge+checkpoint as every
    other entry), it just cannot report agreement until the founder's marks
    exist."""
    sheet_path = calibration_dir / "labels" / "label_sheet.xlsx"
    if not sheet_path.is_file():
        return None

    marks = load_calibration_marks(sheet_path)
    by_name_key: dict[str, dict[str, Any]] = {}
    for judgment in sorted(judgments_by_key.values(), key=lambda j: j["judged_at"]):
        by_name_key[judgment["name_key"]] = judgment  # latest judgment per name_key wins

    joined = 0
    agreements = 0
    unmarked = 0
    unjudged = 0
    for name_key, founder_grounded in marks.items():
        if founder_grounded is None:
            unmarked += 1
            continue
        judgment = by_name_key.get(name_key)
        if judgment is None:
            unjudged += 1
            continue
        joined += 1
        if judgment["grounded"] == founder_grounded:
            agreements += 1

    return {
        "sheet_path": str(sheet_path),
        "marked": len(marks) - unmarked,
        "unjudged": unjudged,
        "joined": joined,
        "agreement": (agreements / joined) if joined else None,
    }


# ---------------------------------------------------------------------------
# The null re-ask: bypasses the checkpoint on purpose
# ---------------------------------------------------------------------------


def _recheck_nulls(
    sample: list[dict[str, Any]],
    answers_by_chunk_id: dict[str, dict[str, Any]],
    author_year: dict[str, tuple[Any, Any]],
    client: LLMClient | None,
) -> dict[str, Any]:
    """Re-ask each of `sample`'s null entries at Gather's own settings,
    calling `axial.gather._gather_one` directly on a freshly rebuilt
    `GatherJob` -- never `run_gather`, never `disagreements.jsonl`. Report
    how many flip from null to a real finding."""
    if not sample:
        return {"sampled": 0, "inconclusive": 0, "flipped": 0, "flip_rate": None, "flips": []}

    flips: list[dict[str, Any]] = []
    inconclusive = 0
    for record in sample:
        packets = build_packets(record["chunk_ids"], answers_by_chunk_id, author_year)
        if not packets:
            inconclusive += 1
            continue
        batches = split_into_batches(packets)
        job = GatherJob(record["canonical"], tuple(tuple(batch) for batch in batches))
        _job, new_record, failure_reason = _gather_one(job, client)
        if new_record is None:
            print(
                f"gather-eval: null re-ask of {record['canonical']!r} failed: {failure_reason}",
                file=sys.stderr,
            )
            inconclusive += 1
            continue
        if new_record["disagreement"] is not None:
            flips.append(
                {
                    "name_key": record["name_key"],
                    "canonical": record["canonical"],
                    "new_disagreement": new_record["disagreement"],
                }
            )

    scored = len(sample) - inconclusive
    return {
        "sampled": len(sample),
        "inconclusive": inconclusive,
        "flipped": len(flips),
        "flip_rate": (len(flips) / scored) if scored else None,
        "flips": flips,
    }


# ---------------------------------------------------------------------------
# The full run: judge everything scoreable, score calibration, re-ask nulls
# ---------------------------------------------------------------------------


def _summarize(
    scored: list[dict[str, Any]],
) -> tuple[float | None, dict[str, float | None], float | None]:
    if not scored:
        empty_bands = {label: None for label, _low, _high in MEMBER_COUNT_BANDS}
        return None, empty_bands, None

    grounding_rate = sum(1 for j in scored if j["grounded"]) / len(scored)
    conflict_rate = sum(1 for j in scored if j["conflict_grounded"]) / len(scored)

    by_band: dict[str, float | None] = {}
    for label, _low, _high in MEMBER_COUNT_BANDS:
        band_scored = [j for j in scored if j["band"] == label]
        by_band[label] = (
            (sum(1 for j in band_scored if j["grounded"]) / len(band_scored))
            if band_scored
            else None
        )

    return grounding_rate, by_band, conflict_rate


def run_gather_eval_score(
    *,
    disagreements_path: Path | None = None,
    judgments_path: Path | None = None,
    calibration_dir: Path | None = None,
    answers_dir: Path | None = None,
    source_meta_dir: Path | None = None,
    config_path: Path = DEFAULT_PIPELINE_CONFIG_PATH,
    client: LLMClient | None = None,
    limit: int | None = None,
    null_sample_size: int = NULL_SAMPLE_SIZE,
    seed: int = DEFAULT_SEED,
    workers: int = DEFAULT_WORKERS,
) -> dict[str, Any]:
    """Judge every scoreable disagreement entry not already in the checkpoint
    (`judgments_path`), score the founder-marked calibration sheet if one has
    been returned, and re-ask a seeded sample of null entries bypassing
    Gather's own checkpoint. `limit` is a cheap smoke flag on the main judge
    loop only (DEC-53: never the sampling mechanism) -- the calibration join
    and the null re-ask always use their own seeded samples, untouched by it.
    """
    disagreements_path = (
        Path(disagreements_path) if disagreements_path is not None else DEFAULT_DISAGREEMENTS_PATH
    )
    judgments_path = Path(judgments_path) if judgments_path is not None else DEFAULT_JUDGMENTS_PATH
    calibration_dir = (
        Path(calibration_dir) if calibration_dir is not None else DEFAULT_GATHER_EVAL_DIR
    )
    answers_dir = (
        Path(answers_dir) if answers_dir is not None else _default_answers_dir(config_path)
    )
    source_meta_dir = Path(source_meta_dir) if source_meta_dir is not None else SOURCE_META_DIR

    all_records = load_checkpoint_records(disagreements_path, DisagreementRecordsCorruptError)
    if not all_records:
        raise EmptyDisagreementsError(disagreements_path)
    scoreable = [record for record in all_records if record.get("disagreement") is not None]
    null_entries = [record for record in all_records if record.get("disagreement") is None]

    answers_by_chunk_id, author_year = load_gather_context(answers_dir, source_meta_dir)

    judgments_by_key = load_judgments(judgments_path)
    pending = [record for record in scoreable if eval_key(record) not in judgments_by_key]
    pending.sort(key=lambda r: r["canonical"])
    to_attempt = pending if limit is None else pending[:limit]

    null_sample = seeded_sample(null_entries, null_sample_size, seed)

    if client is None and (to_attempt or null_sample):
        client = get_client(config_path=config_path)

    called = 0
    failed = 0
    if to_attempt:
        with ThreadPoolExecutor(max_workers=max(workers, 1)) as executor:
            futures = {
                executor.submit(
                    _judge_one,
                    record,
                    build_packets(record["chunk_ids"], answers_by_chunk_id, author_year),
                    client,
                ): record
                for record in to_attempt
            }
            for future in as_completed(futures):
                record, judgment, failure_reason = future.result()
                if judgment is None:
                    print(
                        f"gather-eval: {record['canonical']!r} judge failed: {failure_reason}",
                        file=sys.stderr,
                    )
                    failed += 1
                    continue
                append_checkpoint_record(judgments_path, judgment)
                judgments_by_key[judgment["eval_key"]] = judgment
                called += 1

    live_keys = {eval_key(record) for record in scoreable}
    scored = [
        judgment for judgment in judgments_by_key.values() if judgment["eval_key"] in live_keys
    ]

    grounding_rate, grounding_by_band, conflict_rate = _summarize(scored)
    calibration = _score_calibration(calibration_dir, judgments_by_key)
    # `_recheck_nulls` returns before ever touching `client` when
    # `null_sample` is empty, so a `None` client (never built above, because
    # neither `to_attempt` nor `null_sample` needed one) is safe here.
    null_result = _recheck_nulls(null_sample, answers_by_chunk_id, author_year, client)

    failures = sorted((j for j in scored if not j["grounded"]), key=lambda j: j["canonical"])

    report = {
        "disagreements_path": str(disagreements_path),
        "judgments_path": str(judgments_path),
        "scoreable": len(scoreable),
        "scored": len(scored),
        "asked": called,
        "reused": len(scored) - called,
        "failed": failed,
        "limit": limit,
        "grounding_rate": grounding_rate,
        "grounding_rate_by_band": grounding_by_band,
        "conflict_rate": conflict_rate,
        "calibration": calibration,
        "null_flip": null_result,
        "failures": [
            {
                "name_key": j["name_key"],
                "canonical": j["canonical"],
                "band": j["band"],
                "disagreement": j["disagreement"],
                "attribution_grounded": j["attribution_grounded"],
                "conflict_grounded": j["conflict_grounded"],
                "explanation": j["explanation"],
            }
            for j in failures
        ],
    }

    # A persisted artifact alongside the ephemeral stdout/run.jsonl -- the
    # issue's own "what it reports" list (grounding/conflict rate by band,
    # agreement, null flip rate, quoted failures), so the operator writing
    # the real corpus run's summary.md has a single file to read from
    # (mirrors `axial.eval.run_eval`'s own `eval_report.json`).
    calibration_dir.mkdir(parents=True, exist_ok=True)
    report_path = calibration_dir / "report.json"
    report_path.write_text(json.dumps(report, indent=2, sort_keys=True), encoding="utf-8")
    report["report_path"] = str(report_path)

    return report
