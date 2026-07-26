"""Tests for the one-off source-rename migration (axial.rename_source_ids).

Every test builds its own throwaway corpus under `tmp_path`; nothing here
reads or writes the real `data/`.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from axial.envelope import compute_source_id, content_digest
from axial.gold import LABEL_SHEET_NAME, VOCAB_SHEET_NAME
from axial.paths import (
    artifact_note_path,
    budgeted_chunk_filename,
    chunk_note_path,
    path_overage,
)
from axial.rename_source_ids import (
    AmbiguousHashError,
    ContentHashMismatchError,
    DestinationExistsError,
    FilePlan,
    MalformedRecordError,
    UnmatchedNewSourceError,
    UnmatchedOldSourceIdError,
    apply_file,
    apply_plan,
    build_id_map,
    build_plan,
    format_plan,
    split_note,
    substitute_ids_in_text,
)
from axial.vault import render_note

# A pre-rename download stem, hyphens and all, so the mapping is exercised
# against a stem the hash-suffix anchor has to see past. Shorter than the real
# 181-character worst case, because the real one plus a pytest tmp_path is
# itself over Windows' MAX_PATH -- the very defect this migration exists to
# retire.
LONG_STEM = "Benjamin Thomas White - The Emergence of Minorities - libgen.li"
NEW_STEM = "white-2011"
SOURCE_BYTES = b"%PDF-1.4 pretend book bytes"

# How far over the MAX_PATH budget the fixture's prose note filename is built
# to be. Sized at runtime against the *resolved* vault directory rather than
# written as a fixed slug length, because `path_overage` measures the real
# directory and a pytest tmp_path is ~90 characters on Windows and ~50 on a
# Linux CI runner -- a fixed length silently stops reproducing the condition
# on the other platform (it did: CI went red on exactly that). The margin is
# smaller than the 53 characters LONG_STEM loses when it becomes NEW_STEM, so
# the note is guaranteed budgeted before the rename and unbudgeted after it,
# on every platform.
_OVER_BUDGET_BY = 20


def _slug_over_budget(prose_dir: Path, source_id: str, over_by: int) -> str:
    """A section slug sized so that `<source_id>_2_<slug>_001.md` is exactly
    `over_by` characters over `prose_dir`'s note-path budget."""
    prose_dir.mkdir(parents=True, exist_ok=True)
    padding = over_by - path_overage(prose_dir, f"{source_id}_2__001.md")
    assert padding > 0, f"{prose_dir} is already over budget with an empty slug"
    return ("state-formation-and-comparative-development-" * 20)[:padding]


@dataclass
class Corpus:
    """Paths and ids of one throwaway corpus, before migration."""

    sources_dir: Path
    data_dir: Path
    vault_dir: Path
    gold_dir: Path
    cases_dir: Path
    analyses_dir: Path
    pin_path: Path
    old_id: str
    new_id: str
    old_chunk_id: str
    old_other_chunk_id: str
    old_artifact_id: str
    prose_note: Path
    artifact_note: Path
    gold_record: Path

    def plan(self):
        return build_plan(
            sources_dir=self.sources_dir,
            data_dir=self.data_dir,
            vault_dir=self.vault_dir,
            gold_dir=self.gold_dir,
            cases_dir=self.cases_dir,
            analyses_dir=self.analyses_dir,
            pin_path=self.pin_path,
        )

    def migrate(self):
        plan = self.plan()
        apply_plan(plan)
        return plan

    def rename(self, id_: str) -> str:
        """The id `id_` becomes after the migration."""
        return self.new_id + id_[len(self.old_id) :]


def _write_json(path: Path, payload) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def _write_jsonl(path: Path, records: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "".join(json.dumps(record) + "\n" for record in records),
        encoding="utf-8",
    )


def _write_gold_record(gold_dir: Path, source_id: str, chunk_id: str) -> Path:
    """One sampled gold chunk record, written the way `run_gold_sample` writes
    it: named by chunk_id, budgeted down when that would blow the path limit
    (which it does here, by construction), so its id lives only inside."""
    chunks_dir = gold_dir / "chunks"
    chunks_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{chunk_id}.json"
    if path_overage(chunks_dir, filename) > 0:
        filename = budgeted_chunk_filename(chunks_dir, source_id, chunk_id, extension=".json")
    path = chunks_dir / filename
    path.write_text(
        json.dumps(
            {"chunk_id": chunk_id, "source": source_id, "section": "Abstract"},
            indent=2,
            sort_keys=True,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return path


def _write_label_sheet(path: Path, source_id: str, chunk_id: str) -> None:
    """A label sheet shaped like the real one: a `label_sheet` worksheet whose
    first two columns are the id provenance and whose later columns hold the
    merged human labels, plus the hidden `vocab` sheet its dropdown data
    validation points at."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = LABEL_SHEET_NAME
    for column, name in enumerate(("chunk_id", "source", "section", "claim_type_gold"), start=1):
        sheet.cell(row=1, column=column, value=name)
    for column, value in enumerate((chunk_id, source_id, "Abstract", "ideology"), start=1):
        sheet.cell(row=2, column=column, value=value)

    vocab = workbook.create_sheet(VOCAB_SHEET_NAME)
    vocab.sheet_state = "hidden"
    vocab.cell(row=1, column=1, value="claim_type")
    for row, value in enumerate(("ideology", "coercion"), start=2):
        vocab.cell(row=row, column=1, value=value)

    validation = DataValidation(
        type="list", formula1=f"{VOCAB_SHEET_NAME}!$A$2:$A$3", allow_blank=True
    )
    sheet.add_data_validation(validation)
    validation.add("D2:D2")

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _build_corpus(
    tmp_path: Path,
    *,
    new_stem: str = NEW_STEM,
    old_stem: str = LONG_STEM,
    source_bytes: bytes = SOURCE_BYTES,
) -> Corpus:
    """A miniature but structurally faithful corpus: one renamed raw source,
    one derived file per pipeline directory, one prose note, one artifact
    note, the gold set (chunk record, dispatch, completed labelling output,
    label sheet), one sim eval case, one benchmark analysis log, and a pin."""
    sources_dir = tmp_path / "sources"
    sources_dir.mkdir()
    raw = sources_dir / f"{new_stem}.pdf"
    raw.write_bytes(source_bytes)
    new_id = compute_source_id(raw)
    old_id = f"{old_stem}-{new_id[-12:]}"

    data_dir = tmp_path / "data"
    vault_dir = data_dir / "vault"
    gold_dir = data_dir / "gold"
    slug = _slug_over_budget(vault_dir / "prose", old_id, _OVER_BUDGET_BY)
    chunk_id = f"{old_id}_2_{slug}_001"
    other_chunk_id = f"{old_id}_3_introduction_001"
    artifact_id = f"{old_id}_art_12.3"

    _write_json(data_dir / "trees" / f"{old_id}.json", {"title": "no ids in here"})
    _write_json(data_dir / "envelopes" / f"{old_id}.json", {"source_id": old_id})
    _write_json(
        data_dir / "source_meta" / f"{old_id}.json",
        {"source_id": old_id, "file_hash": content_digest(raw)},
    )
    _write_jsonl(
        data_dir / "chunks" / f"{old_id}.jsonl",
        [{"chunk_id": chunk_id, "text": "prose"}, {"chunk_id": other_chunk_id, "text": "more"}],
    )
    _write_jsonl(
        data_dir / "chunks" / f"{old_id}.skips.jsonl",
        [{"section": "Notes", "reason": "apparatus: endnotes"}],
    )
    _write_jsonl(
        data_dir / "tags" / f"{old_id}.jsonl",
        [{"chunk_id": chunk_id, "source_id": old_id, "field": "field:sociology"}],
    )
    _write_jsonl(
        data_dir / "tags" / "_quarantine_log.jsonl",
        [{"source_id": old_id, "chunk_id": chunk_id, "reason": "parse_error"}],
    )
    _write_jsonl(
        data_dir / "xref" / f"{old_id}.jsonl",
        [{"chunk_id": chunk_id, "target_chunk_id": other_chunk_id, "kind": "prose"}],
    )
    _write_jsonl(
        data_dir / "artifacts" / f"{old_id}.jsonl",
        [{"artifact_id": artifact_id, "source_id": old_id, "artifact_role": "table"}],
    )

    prose_note = chunk_note_path(vault_dir, old_id, chunk_id)
    prose_note.parent.mkdir(parents=True, exist_ok=True)
    prose_note.write_text(
        render_note(
            {
                "chunk_id": chunk_id,
                "section": "Abstract",
                "chunk_text": "The key representatives of the bellicist approach.",
                "artifact_refs": [artifact_id],
            },
            "# Abstract\n\nThe key representatives of the bellicist approach.\n",
        ),
        encoding="utf-8",
    )
    artifact_note = artifact_note_path(vault_dir, old_id, artifact_id)
    artifact_note.parent.mkdir(parents=True, exist_ok=True)
    artifact_note.write_text(
        render_note(
            {
                "artifact_id": artifact_id,
                "source_id": old_id,
                "artifact_role": "table",
                "cited_by": [chunk_id],
            },
            "# Table 3\n\n| a | b |\n",
        ),
        encoding="utf-8",
    )

    gold_record = _write_gold_record(gold_dir, old_id, chunk_id)
    _write_json(
        gold_dir / "dispatch" / "blind_full_120.json",
        [{"chunk_id": chunk_id, "chunk_text": "prose", "source": old_id}],
    )
    # completed human labelling: the chunk_id is the object KEY here
    _write_json(
        gold_dir / "dispatch" / "out" / "blind_draw_1.json",
        {chunk_id: {"claim_type": "ideology-as-system", "notes": "human judgement"}},
    )
    _write_label_sheet(gold_dir / "label_sheet.xlsx", old_id, chunk_id)

    cases_dir = tmp_path / "cases"
    _write_json(
        cases_dir / "P1-01.json",
        {"case_id": "P1-01", "required_citation_source_ids": [old_id], "notes": "keep me"},
    )

    analyses_dir = data_dir / "logs" / "bench" / "analyses"
    _write_json(
        analyses_dir / "P1-01" / "draw0" / "abc123.json",
        {"grounds": [{"chunk_id": chunk_id}, {"chunk_id": other_chunk_id}]},
    )

    pin_path = tmp_path / "pin.json"
    _write_json(
        pin_path,
        {
            "ingest_code_sha": "deadbeef",
            "sources": [{"content_hash": content_digest(raw), "source_id": old_id}],
        },
    )

    return Corpus(
        sources_dir=sources_dir,
        data_dir=data_dir,
        vault_dir=vault_dir,
        gold_dir=gold_dir,
        cases_dir=cases_dir,
        analyses_dir=analyses_dir,
        pin_path=pin_path,
        old_id=old_id,
        new_id=new_id,
        old_chunk_id=chunk_id,
        old_other_chunk_id=other_chunk_id,
        old_artifact_id=artifact_id,
        prose_note=prose_note,
        artifact_note=artifact_note,
        gold_record=gold_record,
    )


# ------------------------------------------------------------------ mapping


def test_id_map_matches_on_hash_suffix_despite_hyphenated_stems(tmp_path: Path):
    corpus = _build_corpus(tmp_path)

    id_map = build_id_map(corpus.sources_dir, corpus.data_dir)

    # both stems contain hyphens; only the shared 12-char hash suffix matched
    assert "-" in corpus.old_id[: -len("-000000000000")]
    assert id_map == {corpus.old_id: corpus.new_id}
    assert corpus.old_id[-12:] == corpus.new_id[-12:]


def test_old_source_id_with_no_matching_raw_file_is_refused(tmp_path: Path):
    corpus = _build_corpus(tmp_path)
    orphan = f"some-other-book-{'a' * 12}"
    (corpus.data_dir / "trees" / f"{orphan}.json").write_text("{}", encoding="utf-8")

    with pytest.raises(UnmatchedOldSourceIdError) as excinfo:
        build_id_map(corpus.sources_dir, corpus.data_dir)

    assert orphan in str(excinfo.value)


def test_raw_file_with_no_derived_data_is_refused(tmp_path: Path):
    corpus = _build_corpus(tmp_path)
    (corpus.sources_dir / "unbuilt-2020.pdf").write_bytes(b"never ingested")

    with pytest.raises(UnmatchedNewSourceError) as excinfo:
        build_id_map(corpus.sources_dir, corpus.data_dir)

    assert "unbuilt-2020" in str(excinfo.value)


def test_two_old_ids_sharing_one_hash_are_refused(tmp_path: Path):
    corpus = _build_corpus(tmp_path)
    twin = f"a-different-old-stem-{corpus.old_id[-12:]}"
    (corpus.data_dir / "trees" / f"{twin}.json").write_text("{}", encoding="utf-8")

    with pytest.raises(AmbiguousHashError) as excinfo:
        build_id_map(corpus.sources_dir, corpus.data_dir)

    assert excinfo.value.hash12 == corpus.old_id[-12:]
    assert twin in str(excinfo.value)


def test_pin_content_hash_mismatch_refuses_and_writes_nothing(tmp_path: Path):
    corpus = _build_corpus(tmp_path)
    # the same 12-char prefix, a different full hash: the bytes moved
    pinned = json.loads(corpus.pin_path.read_text(encoding="utf-8"))
    pinned["sources"][0]["content_hash"] = corpus.old_id[-12:] + "f" * 52
    corpus.pin_path.write_text(json.dumps(pinned), encoding="utf-8")
    before = corpus.prose_note.read_text(encoding="utf-8")

    with pytest.raises(ContentHashMismatchError):
        corpus.plan()

    assert corpus.prose_note.read_text(encoding="utf-8") == before


# ------------------------------------------------------------------ renames


def test_derived_files_are_renamed_and_rewritten(tmp_path: Path):
    corpus = _build_corpus(tmp_path)

    corpus.migrate()

    for subdir, suffix in (
        ("trees", ".json"),
        ("envelopes", ".json"),
        ("source_meta", ".json"),
        ("chunks", ".jsonl"),
        ("chunks", ".skips.jsonl"),
        ("tags", ".jsonl"),
        ("xref", ".jsonl"),
        ("artifacts", ".jsonl"),
    ):
        assert (corpus.data_dir / subdir / f"{corpus.new_id}{suffix}").is_file()
        assert not (corpus.data_dir / subdir / f"{corpus.old_id}{suffix}").exists()

    meta = json.loads(
        (corpus.data_dir / "source_meta" / f"{corpus.new_id}.json").read_text(encoding="utf-8")
    )
    assert meta["source_id"] == corpus.new_id
    assert meta["file_hash"] == content_digest(corpus.sources_dir / f"{NEW_STEM}.pdf")


def test_jsonl_records_including_xref_pairs_are_rewritten(tmp_path: Path):
    corpus = _build_corpus(tmp_path)

    corpus.migrate()

    xref = json.loads(
        (corpus.data_dir / "xref" / f"{corpus.new_id}.jsonl").read_text(encoding="utf-8").strip()
    )
    assert xref == {
        "chunk_id": corpus.rename(corpus.old_chunk_id),
        "target_chunk_id": corpus.rename(corpus.old_other_chunk_id),
        "kind": "prose",
    }

    tags = json.loads(
        (corpus.data_dir / "tags" / f"{corpus.new_id}.jsonl").read_text(encoding="utf-8").strip()
    )
    assert tags["chunk_id"] == corpus.rename(corpus.old_chunk_id)
    assert tags["source_id"] == corpus.new_id


def test_corpus_wide_log_keeps_its_name_but_gets_its_ids_rewritten(tmp_path: Path):
    corpus = _build_corpus(tmp_path)

    corpus.migrate()

    log = corpus.data_dir / "tags" / "_quarantine_log.jsonl"
    record = json.loads(log.read_text(encoding="utf-8").strip())
    assert record["chunk_id"] == corpus.rename(corpus.old_chunk_id)
    assert record["source_id"] == corpus.new_id


def test_eval_cases_and_benchmark_analyses_are_rewritten(tmp_path: Path):
    corpus = _build_corpus(tmp_path)

    corpus.migrate()

    case = json.loads((corpus.cases_dir / "P1-01.json").read_text(encoding="utf-8"))
    assert case["required_citation_source_ids"] == [corpus.new_id]
    assert case["notes"] == "keep me"

    analysis = json.loads(
        (corpus.analyses_dir / "P1-01" / "draw0" / "abc123.json").read_text(encoding="utf-8")
    )
    assert [ground["chunk_id"] for ground in analysis["grounds"]] == [
        corpus.rename(corpus.old_chunk_id),
        corpus.rename(corpus.old_other_chunk_id),
    ]


# -------------------------------------------------------------- vault notes


def test_budgeted_note_filename_is_resolved_by_frontmatter_not_filename(tmp_path: Path):
    corpus = _build_corpus(tmp_path)
    prose_dir = corpus.vault_dir / "prose"
    # Preconditions, asserted loudly so a platform whose temp path no longer
    # reproduces them fails as "this fixture stopped being over budget" rather
    # than as "the tool did not shorten the filename".
    assert path_overage(prose_dir, f"{corpus.old_chunk_id}.md") == _OVER_BUDGET_BY
    assert corpus.prose_note.name != f"{corpus.old_chunk_id}.md"
    assert path_overage(prose_dir, f"{corpus.rename(corpus.old_chunk_id)}.md") < 0

    corpus.migrate()

    assert not corpus.prose_note.exists()
    new_note = chunk_note_path(corpus.vault_dir, corpus.new_id, corpus.rename(corpus.old_chunk_id))
    assert new_note.is_file()
    frontmatter, body = split_note(new_note, new_note.read_text(encoding="utf-8"))
    assert frontmatter["chunk_id"] == corpus.rename(corpus.old_chunk_id)
    assert frontmatter["artifact_refs"] == [corpus.rename(corpus.old_artifact_id)]
    assert body == "# Abstract\n\nThe key representatives of the bellicist approach.\n"


def test_artifact_note_frontmatter_ids_are_rewritten(tmp_path: Path):
    corpus = _build_corpus(tmp_path)
    # precondition: this note's filename fits, so it exercises the plain
    # (unbudgeted) path and its destination is genuinely a different file
    assert path_overage(corpus.vault_dir / "artifacts", f"{corpus.old_artifact_id}.md") <= 0

    corpus.migrate()

    new_note = artifact_note_path(
        corpus.vault_dir, corpus.new_id, corpus.rename(corpus.old_artifact_id)
    )
    frontmatter, _body = split_note(new_note, new_note.read_text(encoding="utf-8"))
    assert frontmatter["artifact_id"] == corpus.rename(corpus.old_artifact_id)
    assert frontmatter["source_id"] == corpus.new_id
    assert frontmatter["cited_by"] == [corpus.rename(corpus.old_chunk_id)]
    assert not corpus.artifact_note.exists()


def test_note_whose_id_belongs_to_no_known_source_is_refused(tmp_path: Path):
    corpus = _build_corpus(tmp_path)
    stray = corpus.vault_dir / "prose" / "stray.md"
    stray.write_text(
        render_note({"chunk_id": "ghost-000000000000_1_x_001"}, "# x\n"), encoding="utf-8"
    )

    with pytest.raises(MalformedRecordError) as excinfo:
        corpus.plan()

    assert "ghost-000000000000_1_x_001" in str(excinfo.value)


def test_half_applied_corpus_is_refused_rather_than_double_renamed(tmp_path: Path):
    """An interrupted run leaves old- and new-named files side by side. That
    is not a resumable state, so it is refused by name, loudly."""
    corpus = _build_corpus(tmp_path)
    (corpus.data_dir / "trees" / f"{corpus.new_id}.json").write_text("{}", encoding="utf-8")

    with pytest.raises(AmbiguousHashError) as excinfo:
        corpus.plan()

    assert excinfo.value.hash12 == corpus.new_id[-12:]


def test_existing_destination_is_never_overwritten(tmp_path: Path):
    old = f"old-stem-{'a' * 12}"
    new = f"new-{'a' * 12}"
    (tmp_path / f"{old}.json").write_text('{"source_id": "old"}', encoding="utf-8")
    squatter = tmp_path / f"{new}.json"
    squatter.write_text('{"do": "not clobber"}', encoding="utf-8")

    with pytest.raises(DestinationExistsError):
        apply_file(FilePlan(src=tmp_path / f"{old}.json", dest=squatter), {old: new})

    assert json.loads(squatter.read_text(encoding="utf-8")) == {"do": "not clobber"}


# ----------------------------------------------------------------- gold set


def test_budgeted_gold_record_is_renamed_by_the_id_inside_it(tmp_path: Path):
    corpus = _build_corpus(tmp_path)
    chunks_dir = corpus.gold_dir / "chunks"
    # precondition, same shape as the vault note's: this record's filename was
    # shortened at write time, so its id is only recoverable from the record
    assert path_overage(chunks_dir, f"{corpus.old_chunk_id}.json") > 0
    assert corpus.gold_record.name != f"{corpus.old_chunk_id}.json"

    corpus.migrate()

    assert not corpus.gold_record.exists()
    new_record = chunks_dir / f"{corpus.rename(corpus.old_chunk_id)}.json"
    assert new_record.is_file()
    payload = json.loads(new_record.read_text(encoding="utf-8"))
    assert payload["chunk_id"] == corpus.rename(corpus.old_chunk_id)
    assert payload["source"] == corpus.new_id
    assert payload["section"] == "Abstract"


def test_gold_dispatch_and_completed_labelling_are_rewritten(tmp_path: Path):
    corpus = _build_corpus(tmp_path)

    corpus.migrate()

    dispatch = json.loads(
        (corpus.gold_dir / "dispatch" / "blind_full_120.json").read_text(encoding="utf-8")
    )
    assert dispatch[0]["chunk_id"] == corpus.rename(corpus.old_chunk_id)
    assert dispatch[0]["source"] == corpus.new_id

    out = json.loads(
        (corpus.gold_dir / "dispatch" / "out" / "blind_draw_1.json").read_text(encoding="utf-8")
    )
    # the human labels are keyed by chunk_id; the key moves, the labels do not
    assert list(out) == [corpus.rename(corpus.old_chunk_id)]
    assert out[corpus.rename(corpus.old_chunk_id)] == {
        "claim_type": "ideology-as-system",
        "notes": "human judgement",
    }


def test_label_sheet_round_trip_preserves_vocab_and_validations(tmp_path: Path):
    corpus = _build_corpus(tmp_path)
    sheet_path = corpus.gold_dir / "label_sheet.xlsx"
    before = load_workbook(sheet_path)

    corpus.migrate()

    after = load_workbook(sheet_path)
    assert after.sheetnames == before.sheetnames
    assert after[VOCAB_SHEET_NAME].sheet_state == "hidden"
    assert [[cell.value for cell in row] for row in after[VOCAB_SHEET_NAME].iter_rows()] == [
        [cell.value for cell in row] for row in before[VOCAB_SHEET_NAME].iter_rows()
    ]
    assert [
        (dv.type, dv.formula1, str(dv.sqref))
        for dv in after[LABEL_SHEET_NAME].data_validations.dataValidation
    ] == [
        (dv.type, dv.formula1, str(dv.sqref))
        for dv in before[LABEL_SHEET_NAME].data_validations.dataValidation
    ]


def test_label_sheet_rewrites_only_the_id_columns(tmp_path: Path):
    corpus = _build_corpus(tmp_path)

    corpus.migrate()

    sheet = load_workbook(corpus.gold_dir / "label_sheet.xlsx")[LABEL_SHEET_NAME]
    header = [cell.value for cell in sheet[1]]
    row = [cell.value for cell in sheet[2]]
    assert header == ["chunk_id", "source", "section", "claim_type_gold"]
    assert row == [
        corpus.rename(corpus.old_chunk_id),
        corpus.new_id,
        "Abstract",
        "ideology",  # the merged human label is untouched
    ]


def test_gold_record_whose_id_belongs_to_no_known_source_is_refused(tmp_path: Path):
    corpus = _build_corpus(tmp_path)
    stray = corpus.gold_dir / "chunks" / "stray.json"
    stray.write_text(json.dumps({"chunk_id": "ghost-000000000000_1_x_001"}), encoding="utf-8")

    with pytest.raises(MalformedRecordError) as excinfo:
        corpus.plan()

    assert "ghost-000000000000_1_x_001" in str(excinfo.value)


# ------------------------------------------------------- dry run + idempotency


def test_dry_run_writes_nothing_but_reports_the_full_plan(tmp_path: Path):
    corpus = _build_corpus(tmp_path)
    before = {path: path.read_bytes() for path in sorted(tmp_path.rglob("*")) if path.is_file()}

    plan = corpus.plan()
    rendered = format_plan(plan)

    assert {path: path.read_bytes() for path in sorted(tmp_path.rglob("*")) if path.is_file()} == (
        before
    )
    assert corpus.old_id in rendered
    assert corpus.new_id in rendered
    assert str(corpus.data_dir / "chunks") in rendered
    assert str(corpus.vault_dir / "prose") in rendered
    assert str(corpus.gold_dir / "chunks") in rendered
    assert str(corpus.gold_dir / "label_sheet.xlsx") in rendered


def test_second_apply_is_a_no_op(tmp_path: Path):
    corpus = _build_corpus(tmp_path)
    corpus.migrate()
    after_first = {
        path: path.read_bytes() for path in sorted(tmp_path.rglob("*")) if path.is_file()
    }

    second = corpus.plan()
    apply_plan(second)

    assert second.id_map == {corpus.new_id: corpus.new_id}
    assert {
        path: path.read_bytes() for path in sorted(tmp_path.rglob("*")) if path.is_file()
    } == after_first


# ------------------------------------------------------------- substitution


def test_substitution_handles_json_escaped_non_ascii_stems():
    old = "Ugur Umit Ungor - Paramilitarism—notes-abcdef123456"
    new = "ungor-2020-abcdef123456"
    text = json.dumps({"chunk_id": f"{old}_1_intro_001"})

    rewritten = substitute_ids_in_text(text, {old: new})

    assert json.loads(rewritten) == {"chunk_id": f"{new}_1_intro_001"}
