"""Gold-set scoring harness: join the tagger's own sampled chunk records
against the Academic's returned answer key and write a machine-readable
scoring report (PRD §8 P0-10, §10 success metrics & eval; issue #135).

Offline by construction: this pass reads two things already on disk --
`data/gold/chunks/*.json` (the tagger's own output, written by `axial gold
sample`, #53) and `data/gold/labels/label_sheet.xlsx` (the Academic's
returned answer key, handed back off-pipeline per §7.6) -- and joins them on
`chunk_id`. No LLM call, no network, deterministic.

Report contract (locked by tests/test_eval.py, DEC-1):

    {
      "per_axis_agreement": {axis: fraction},
      "tag_counts": {axis: {tag: n}},
      "never_applied": {axis: [tags]},
      "disagreements": [{"chunk_id", "axis", "tagger", "academic"}, ...]
    }

Two report sections are additive, not locked by the outer test but required
by the plan (plans/eval/01-score-gold-set.md:68,74) and stage-2 review of
#135:

    {
      ...,
      "addition_candidates": [{"chunk_id", "axis", "value"}, ...],
      "unmatched": {"sheet_only": [chunk_id, ...], "chunks_only": [chunk_id, ...]}
    }

A third section, `per_polity_score`, scores the free-text, many-valued
`polities_touched` facet as a SET-based metric (locked by
tests/eval/test_eval_polities.py, issue #215) -- it does not fit the
single-value categorical-axis machinery above:

    {
      "per_polity_score": {
        "micro": {"precision": float, "recall": float, "f1": float},
        "macro_f1": float,
        "both_empty_matches": int,
        "per_chunk": [{"chunk_id", "tp", "fp", "fn", "f1"}, ...]
      }
    }

Design choices for `per_polity_score` (#215):
  - Both sides are folded through the #205 canonical alias map
    (`axial.polity_canonical`) before set comparison, so surface variants of
    the same referent (e.g. `USSR` vs `Soviet Union`) count as one polity, a
    true positive rather than a false-positive-plus-false-negative. An
    unmapped verbatim's own casefold+whitespace-normalized form is its key
    (still compares exactly). Missing map file -> graceful degradation:
    every polity falls back to its own casefold+whitespace key, non-fatal.
  - Scope: only chunks with BOTH a row on the returned sheet AND a tagger
    record are scored (the join intersection) -- a tagger chunk with no
    sheet row is excluded entirely (it already surfaces in
    `unmatched.chunks_only`), never treated as a both-empty match.
  - Unlike the categorical axes, an empty Academic `polities_touched` cell
    on a returned row is a REAL answer ("no engaged polity"), not a
    non-answer to exclude -- see gold.py's Academic-facing README text.
  - Both-empty (tagger and Academic both list no polity) is a perfect
    per-chunk match (f1 = 1.0, counted in `both_empty_matches`) but
    contributes nothing to the pooled micro TP/FP/FN.
  - Micro = pooled TP/FP/FN across every scored chunk; macro_f1 = mean of
    each chunk's own F1 (both-empty chunks' 1.0 included).

Design choices (not locked by the outer test, made here):
  - `per_axis_agreement` denominator: only chunks carrying a non-empty
    Academic label on that axis; an empty Academic cell excludes the
    chunk/axis pair entirely (neither an agreement nor a disagreement).
  - `tag_counts`: an "application" is counted from BOTH sides -- once for
    the tagger's own value on that axis (always present) and once for the
    Academic's value when non-empty -- so a tag both sides used on the same
    chunk counts twice. This keeps `tag_counts` and `never_applied`
    semantically aligned: `never_applied` is exactly the schema-vocabulary
    tags with a combined count of zero across BOTH parties, matching the
    spec text ("applied by neither the tagger nor the Academic").
  - `disagreements` are sorted by `chunk_id` then by the axis's position in
    `gold.AXIS_COLUMNS`, for a stable, deterministic row order.
  - `addition_candidates` (plan :74, "tag coverage" addition half): when a
    non-empty Academic value on an axis does not match the tagger's value
    AND is not a member of that axis's schema vocabulary, it is reported
    here instead of in `disagreements` -- the plan's "not a plain mismatch"
    call. The tagger only ever emits controlled-vocabulary values (schema-
    validated upstream in the tag pass), so an out-of-vocab value can only
    originate from the Academic side; there is nothing to gain from also
    double-reporting it as an ordinary mismatch. Same sort order as
    `disagreements`. Per-axis agreement math is unaffected either way -- a
    non-empty Academic cell still counts in that axis's denominator as a
    non-match regardless of which list it lands in.
  - `unmatched` (plan :68, join-miss surfacing): `sheet_only` is every
    chunk_id present on the returned sheet with no matching tagger chunk
    record (e.g. the sheet was returned against an earlier, since-rewritten
    `axial gold sample` run); `chunks_only` is every tagger chunk record
    with no row on the returned sheet at all. Both sorted for determinism.
    A non-empty `unmatched` is not fatal -- exit code stays 0, since a
    partial join still gives a partial, useful report -- but `run_eval`
    also prints a stderr warning naming the counts/a sample of ids so an
    operator running `axial eval` cannot silently miss a stale join
    (mirrors the codebase's anti-silent-failure stderr-warning convention,
    e.g. `gold.log_polity_not_in_list`, `extract._log_fallback`).
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from axial.gold import (
    AXIS_COLUMNS,
    SHEET_COLUMNS,
    MissingChunksError,
    _axis_vocabularies,
    _default_gold_dir,
    _load_gold_records,
)
from axial.llm import DEFAULT_PIPELINE_CONFIG_PATH
from axial.polity_canonical import (
    MissingPolityCanonicalFileError,
    PolityCanonical,
    canonicalize,
    load_polity_canonical,
)
from axial.tag import DEFAULT_DOMAIN_DIR

# Re-exported so callers (the CLI) don't need to import from axial.gold too.
__all__ = [
    "EvalError",
    "MissingChunksError",
    "MissingReturnedSheetError",
    "run_eval",
]


class EvalError(Exception):
    """Base class for all gold-set scoring errors."""


class MissingReturnedSheetError(EvalError):
    """Raised when no returned answer-key sheet exists at
    `data/gold/labels/label_sheet.xlsx` -- the Academic must place their
    filled sheet there (§7.6's handoff) before `axial eval` can score it. No
    report is written when this is raised."""

    def __init__(self, sheet_path: Path):
        self.sheet_path = sheet_path
        super().__init__(
            f"no returned label sheet found at {sheet_path}; place the "
            f"Academic's filled label_sheet.xlsx under data/gold/labels/ "
            f"and re-run `axial eval`"
        )


_CHUNK_ID_COLUMN = SHEET_COLUMNS.index("chunk_id") + 1
_AXIS_COLUMNS_INDEX = {axis: SHEET_COLUMNS.index(axis) + 1 for axis in AXIS_COLUMNS}
_POLITIES_COLUMN = SHEET_COLUMNS.index("polities_touched") + 1


def _normalize_cell(value: Any) -> str | None:
    """A truly-blank Academic cell reads back as `None` (or, defensively, an
    all-whitespace string); both count as "no answer" -- excluded from the
    agreement denominator rather than treated as a value."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _load_academic_labels(sheet_path: Path) -> dict[str, dict[str, str | None]]:
    """Read the returned answer-key sheet's data rows into
    `{chunk_id: {axis: value_or_None}}`."""
    worksheet = load_workbook(sheet_path, read_only=True).worksheets[0]
    labels: dict[str, dict[str, str | None]] = {}
    for row in worksheet.iter_rows(min_row=2):
        chunk_id = _normalize_cell(row[_CHUNK_ID_COLUMN - 1].value)
        if chunk_id is None:
            continue
        labels[chunk_id] = {
            axis: _normalize_cell(row[column - 1].value)
            for axis, column in _AXIS_COLUMNS_INDEX.items()
        }
    return labels


def _parse_polities_cell(value: Any) -> list[str]:
    """Parse a returned sheet's free-text `polities_touched` cell
    (`"; "`-joined, per `gold.build_workbook`) into a list of polity
    strings: split on `;`, strip each, drop empties. A truly-blank cell
    (`None`/whitespace) parses to an empty list -- a REAL "no engaged
    polity" answer for this facet, not a non-answer (see module
    docstring)."""
    text = _normalize_cell(value)
    if text is None:
        return []
    return [part.strip() for part in text.split(";") if part.strip()]


def _load_academic_polities(sheet_path: Path) -> dict[str, list[str]]:
    """Read the returned answer-key sheet's `polities_touched` column into
    `{chunk_id: [polity, ...]}`. A parallel reader to `_load_academic_labels`
    (kept separate so the existing four-axis reader stays byte-for-byte
    intact) -- same row/chunk_id join, different column."""
    worksheet = load_workbook(sheet_path, read_only=True).worksheets[0]
    polities: dict[str, list[str]] = {}
    for row in worksheet.iter_rows(min_row=2):
        chunk_id = _normalize_cell(row[_CHUNK_ID_COLUMN - 1].value)
        if chunk_id is None:
            continue
        polities[chunk_id] = _parse_polities_cell(row[_POLITIES_COLUMN - 1].value)
    return polities


def _load_polity_canonical_or_none(domain_dir: str | Path) -> PolityCanonical | None:
    """Load `<domain_dir>/polity_canonical.yaml` (issue #205), or `None` when
    the file is simply absent (`MissingPolityCanonicalFileError`) --
    graceful degradation, non-fatal: `_polity_fold_key` falls back to
    casefold+whitespace normalization for every polity when no map is
    loaded. Any OTHER `PolityCanonicalError` (malformed YAML, ambiguous
    alias) is a real authoring error and is left to propagate."""
    try:
        return load_polity_canonical(domain_dir)
    except MissingPolityCanonicalFileError:
        print(
            f"warning: no polity canonical map found under {Path(domain_dir)} "
            f"(#205) -- scoring polities_touched without alias folding (surface "
            f"strings compared casefold+whitespace-normalized only)",
            file=sys.stderr,
        )
        return None


_WHITESPACE_PATTERN = re.compile(r"\s+")


def _polity_fold_key(verbatim: str, cmap: PolityCanonical | None) -> str:
    """The comparison key for one polity verbatim: when `cmap` resolves it
    to a mapped node, the node's own canonical name; otherwise (or when no
    map is loaded at all) a casefold+whitespace-normalized form of the
    verbatim itself, so unmapped/unfolded polities still compare exactly."""
    if cmap is not None:
        result = canonicalize(verbatim, cmap)
        if result.status == "mapped" and result.canonical is not None:
            return result.canonical
    return _WHITESPACE_PATTERN.sub(" ", verbatim.strip()).casefold()


def _fold_polity_set(names: list[str], cmap: PolityCanonical | None) -> set[str]:
    """Fold a list of polity verbatims into their comparison-key set (see
    `_polity_fold_key`)."""
    return {_polity_fold_key(name, cmap) for name in names}


def _score_polity_chunk(tagger_keys: set[str], academic_keys: set[str]) -> dict[str, Any]:
    """Set-based precision/recall/F1 for one chunk's folded tagger vs.
    Academic polity key sets. Both empty -> a perfect match (f1 = 1.0,
    "no engaged polity" agreed); a half-empty chunk falls out to f1 = 0.0
    (tp = 0) with no special case needed."""
    tp = len(tagger_keys & academic_keys)
    fp = len(tagger_keys - academic_keys)
    fn = len(academic_keys - tagger_keys)

    if not tagger_keys and not academic_keys:
        f1 = 1.0
    else:
        precision = tp / (tp + fp) if (tp + fp) else 0.0
        recall = tp / (tp + fn) if (tp + fn) else 0.0
        f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) else 0.0

    return {"tp": tp, "fp": fp, "fn": fn, "f1": f1}


def _build_polity_score(
    tagger_records: list[dict[str, Any]],
    academic_polities: dict[str, list[str]],
    cmap: PolityCanonical | None,
) -> dict[str, Any]:
    """Score the many-valued `polities_touched` facet as a set-based metric
    (issue #215). Scored only over chunks with BOTH a tagger record AND a
    row on the returned sheet (the join intersection) -- see module
    docstring for the scope/both-empty/fold design choices."""
    per_chunk: list[dict[str, Any]] = []
    both_empty_matches = 0
    pooled_tp = pooled_fp = pooled_fn = 0
    f1_values: list[float] = []

    for record in tagger_records:
        chunk_id = record.get("chunk_id")
        if chunk_id not in academic_polities:
            continue  # no sheet row -- excluded, not a both-empty match

        tagger_keys = _fold_polity_set(record.get("polities_touched") or [], cmap)
        academic_keys = _fold_polity_set(academic_polities[chunk_id], cmap)

        scored = _score_polity_chunk(tagger_keys, academic_keys)
        per_chunk.append({"chunk_id": chunk_id, **scored})
        f1_values.append(scored["f1"])

        if not tagger_keys and not academic_keys:
            both_empty_matches += 1
        else:
            pooled_tp += scored["tp"]
            pooled_fp += scored["fp"]
            pooled_fn += scored["fn"]

    micro_precision = pooled_tp / (pooled_tp + pooled_fp) if (pooled_tp + pooled_fp) else 0.0
    micro_recall = pooled_tp / (pooled_tp + pooled_fn) if (pooled_tp + pooled_fn) else 0.0
    micro_f1 = (
        (2 * micro_precision * micro_recall / (micro_precision + micro_recall))
        if (micro_precision + micro_recall)
        else 0.0
    )
    macro_f1 = (sum(f1_values) / len(f1_values)) if f1_values else 0.0

    per_chunk.sort(key=lambda row: row["chunk_id"])

    return {
        "micro": {"precision": micro_precision, "recall": micro_recall, "f1": micro_f1},
        "macro_f1": macro_f1,
        "both_empty_matches": both_empty_matches,
        "per_chunk": per_chunk,
    }


def _build_report(
    tagger_records: list[dict[str, Any]],
    academic_by_chunk: dict[str, dict[str, str | None]],
    vocabularies: dict[str, list[str]],
) -> dict[str, Any]:
    """Join the tagger records against the Academic's labels on `chunk_id`
    and compute the locked report shape. See module docstring for the
    denominator/tag-counting/ordering choices."""
    agreement_counts = {axis: 0 for axis in AXIS_COLUMNS}
    agreement_denoms = {axis: 0 for axis in AXIS_COLUMNS}
    tag_counts: dict[str, dict[str, int]] = {axis: {} for axis in AXIS_COLUMNS}
    disagreements: list[dict[str, Any]] = []
    addition_candidates: list[dict[str, Any]] = []

    tagger_chunk_ids = {record.get("chunk_id") for record in tagger_records}

    for record in tagger_records:
        chunk_id = record.get("chunk_id")
        academic = academic_by_chunk.get(chunk_id, {})
        for axis in AXIS_COLUMNS:
            tagger_value = record.get(axis)
            if tagger_value:
                tag_counts[axis][tagger_value] = tag_counts[axis].get(tagger_value, 0) + 1

            academic_value = academic.get(axis)
            if academic_value is None:
                continue  # Academic non-answer -- excluded from this axis's denominator

            tag_counts[axis][academic_value] = tag_counts[axis].get(academic_value, 0) + 1
            agreement_denoms[axis] += 1
            if tagger_value == academic_value:
                agreement_counts[axis] += 1
            elif academic_value not in vocabularies.get(axis, []):
                # Out-of-vocab Academic value: an addition candidate, not a
                # plain mismatch (plan :74) -- see module docstring.
                addition_candidates.append(
                    {"chunk_id": chunk_id, "axis": axis, "value": academic_value}
                )
            else:
                disagreements.append(
                    {
                        "chunk_id": chunk_id,
                        "axis": axis,
                        "tagger": tagger_value,
                        "academic": academic_value,
                    }
                )

    per_axis_agreement = {
        axis: (agreement_counts[axis] / agreement_denoms[axis]) if agreement_denoms[axis] else 0.0
        for axis in AXIS_COLUMNS
    }

    never_applied = {
        axis: [tag for tag in vocabularies.get(axis, []) if tag_counts[axis].get(tag, 0) == 0]
        for axis in AXIS_COLUMNS
    }

    disagreements.sort(key=lambda row: (row["chunk_id"], AXIS_COLUMNS.index(row["axis"])))
    addition_candidates.sort(key=lambda row: (row["chunk_id"], AXIS_COLUMNS.index(row["axis"])))

    sheet_only = sorted(
        chunk_id for chunk_id in academic_by_chunk if chunk_id not in tagger_chunk_ids
    )
    chunks_only = sorted(
        chunk_id for chunk_id in tagger_chunk_ids if chunk_id not in academic_by_chunk
    )

    return {
        "per_axis_agreement": per_axis_agreement,
        "tag_counts": tag_counts,
        "never_applied": never_applied,
        "disagreements": disagreements,
        "addition_candidates": addition_candidates,
        "unmatched": {"sheet_only": sheet_only, "chunks_only": chunks_only},
    }


def _sample_ids(ids: list[str], limit: int = 5) -> str:
    """Format a bounded, human-readable sample of ids for a stderr warning
    (the full list can be long; the count already conveys the scale)."""
    sample = ", ".join(ids[:limit])
    if len(ids) > limit:
        sample += f", ... (+{len(ids) - limit} more)"
    return sample


def _warn_unmatched(unmatched: dict[str, list[str]]) -> None:
    """Print a stderr warning naming a two-directional join miss between the
    returned sheet and the tagger's own chunk records. Non-fatal by design
    -- `run_eval` still writes a (partial) report and returns exit 0 -- but
    silent here would hide a stale-sheet-vs-rewritten-chunks join (`gold
    run_gold_sample` clears+rewrites `data/gold/chunks/` on every run)."""
    sheet_only = unmatched.get("sheet_only") or []
    chunks_only = unmatched.get("chunks_only") or []
    if sheet_only:
        print(
            f"warning: {len(sheet_only)} chunk_id(s) on the returned sheet "
            f"have no matching record under data/gold/chunks/ (the sheet may "
            f"be stale against a since-rewritten sample): {_sample_ids(sheet_only)}",
            file=sys.stderr,
        )
    if chunks_only:
        print(
            f"warning: {len(chunks_only)} sampled chunk record(s) under "
            f"data/gold/chunks/ have no row on the returned sheet: "
            f"{_sample_ids(chunks_only)}",
            file=sys.stderr,
        )


def run_eval(
    gold_dir: Path | None = None,
    domain_dir: str | Path = DEFAULT_DOMAIN_DIR,
    config_path: Path = DEFAULT_PIPELINE_CONFIG_PATH,
) -> Path:
    """Score the Academic's returned answer key
    (`<gold_dir>/labels/label_sheet.xlsx`) against the tagger's own sampled
    chunk records (`<gold_dir>/chunks/*.json`), writing
    `<gold_dir>/labels/eval_report.json`. Returns the written path.

    Raises `MissingChunksError` when no sampled chunk records exist (`axial
    gold sample` must run first) and `MissingReturnedSheetError` when no
    returned sheet has been placed under `<gold_dir>/labels/` yet -- no
    report is written in either case. Offline and deterministic: no LLM
    call, no network."""
    if gold_dir is None:
        gold_dir = _default_gold_dir(config_path)

    chunks_dir = gold_dir / "chunks"
    tagger_records = _load_gold_records(chunks_dir) if chunks_dir.is_dir() else []
    if not tagger_records:
        raise MissingChunksError(chunks_dir)

    labels_dir = gold_dir / "labels"
    sheet_path = labels_dir / "label_sheet.xlsx"
    if not sheet_path.is_file():
        raise MissingReturnedSheetError(sheet_path)

    academic_by_chunk = _load_academic_labels(sheet_path)
    vocabularies = _axis_vocabularies(domain_dir)

    report = _build_report(tagger_records, academic_by_chunk, vocabularies)
    _warn_unmatched(report["unmatched"])

    academic_polities = _load_academic_polities(sheet_path)
    cmap = _load_polity_canonical_or_none(domain_dir)
    report["per_polity_score"] = _build_polity_score(tagger_records, academic_polities, cmap)

    labels_dir.mkdir(parents=True, exist_ok=True)
    report_path = labels_dir / "eval_report.json"
    report_path.write_text(
        json.dumps(report, indent=2, sort_keys=True, ensure_ascii=False), encoding="utf-8"
    )
    return report_path
