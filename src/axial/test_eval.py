"""Inner unit tests for the gold-set scoring harness (src/axial/eval.py,
issue #135).

The outer acceptance test (tests/test_eval.py) drives `axial eval` end to
end through a subprocess and pins the exact report shape/values for one
hand-derived fixture. These unit tests exercise the pieces underneath it in
isolation: cell normalization, the workbook reader, the join/aggregation
logic (agreement denominators, both-sides tag counting, disagreement
ordering) and the two typed errors -- including edge cases the outer
fixture doesn't need to cover (an axis with zero Academic answers, a
whitespace-only "empty" cell, a chunk in the tagger output the sheet never
mentions).
"""

from __future__ import annotations

import json

import pytest
from openpyxl import Workbook

from axial.eval import (
    EvalError,
    MissingChunksError,
    MissingReturnedSheetError,
    _build_report,
    _load_academic_labels,
    _normalize_cell,
    _warn_unmatched,
    run_eval,
)
from axial.gold import AXIS_COLUMNS, SHEET_COLUMNS, build_workbook


def _record(
    chunk_id, field="state", scope="scope:general", claim="state-formation", school="bellicist"
):
    return {
        "chunk_id": chunk_id,
        "source": "src-a",
        "section": "Body",
        "chunk_text": f"text {chunk_id}",
        "field": field,
        "empirical_scope": scope,
        "role_in_argument": "role:claim",
        "claim_type": claim,
        "theory_school": school,
    }


class TestNormalizeCell:
    def test_none_is_none(self):
        assert _normalize_cell(None) is None

    def test_empty_string_is_none(self):
        assert _normalize_cell("") is None

    def test_whitespace_only_is_none(self):
        assert _normalize_cell("   ") is None

    def test_value_is_stripped(self):
        assert _normalize_cell("  state  ") == "state"

    def test_non_string_value_is_stringified(self):
        assert _normalize_cell(42) == "42"


class TestMissingReturnedSheetError:
    def test_is_eval_error(self):
        assert issubclass(MissingReturnedSheetError, EvalError)

    def test_message_names_labels_dir(self, tmp_path):
        from pathlib import Path

        sheet_path = tmp_path / "data" / "gold" / "labels" / "label_sheet.xlsx"
        error = MissingReturnedSheetError(sheet_path)
        assert "data/gold/labels/" in str(error)
        assert error.sheet_path == sheet_path


class TestBuildReport:
    def test_agreement_excludes_empty_academic_cells(self):
        tagger_records = [_record("c1", field="state"), _record("c2", field="violence")]
        academic = {
            "c1": {
                "field": "state",
                "empirical_scope": None,
                "claim_type": None,
                "theory_school": None,
            },
            "c2": {
                "field": None,
                "empirical_scope": None,
                "claim_type": None,
                "theory_school": None,
            },
        }
        vocabularies = {axis: [] for axis in AXIS_COLUMNS}

        report = _build_report(tagger_records, academic, vocabularies)

        # Only c1 answered `field`; it agrees -> 1/1.
        assert report["per_axis_agreement"]["field"] == 1.0
        # No chunk answered the other three axes -> zero denominator -> 0.0,
        # never a divide-by-zero crash.
        assert report["per_axis_agreement"]["empirical_scope"] == 0.0
        assert report["per_axis_agreement"]["claim_type"] == 0.0
        assert report["per_axis_agreement"]["theory_school"] == 0.0
        assert report["disagreements"] == []

    def test_disagreement_rows_and_no_false_positives(self):
        tagger_records = [_record("c1", field="state")]
        academic = {
            "c1": {
                "field": "violence",
                "empirical_scope": "scope:general",
                "claim_type": "state-formation",
                "theory_school": "bellicist",
            },
        }
        # `field`'s mismatched Academic value ("violence") must be in-vocab
        # here so this test exercises a plain disagreement, not the
        # out-of-vocab addition-candidate path (covered separately below).
        vocabularies = {axis: [] for axis in AXIS_COLUMNS}
        vocabularies["field"] = ["state", "violence"]

        report = _build_report(tagger_records, academic, vocabularies)

        assert report["disagreements"] == [
            {"chunk_id": "c1", "axis": "field", "tagger": "state", "academic": "violence"}
        ]
        # The three agreeing axes must not leak into disagreements.
        reported_axes = {row["axis"] for row in report["disagreements"]}
        assert "empirical_scope" not in reported_axes
        assert "claim_type" not in reported_axes
        assert "theory_school" not in reported_axes

    def test_disagreements_sorted_by_chunk_then_axis_order(self):
        tagger_records = [
            _record("c2", field="state", claim="state-formation"),
            _record("c1", field="state", claim="state-formation"),
        ]
        academic = {
            "c2": {
                "field": "violence",
                "empirical_scope": "scope:general",
                "claim_type": "state-capacity",
                "theory_school": "bellicist",
            },
            "c1": {
                "field": "violence",
                "empirical_scope": "scope:general",
                "claim_type": "state-capacity",
                "theory_school": "bellicist",
            },
        }
        # Both mismatched axes' Academic values must be in-vocab so this
        # test exercises plain disagreements, not addition candidates.
        vocabularies = {axis: [] for axis in AXIS_COLUMNS}
        vocabularies["field"] = ["state", "violence"]
        vocabularies["claim_type"] = ["state-formation", "state-capacity"]

        report = _build_report(tagger_records, academic, vocabularies)
        rows = [(row["chunk_id"], row["axis"]) for row in report["disagreements"]]
        # c1 sorts before c2; within a chunk, field precedes claim_type
        # (their order in AXIS_COLUMNS).
        assert rows == [
            ("c1", "field"),
            ("c1", "claim_type"),
            ("c2", "field"),
            ("c2", "claim_type"),
        ]

    def test_tag_counts_combine_both_sides(self):
        # Tagger and Academic both say "state" on c1 -> counted twice;
        # tagger says "violence" on c2, Academic gives no answer -> counted once.
        tagger_records = [_record("c1", field="state"), _record("c2", field="violence")]
        academic = {
            "c1": {
                "field": "state",
                "empirical_scope": None,
                "claim_type": None,
                "theory_school": None,
            },
            "c2": {
                "field": None,
                "empirical_scope": None,
                "claim_type": None,
                "theory_school": None,
            },
        }
        vocabularies = {axis: [] for axis in AXIS_COLUMNS}

        report = _build_report(tagger_records, academic, vocabularies)

        assert report["tag_counts"]["field"] == {"state": 2, "violence": 1}

    def test_never_applied_names_zero_combined_count_tags(self):
        tagger_records = [_record("c1", field="state")]
        academic = {
            "c1": {
                "field": "state",
                "empirical_scope": None,
                "claim_type": None,
                "theory_school": None,
            },
        }
        vocabularies = {
            "field": ["state", "violence", "ideology"],
            "empirical_scope": [],
            "claim_type": [],
            "theory_school": [],
        }

        report = _build_report(tagger_records, academic, vocabularies)

        assert report["never_applied"]["field"] == ["violence", "ideology"]

    def test_chunk_absent_from_sheet_treated_as_all_axes_unanswered(self):
        # A tagger chunk with no matching row on the returned sheet at all
        # (join miss) must not crash -- every axis is simply unanswered.
        tagger_records = [_record("c1")]
        academic = {}
        vocabularies = {axis: [] for axis in AXIS_COLUMNS}

        report = _build_report(tagger_records, academic, vocabularies)

        assert report["disagreements"] == []
        assert all(fraction == 0.0 for fraction in report["per_axis_agreement"].values())
        # The join miss itself must be surfaced, not silently absorbed.
        assert report["unmatched"]["chunks_only"] == ["c1"]
        assert report["unmatched"]["sheet_only"] == []

    def test_unmatched_names_sheet_rows_with_no_chunk_record(self):
        # A returned-sheet chunk_id with no matching tagger record at all
        # (e.g. the sheet answers a since-rewritten sample) -- the other
        # join-miss direction from the test above.
        tagger_records = [_record("c1")]
        academic = {
            "c1": {
                "field": "state",
                "empirical_scope": "scope:general",
                "claim_type": "state-formation",
                "theory_school": "bellicist",
            },
            "stale-c9": {
                "field": "state",
                "empirical_scope": "scope:general",
                "claim_type": "state-formation",
                "theory_school": "bellicist",
            },
        }
        vocabularies = {axis: [] for axis in AXIS_COLUMNS}

        report = _build_report(tagger_records, academic, vocabularies)

        assert report["unmatched"]["sheet_only"] == ["stale-c9"]
        assert report["unmatched"]["chunks_only"] == []

    def test_unmatched_both_directions_sorted(self):
        tagger_records = [_record("only-in-chunks-b"), _record("only-in-chunks-a")]
        academic = {
            "only-in-sheet-b": {
                "field": "state",
                "empirical_scope": None,
                "claim_type": None,
                "theory_school": None,
            },
            "only-in-sheet-a": {
                "field": "state",
                "empirical_scope": None,
                "claim_type": None,
                "theory_school": None,
            },
        }
        vocabularies = {axis: [] for axis in AXIS_COLUMNS}

        report = _build_report(tagger_records, academic, vocabularies)

        assert report["unmatched"]["sheet_only"] == ["only-in-sheet-a", "only-in-sheet-b"]
        assert report["unmatched"]["chunks_only"] == ["only-in-chunks-a", "only-in-chunks-b"]

    def test_out_of_vocab_academic_value_is_an_addition_candidate_not_a_disagreement(self):
        tagger_records = [_record("c1", field="state")]
        academic = {
            "c1": {
                "field": "unrecognized-new-field",  # outside the vocabulary below
                "empirical_scope": None,
                "claim_type": None,
                "theory_school": None,
            },
        }
        vocabularies = {axis: [] for axis in AXIS_COLUMNS}
        vocabularies["field"] = ["state", "violence"]

        report = _build_report(tagger_records, academic, vocabularies)

        assert report["addition_candidates"] == [
            {"chunk_id": "c1", "axis": "field", "value": "unrecognized-new-field"}
        ]
        # Not double-reported as a plain mismatch (plan's "not a plain
        # mismatch" -- see module docstring).
        assert report["disagreements"] == []
        # Agreement math is unaffected: still a non-match in the denominator.
        assert report["per_axis_agreement"]["field"] == 0.0

    def test_in_vocab_mismatch_stays_a_plain_disagreement_not_addition_candidate(self):
        tagger_records = [_record("c1", field="state")]
        academic = {
            "c1": {
                "field": "violence",  # in-vocab mismatch
                "empirical_scope": None,
                "claim_type": None,
                "theory_school": None,
            },
        }
        vocabularies = {axis: [] for axis in AXIS_COLUMNS}
        vocabularies["field"] = ["state", "violence"]

        report = _build_report(tagger_records, academic, vocabularies)

        assert report["disagreements"] == [
            {"chunk_id": "c1", "axis": "field", "tagger": "state", "academic": "violence"}
        ]
        assert report["addition_candidates"] == []


class TestWarnUnmatched:
    def test_no_warning_when_both_lists_empty(self, capsys):
        _warn_unmatched({"sheet_only": [], "chunks_only": []})
        captured = capsys.readouterr()
        assert captured.err == ""

    def test_warns_on_sheet_only(self, capsys):
        _warn_unmatched({"sheet_only": ["stale-c9"], "chunks_only": []})
        captured = capsys.readouterr()
        assert captured.out == "", "the warning must go to stderr, not stdout"
        assert "stale-c9" in captured.err
        assert "1" in captured.err

    def test_warns_on_chunks_only(self, capsys):
        _warn_unmatched({"sheet_only": [], "chunks_only": ["c1", "c2"]})
        captured = capsys.readouterr()
        assert captured.out == ""
        assert "c1" in captured.err
        assert "c2" in captured.err
        assert "2" in captured.err

    def test_warns_on_both_directions_independently(self, capsys):
        _warn_unmatched({"sheet_only": ["stale-c9"], "chunks_only": ["c1"]})
        captured = capsys.readouterr()
        assert "stale-c9" in captured.err
        assert "c1" in captured.err


class TestLoadAcademicLabels:
    def test_reads_values_and_blank_cells(self, tmp_path):
        vocabularies = {axis: ["state", "violence"] for axis in AXIS_COLUMNS}
        records = [_record("c1"), _record("c2")]
        workbook = build_workbook(records, vocabularies)
        worksheet = workbook.worksheets[0]
        field_col = SHEET_COLUMNS.index("field") + 1
        claim_col = SHEET_COLUMNS.index("claim_type") + 1
        # Row 2 (c1): fill both; row 3 (c2): leave field truly blank.
        worksheet.cell(row=2, column=field_col).value = "state"
        worksheet.cell(row=2, column=claim_col).value = "state-formation"
        worksheet.cell(row=3, column=field_col).value = None
        worksheet.cell(row=3, column=claim_col).value = "state-formation"

        sheet_path = tmp_path / "label_sheet.xlsx"
        workbook.save(sheet_path)

        labels = _load_academic_labels(sheet_path)

        assert labels["c1"]["field"] == "state"
        assert labels["c1"]["claim_type"] == "state-formation"
        assert labels["c2"]["field"] is None
        assert labels["c2"]["claim_type"] == "state-formation"


class TestRunEval:
    def test_missing_chunks_raises(self, tmp_path):
        gold_dir = tmp_path / "data" / "gold"
        with pytest.raises(MissingChunksError):
            run_eval(gold_dir=gold_dir)

    def test_missing_returned_sheet_raises_and_writes_no_report(self, tmp_path):
        gold_dir = tmp_path / "data" / "gold"
        chunks_dir = gold_dir / "chunks"
        chunks_dir.mkdir(parents=True)
        (chunks_dir / "c1.json").write_text(
            json.dumps(_record("c1"), indent=2, sort_keys=True), encoding="utf-8"
        )

        with pytest.raises(MissingReturnedSheetError) as excinfo:
            run_eval(gold_dir=gold_dir)

        assert "data/gold/labels/" in str(excinfo.value)
        assert not (gold_dir / "labels" / "eval_report.json").is_file()

    def test_end_to_end_writes_deterministic_report(self, tmp_path):
        from axial.gold import _axis_vocabularies
        from axial.tag import DEFAULT_DOMAIN_DIR

        gold_dir = tmp_path / "data" / "gold"
        chunks_dir = gold_dir / "chunks"
        chunks_dir.mkdir(parents=True)
        records = [_record("c1", field="state"), _record("c2", field="violence")]
        for record in records:
            (chunks_dir / f"{record['chunk_id']}.json").write_text(
                json.dumps(record, indent=2, sort_keys=True), encoding="utf-8"
            )

        vocabularies = _axis_vocabularies(DEFAULT_DOMAIN_DIR)
        workbook = build_workbook(records, vocabularies)
        worksheet = workbook.worksheets[0]
        field_col = SHEET_COLUMNS.index("field") + 1
        # Row 2 (c1) agrees; row 3 (c2) disagrees.
        worksheet.cell(row=2, column=field_col).value = "state"
        worksheet.cell(row=3, column=field_col).value = "ideology"

        labels_dir = gold_dir / "labels"
        labels_dir.mkdir(parents=True)
        workbook.save(labels_dir / "label_sheet.xlsx")

        report_path = run_eval(gold_dir=gold_dir, domain_dir=DEFAULT_DOMAIN_DIR)

        assert report_path == labels_dir / "eval_report.json"
        first = json.loads(report_path.read_text(encoding="utf-8"))

        second_path = run_eval(gold_dir=gold_dir, domain_dir=DEFAULT_DOMAIN_DIR)
        second = json.loads(second_path.read_text(encoding="utf-8"))

        assert first == second
        assert first["per_axis_agreement"]["field"] == 0.5
