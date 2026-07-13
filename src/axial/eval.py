"""Gold-set scoring harness: join the tagger's own sampled chunk records
against the Academic's returned answer key and write a machine-readable
scoring report (PRD §8 P0-10, §10 success metrics & eval; issue #135).

Offline by construction: this pass reads two things already on disk --
`data/gold/chunks/*.json` (the tagger's own output, written by `axial gold
sample`, #53) and `data/gold/labels/label_sheet.xlsx` (the Academic's
returned answer key, handed back off-pipeline per §7.6) -- and joins them on
`chunk_id`. No LLM call, no network, deterministic.

Report contract (locked by tests/test_eval.py, DEC-1):

    {
      "per_axis_agreement": {axis: fraction},
      "tag_counts": {axis: {tag: n}},
      "never_applied": {axis: [tags]},
      "disagreements": [{"chunk_id", "axis", "tagger", "academic"}, ...]
    }

Two report sections are additive, not locked by the outer test but required
by the plan (plans/eval/01-score-gold-set.md:68,74) and stage-2 review of
#135:

    {
      ...,
      "addition_candidates": [{"chunk_id", "axis", "value"}, ...],
      "unmatched": {"sheet_only": [chunk_id, ...], "chunks_only": [chunk_id, ...]}
    }

Design choices (not locked by the outer test, made here):
  - `per_axis_agreement` denominator: only chunks carrying a non-empty
    Academic label on that axis; an empty Academic cell excludes the
    chunk/axis pair entirely (neither an agreement nor a disagreement).
  - `tag_counts`: an "application" is counted from BOTH sides -- once for
    the tagger's own value on that axis (always present) and once for the
    Academic's value when non-empty -- so a tag both sides used on the same
    chunk counts twice. This keeps `tag_counts` and `never_applied`
    semantically aligned: `never_applied` is exactly the schema-vocabulary
    tags with a combined count of zero across BOTH parties, matching the
    spec text ("applied by neither the tagger nor the Academic").
  - `disagreements` are sorted by `chunk_id` then by the axis's position in
    `gold.AXIS_COLUMNS`, for a stable, deterministic row order.
  - `addition_candidates` (plan :74, "tag coverage" addition half): when a
    non-empty Academic value on an axis does not match the tagger's value
    AND is not a member of that axis's schema vocabulary, it is reported
    here instead of in `disagreements` -- the plan's "not a plain mismatch"
    call. The tagger only ever emits controlled-vocabulary values (schema-
    validated upstream in the tag pass), so an out-of-vocab value can only
    originate from the Academic side; there is nothing to gain from also
    double-reporting it as an ordinary mismatch. Same sort order as
    `disagreements`. Per-axis agreement math is unaffected either way -- a
    non-empty Academic cell still counts in that axis's denominator as a
    non-match regardless of which list it lands in.
  - `unmatched` (plan :68, join-miss surfacing): `sheet_only` is every
    chunk_id present on the returned sheet with no matching tagger chunk
    record (e.g. the sheet was returned against an earlier, since-rewritten
    `axial gold sample` run); `chunks_only` is every tagger chunk record
    with no row on the returned sheet at all. Both sorted for determinism.
    A non-empty `unmatched` is not fatal -- exit code stays 0, since a
    partial join still gives a partial, useful report -- but `run_eval`
    also prints a stderr warning naming the counts/a sample of ids so an
    operator running `axial eval` cannot silently miss a stale join
    (mirrors the codebase's anti-silent-failure stderr-warning convention,
    e.g. `gold.log_country_not_in_list`, `extract._log_fallback`).
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from axial.gold import (
    AXIS_COLUMNS,
    SHEET_COLUMNS,
    MissingChunksError,
    _axis_vocabularies,
    _default_gold_dir,
    _load_gold_records,
)
from axial.llm import DEFAULT_PIPELINE_CONFIG_PATH
from axial.tag import DEFAULT_DOMAIN_DIR

# Re-exported so callers (the CLI) don't need to import from axial.gold too.
__all__ = [
    "EvalError",
    "MissingChunksError",
    "MissingReturnedSheetError",
    "run_eval",
]


class EvalError(Exception):
    """Base class for all gold-set scoring errors."""


class MissingReturnedSheetError(EvalError):
    """Raised when no returned answer-key sheet exists at
    `data/gold/labels/label_sheet.xlsx` -- the Academic must place their
    filled sheet there (§7.6's handoff) before `axial eval` can score it. No
    report is written when this is raised."""

    def __init__(self, sheet_path: Path):
        self.sheet_path = sheet_path
        super().__init__(
            f"no returned label sheet found at {sheet_path}; place the "
            f"Academic's filled label_sheet.xlsx under data/gold/labels/ "
            f"and re-run `axial eval`"
        )


_CHUNK_ID_COLUMN = SHEET_COLUMNS.index("chunk_id") + 1
_AXIS_COLUMNS_INDEX = {axis: SHEET_COLUMNS.index(axis) + 1 for axis in AXIS_COLUMNS}


def _normalize_cell(value: Any) -> str | None:
    """A truly-blank Academic cell reads back as `None` (or, defensively, an
    all-whitespace string); both count as "no answer" -- excluded from the
    agreement denominator rather than treated as a value."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _load_academic_labels(sheet_path: Path) -> dict[str, dict[str, str | None]]:
    """Read the returned answer-key sheet's data rows into
    `{chunk_id: {axis: value_or_None}}`."""
    worksheet = load_workbook(sheet_path, read_only=True).worksheets[0]
    labels: dict[str, dict[str, str | None]] = {}
    for row in worksheet.iter_rows(min_row=2):
        chunk_id = _normalize_cell(row[_CHUNK_ID_COLUMN - 1].value)
        if chunk_id is None:
            continue
        labels[chunk_id] = {
            axis: _normalize_cell(row[column - 1].value)
            for axis, column in _AXIS_COLUMNS_INDEX.items()
        }
    return labels


def _build_report(
    tagger_records: list[dict[str, Any]],
    academic_by_chunk: dict[str, dict[str, str | None]],
    vocabularies: dict[str, list[str]],
) -> dict[str, Any]:
    """Join the tagger records against the Academic's labels on `chunk_id`
    and compute the locked report shape. See module docstring for the
    denominator/tag-counting/ordering choices."""
    agreement_counts = {axis: 0 for axis in AXIS_COLUMNS}
    agreement_denoms = {axis: 0 for axis in AXIS_COLUMNS}
    tag_counts: dict[str, dict[str, int]] = {axis: {} for axis in AXIS_COLUMNS}
    disagreements: list[dict[str, Any]] = []
    addition_candidates: list[dict[str, Any]] = []

    tagger_chunk_ids = {record.get("chunk_id") for record in tagger_records}

    for record in tagger_records:
        chunk_id = record.get("chunk_id")
        academic = academic_by_chunk.get(chunk_id, {})
        for axis in AXIS_COLUMNS:
            tagger_value = record.get(axis)
            if tagger_value:
                tag_counts[axis][tagger_value] = tag_counts[axis].get(tagger_value, 0) + 1

            academic_value = academic.get(axis)
            if academic_value is None:
                continue  # Academic non-answer -- excluded from this axis's denominator

            tag_counts[axis][academic_value] = tag_counts[axis].get(academic_value, 0) + 1
            agreement_denoms[axis] += 1
            if tagger_value == academic_value:
                agreement_counts[axis] += 1
            elif academic_value not in vocabularies.get(axis, []):
                # Out-of-vocab Academic value: an addition candidate, not a
                # plain mismatch (plan :74) -- see module docstring.
                addition_candidates.append(
                    {"chunk_id": chunk_id, "axis": axis, "value": academic_value}
                )
            else:
                disagreements.append(
                    {
                        "chunk_id": chunk_id,
                        "axis": axis,
                        "tagger": tagger_value,
                        "academic": academic_value,
                    }
                )

    per_axis_agreement = {
        axis: (agreement_counts[axis] / agreement_denoms[axis]) if agreement_denoms[axis] else 0.0
        for axis in AXIS_COLUMNS
    }

    never_applied = {
        axis: [tag for tag in vocabularies.get(axis, []) if tag_counts[axis].get(tag, 0) == 0]
        for axis in AXIS_COLUMNS
    }

    disagreements.sort(key=lambda row: (row["chunk_id"], AXIS_COLUMNS.index(row["axis"])))
    addition_candidates.sort(key=lambda row: (row["chunk_id"], AXIS_COLUMNS.index(row["axis"])))

    sheet_only = sorted(
        chunk_id for chunk_id in academic_by_chunk if chunk_id not in tagger_chunk_ids
    )
    chunks_only = sorted(
        chunk_id for chunk_id in tagger_chunk_ids if chunk_id not in academic_by_chunk
    )

    return {
        "per_axis_agreement": per_axis_agreement,
        "tag_counts": tag_counts,
        "never_applied": never_applied,
        "disagreements": disagreements,
        "addition_candidates": addition_candidates,
        "unmatched": {"sheet_only": sheet_only, "chunks_only": chunks_only},
    }


def _sample_ids(ids: list[str], limit: int = 5) -> str:
    """Format a bounded, human-readable sample of ids for a stderr warning
    (the full list can be long; the count already conveys the scale)."""
    sample = ", ".join(ids[:limit])
    if len(ids) > limit:
        sample += f", ... (+{len(ids) - limit} more)"
    return sample


def _warn_unmatched(unmatched: dict[str, list[str]]) -> None:
    """Print a stderr warning naming a two-directional join miss between the
    returned sheet and the tagger's own chunk records. Non-fatal by design
    -- `run_eval` still writes a (partial) report and returns exit 0 -- but
    silent here would hide a stale-sheet-vs-rewritten-chunks join (`gold
    run_gold_sample` clears+rewrites `data/gold/chunks/` on every run)."""
    sheet_only = unmatched.get("sheet_only") or []
    chunks_only = unmatched.get("chunks_only") or []
    if sheet_only:
        print(
            f"warning: {len(sheet_only)} chunk_id(s) on the returned sheet "
            f"have no matching record under data/gold/chunks/ (the sheet may "
            f"be stale against a since-rewritten sample): {_sample_ids(sheet_only)}",
            file=sys.stderr,
        )
    if chunks_only:
        print(
            f"warning: {len(chunks_only)} sampled chunk record(s) under "
            f"data/gold/chunks/ have no row on the returned sheet: "
            f"{_sample_ids(chunks_only)}",
            file=sys.stderr,
        )


def run_eval(
    gold_dir: Path | None = None,
    domain_dir: str | Path = DEFAULT_DOMAIN_DIR,
    config_path: Path = DEFAULT_PIPELINE_CONFIG_PATH,
) -> Path:
    """Score the Academic's returned answer key
    (`<gold_dir>/labels/label_sheet.xlsx`) against the tagger's own sampled
    chunk records (`<gold_dir>/chunks/*.json`), writing
    `<gold_dir>/labels/eval_report.json`. Returns the written path.

    Raises `MissingChunksError` when no sampled chunk records exist (`axial
    gold sample` must run first) and `MissingReturnedSheetError` when no
    returned sheet has been placed under `<gold_dir>/labels/` yet -- no
    report is written in either case. Offline and deterministic: no LLM
    call, no network."""
    if gold_dir is None:
        gold_dir = _default_gold_dir(config_path)

    chunks_dir = gold_dir / "chunks"
    tagger_records = _load_gold_records(chunks_dir) if chunks_dir.is_dir() else []
    if not tagger_records:
        raise MissingChunksError(chunks_dir)

    labels_dir = gold_dir / "labels"
    sheet_path = labels_dir / "label_sheet.xlsx"
    if not sheet_path.is_file():
        raise MissingReturnedSheetError(sheet_path)

    academic_by_chunk = _load_academic_labels(sheet_path)
    vocabularies = _axis_vocabularies(domain_dir)

    report = _build_report(tagger_records, academic_by_chunk, vocabularies)
    _warn_unmatched(report["unmatched"])

    labels_dir.mkdir(parents=True, exist_ok=True)
    report_path = labels_dir / "eval_report.json"
    report_path.write_text(
        json.dumps(report, indent=2, sort_keys=True, ensure_ascii=False), encoding="utf-8"
    )
    return report_path
