"""Gold-set sampling: select a stratified set of tagged prose chunks from
the vault and write one chunk record per selection under `data/gold/chunks/`
(PRD §9 gold corpus & labeling, §8 P0-9; plans/gold/01-gold-sample.md).

Offline by construction: this pass reads the tagged prose notes the
`tag`/`vault` passes already wrote (`data/vault/prose/*.md`) -- no LLM call,
no network. It never recomputes a tag; it copies each note's existing axis
tags into a flat chunk record the label sheet (slice 02) then renders.

Stratification (founder-ratified retarget, spec PR #124): the balancing
strata are **field × empirical_scope × role_in_argument**. The selection
covers every represented value of each of those three axes. source-type
(book/paper), claim_type and theory_school are descriptive, not balancing --
they ride along on whatever is drawn; each source-type declared in
`data/gold/sources.yaml` that is present in the corpus contributes at least
one chunk. Non-substantive back-matter (bibliography, index, references,
endnotes, appendix, front-matter) is excluded from the sampling frame
entirely (`_is_back_matter`), reusing and broadening the chunk-pass filter's
own vocabulary (issue #113).

Selection is deterministic and seedable: notes are read in a stable order
and, given a fixed seed, a re-run reproduces the exact same set. The sample
size sits in a configurable band (default 100-120) clamped to the number of
available chunks. The output directory is cleared before each run, so a
re-run never accumulates stale records.
"""

from __future__ import annotations

import datetime
import json
import random
import re
import shutil
import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from axial.chunk import _BACK_MATTER_TITLES
from axial.codebook import CodebookError, load_codebook
from axial.llm import DEFAULT_PIPELINE_CONFIG_PATH
from axial.tag import DEFAULT_DOMAIN_DIR
from axial.vault import _default_vault_dir

GOLD_DIR = Path("data/gold")

# Appendix I's label-sheet columns, in order. `role_in_argument` is a
# balancing stratum but NOT a sheet column (Appendix I names none); `notes`
# ships empty. `polities_touched` is a pre-labeled, correctable, scored
# free-text facet (same mode as field/empirical_scope: pre-filled with the
# tagger's guess, corrected by typing where wrong) -- not a codebook axis, so
# it carries no dropdown.
SHEET_COLUMNS = (
    "chunk_id",
    "source",
    "section",
    "chunk_text",
    "field",
    "empirical_scope",
    "polities_touched",
    "claim_type",
    "theory_school",
    "notes",
)

# Columns pre-filled from each chunk's own tags (the tagger's guess).
PRELABELED_COLUMNS = ("field", "empirical_scope")

# Columns the Academic labels from scratch (arrive empty, §9 hybrid labeling).
BLIND_COLUMNS = ("claim_type", "theory_school")

# Where the Academic returns the filled sheet (spec directory contract,
# PRODUCT.md §7.5 / repo layout).
LABELS_RETURN_DIR = "data/gold/labels/"

# The four axis columns carrying dropdown validation, in a fixed order; the
# blind pair (claim_type, theory_school) get dropdowns but arrive empty.
AXIS_COLUMNS = ("field", "empirical_scope", "claim_type", "theory_school")

# The pre-labeled, correctable free-text column(s) -- same "pre-filled,
# correct by typing" mode as PRELABELED_COLUMNS, but carrying no dropdown
# (not a codebook axis). Currently just `polities_touched`.
PRELABELED_FREETEXT_COLUMNS = ("polities_touched",)

LABEL_SHEET_NAME = "label_sheet"
VOCAB_SHEET_NAME = "vocab"

DEFAULT_MIN_SIZE = 100
DEFAULT_MAX_SIZE = 120
DEFAULT_SEED = 0

# The three ratified balancing axes (spec PR #124), in a fixed order so the
# coverage pass is deterministic.
STRATA_AXES = ("field", "empirical_scope", "role_in_argument")

# Every field a written gold record carries (Appendix I's substantive columns
# plus role_in_argument, which is a balancing stratum but not a sheet column).
RECORD_FIELDS = (
    "chunk_id",
    "source",
    "section",
    "chunk_text",
    "field",
    "empirical_scope",
    "polities_touched",
    "role_in_argument",
    "claim_type",
    "theory_school",
)

# Non-substantive front/back-matter titles excluded from the gold sampling
# frame, on top of the chunk-pass back-matter vocabulary (`_BACK_MATTER_TITLES`,
# issue #113). #53 excludes a BROADER set than the chunk filter: the chunk
# filter deliberately keeps endnotes/appendix/preface (a false drop there
# loses real content before chunking), but the gold FRAME excludes them --
# they are not substantive argument to be labeled by the Academic.
_GOLD_EXTRA_BACK_MATTER = frozenset(
    {
        "notes",
        "endnotes",
        "end notes",
        "footnotes",
        "notes and references",
        "preface",
        "foreword",
        "acknowledgements",
        "acknowledgments",
        "dedication",
        "epigraph",
        "about the author",
        "about the authors",
        "notes on contributors",
        "glossary",
        "abbreviations",
        "chronology",
        "front matter",
        "back matter",
        "title page",
        "half title",
        "contributors",
    }
)

# Vocabulary-specific suffix words (#204): a title where a qualifier PRECEDES
# a bare back-matter vocab term ("Selected Bibliography", "General Secondary
# Sources") is still back-matter even though it fails the exact/prefix rules
# above. Deliberately narrow -- these two terms are specific enough that an
# `endswith` match never catches an ordinary chapter title ("1984 Reforms"
# does not end in either), and matching on the suffix alone (rather than
# requiring the page/roman-numeral prefix to be stripped first) already
# covers the page-number-prefixed case ("3 General Secondary Sources") since
# a leading digit token cannot affect a suffix comparison.
_BACK_MATTER_SUFFIX_WORDS = (
    "bibliography",
    "secondary sources",
)


class GoldError(Exception):
    """Base class for all gold-set sampling errors."""


class MissingChunksError(GoldError):
    """Raised when no sampled chunk records exist to render into a sheet --
    `axial gold sample` must run first."""

    def __init__(self, chunks_dir: Path):
        self.chunks_dir = chunks_dir
        super().__init__(
            f"no sampled chunk records found under {chunks_dir}; run `axial gold sample` first"
        )


class MissingSheetError(GoldError):
    """Raised when `run_gold_deliver` finds no generated label sheet to
    package -- `axial gold sheet` must run first."""

    def __init__(self, sheet_path: Path):
        self.sheet_path = sheet_path
        super().__init__(f"no label sheet to deliver at {sheet_path}; run `axial gold sheet` first")


class CodebookLoadError(GoldError):
    """Raised when the domain codebook (the dropdown vocabulary source) cannot
    be loaded, so the CLI renders a clean `error: ...` instead of a traceback."""

    def __init__(self, cause: CodebookError):
        self.cause = cause
        super().__init__(str(cause))


class EmptyFrameError(GoldError):
    """Raised when no substantive prose notes are available to sample -- an
    empty vault, or one whose every note is non-substantive back-matter. The
    caller must populate `data/vault/prose/` (run the `tag`/`vault` passes on
    the sources) before sampling."""

    def __init__(self, prose_dir: Path):
        self.prose_dir = prose_dir
        super().__init__(
            f"no substantive prose notes to sample under {prose_dir}; run the "
            f"tag/vault passes on the sources first"
        )


def _default_gold_dir(config_path: Path = DEFAULT_PIPELINE_CONFIG_PATH) -> Path:
    """Resolve the gold directory, mirroring `vault._default_vault_dir`:
    honor `config/pipeline.yaml`'s `paths.gold_dir` when declared, else fall
    back to the cwd-relative `GOLD_DIR` (`data/gold`)."""
    if not config_path.is_file():
        return GOLD_DIR
    with config_path.open("r", encoding="utf-8") as handle:
        document = yaml.safe_load(handle) or {}
    paths_config = document.get("paths", {}) or {}
    configured = paths_config.get("gold_dir")
    return Path(configured) if configured else GOLD_DIR


def _normalize_title(title: str) -> str:
    """Lowercase, collapse whitespace, strip surrounding punctuation -- the
    same normalization the chunk-pass back-matter filter uses (issue #113)."""
    return re.sub(r"\s+", " ", (title or "").lower()).strip(" .:-–—")


_ROMAN_NUMERAL_PREFIX = re.compile(r"^[ivxlcdm]+\.?\s+")

# A leading, purely-numeric page-number token ("154 Notes", "12 Bibliography")
# -- issue #134 gap 1. Stripped before the vocabulary check so page-stamped
# back-matter titles match the same way their bare form does; deliberately
# narrow (digits only, no letters) so an ordinary title that merely starts
# with a year or number ("1984 Reforms") is untouched once its remainder
# fails the vocabulary check.
_PAGE_NUMBER_PREFIX = re.compile(r"^\d+\s+")

# References-family words: when a title's leading roman-numeral ordinal is
# stripped (e.g. "V. Articles and Periodicals" -> "articles and periodicals"),
# the remainder is back-matter only if it names a references/bibliography
# subsection. Deliberately narrow -- an ordinary chapter title that happens to
# start with a roman numeral ("I. The Origins of the State") must NOT match.
_ROMAN_PREFIXED_BACK_MATTER_WORDS = (
    "bibliography",
    "references",
    "articles and periodicals",
    "books",
    "books and articles",
    "primary sources",
    "secondary sources",
    "unpublished sources",
    "archival sources",
    "newspapers and periodicals",
)


def _is_back_matter(section: str) -> bool:
    """True if `section` is non-substantive front/back-matter that must be
    excluded from the gold sampling frame (#53, #131). Reuses the chunk-pass
    vocabulary (`_BACK_MATTER_TITLES`) and broadens it with the gold-frame
    extras plus:
      - an `appendix`/`annex` prefix match (`Appendix A`, `Annex I`);
      - a `notes to page(s) ...` / `note to page ...` prefix match (endnote
        sections titled with a page range, issue #131 false-negative 1);
      - a roman-numeral-ordinal-prefixed references/bibliography subsection
        (`V. Articles and Periodicals`, issue #131 false-negative 2) -- the
        ordinal is stripped and the remainder checked against a narrow
        references-family vocabulary, so an ordinary chapter title that
        merely starts with a roman numeral is never dropped;
      - a leading page-number token before an otherwise-recognized
        back-matter title (`154 Notes`, issue #134 gap 1) -- the numeric
        token is stripped and the remainder re-run through this same
        function, so a title that merely starts with a number but isn't
        otherwise back-matter (`1984 Reforms`) is never dropped;
      - a qualifier preceding a bare references/bibliography vocab term
        (`Selected Bibliography`, `3 General Secondary Sources`, issue #204)
        -- the normalized title's SUFFIX is checked against a narrow
        references-family word list, so ordinary chapter titles are
        untouched while page-prefixed or qualified bibliography/secondary-
        sources sections are still caught."""
    normalized = _normalize_title(section)
    if normalized in _BACK_MATTER_TITLES or normalized in _GOLD_EXTRA_BACK_MATTER:
        return True
    if normalized.startswith("appendix") or normalized.startswith("annex"):
        return True
    if normalized.startswith("notes to page") or normalized.startswith("note to page"):
        return True
    stripped = _ROMAN_NUMERAL_PREFIX.sub("", normalized, count=1)
    if stripped != normalized and stripped in _ROMAN_PREFIXED_BACK_MATTER_WORDS:
        return True
    if any(normalized.endswith(word) for word in _BACK_MATTER_SUFFIX_WORDS):
        return True
    page_stripped = _PAGE_NUMBER_PREFIX.sub("", normalized, count=1)
    if page_stripped != normalized and _is_back_matter(page_stripped):
        return True
    return False


# Minimum-substance guard (#131): a candidate prose chunk's body must clear a
# length floor, an alphabetic-density floor AND a word-shape floor to be
# labelable argument. Thresholds are deliberately generous (a false EXCLUDE
# loses real content from the frame, which is worse than a rare false INCLUDE
# the Academic can just skip). Measured on the gold-test fixtures (real
# 250-480 char argument prose vs. synthetic >=60-char junk, invented text
# only -- copyright):
#   - length < 60 chars: a full citation ("Roe, Invented Book, pp. 10-14.")
#     or a short endnote reads far under this; the substantive fixtures used
#     across the gold tests run 250+ chars. This alone is NOT sufficient: a
#     longer multi-citation footnote string ("Roe, J., Invented Study, pp.
#     12-14; Doe, A., Fictional Work, pp. 88-91.", 71 chars) clears it and
#     must be caught by the checks below (this was the reviewer finding:
#     the original alpha-ratio floor alone was effectively inert against
#     this shape, which measures ~0.63 alpha ratio -- comfortably above the
#     0.6 floor).
#   - alphabetic ratio < 0.6 among non-space characters: still a useful
#     floor for punctuation/digit-dominated fragments, but on its own it
#     does NOT reliably separate realistic citation junk (~0.63-0.74) or
#     OCR/case-substitution garble ("XZQ7t KKtempc;, ZZ QWRTo QQ FALSEo
#     ...", ~0.94 -- almost entirely letters) from real argument prose
#     (~0.90-0.99), which is why a second, independent signal is required.
#   - clean-word-token fraction < 0.7: the fraction of whitespace-split
#     tokens that, once outer punctuation is stripped, are a plausible
#     English word or hyphenated/possessive compound of length >= 3
#     ("state-backed", "Tilly's") with internally case-consistent parts
#     (all-lower, all-upper, or Title case). The length-3 floor matters: it
#     keeps short citation shorthand ("pp.", single-letter initials "J.",
#     "A.") from inflating the score. Real argument prose measures
#     ~0.85-0.93 here (citation-heavy academic prose with parentheticals
#     still clears ~0.85); the seeded chunk-id-bearing placeholder text used
#     by the other gold-sample tests measures ~0.80-0.82 (kept, comfortable
#     margin above the 0.7 floor). Multi-citation junk measures ~0.50-0.64;
#     OCR/case-noise tokens like "GHAROo" or "MlNlSTRYo0F" fail outright,
#     measuring ~0.08 on synthetic garble fixtures.
_MIN_SUBSTANCE_CHARS = 60
_MIN_ALPHA_RATIO = 0.6
_MIN_CLEAN_WORD_FRACTION = 0.7

# A token shorter than this (after stripping edge punctuation and internal
# hyphens/apostrophes) never counts as a clean word -- it excludes citation
# shorthand ("pp.", "n.") and bare initials ("J.", "A.") from inflating the
# clean-word fraction on citation-shaped junk.
_MIN_CLEAN_WORD_LEN = 3

# Punctuation stripped from a token's edges before it is judged word-shaped
# (mirrors the punctuation a citation/footnote commonly wraps a word in).
_TOKEN_EDGE_PUNCT = '.,;:()[]{}"“”‘’—–'


def _is_clean_word_token(token: str) -> bool:
    """True if `token` looks like a plausible English word or a
    hyphenated/possessive compound of them, once its outer punctuation is
    stripped: at least `_MIN_CLEAN_WORD_LEN` letters, and each
    hyphen/apostrophe-delimited part pure letters and internally
    case-consistent (all-lower, all-upper, or Title case) -- this is what
    catches OCR case-substitution noise ("GHAROo", "PRIESTSc") that a pure
    alphabetic-ratio check misses, since such tokens are still almost
    entirely letters. The length floor is what keeps citation shorthand
    ("pp.", "J.", "A.") from counting as a clean word."""
    stripped = token.strip(_TOKEN_EDGE_PUNCT)
    if not stripped:
        return False
    parts = re.split(r"[-']", stripped)
    if any(not part for part in parts):
        return False
    if not all(
        part.isalpha() and (part.islower() or part.isupper() or part.istitle()) for part in parts
    ):
        return False
    return len(stripped.replace("-", "").replace("'", "")) >= _MIN_CLEAN_WORD_LEN


def _clean_word_fraction(text: str) -> float:
    """Fraction of `text`'s whitespace-split tokens that are clean word
    tokens (see `_is_clean_word_token`); 0.0 for an empty/whitespace-only
    text."""
    tokens = text.split()
    if not tokens:
        return 0.0
    clean = sum(1 for token in tokens if _is_clean_word_token(token))
    return clean / len(tokens)


def _is_substantive(chunk_text: str | None) -> bool:
    """True if `chunk_text` clears the length floor, the alphabetic-density
    floor AND the word-shape floor -- i.e. is long enough and prose-like
    enough (not a citation/footnote fragment, not OCR/case-substitution
    garble) to be a labelable argument chunk. See the minimum-substance
    guard comment above for the measured thresholds."""
    if not chunk_text:
        return False
    text = chunk_text.strip()
    if len(text) < _MIN_SUBSTANCE_CHARS:
        return False
    non_space = [ch for ch in text if not ch.isspace()]
    if not non_space:
        return False
    alpha_ratio = sum(1 for ch in non_space if ch.isalpha()) / len(non_space)
    if alpha_ratio < _MIN_ALPHA_RATIO:
        return False
    return _clean_word_fraction(text) >= _MIN_CLEAN_WORD_FRACTION


def source_id_of(chunk_id: str) -> str:
    """Recover a chunk's `source_id` from its `chunk_id` by stripping the
    trailing `_<order>_<slug>_<NNN>` the chunk pass appends
    (`<source_id>_<order>_<slug>_<NNN>`, src/axial/chunk.py). source_id itself
    carries no underscores (`<stem>-<sha256[:12]>`), so the three trailing
    underscore-delimited segments are always the order, slug and index."""
    parts = chunk_id.split("_")
    if len(parts) <= 3:
        return chunk_id
    return "_".join(parts[:-3])


def _split_frontmatter(text: str) -> dict[str, Any] | None:
    """Parse a note's `---`-delimited YAML frontmatter block into a mapping,
    or None if the text is not a well-formed note."""
    lines = text.splitlines()
    if not lines or lines[0].strip() != "---":
        return None
    for index in range(1, len(lines)):
        if lines[index].strip() == "---":
            block = "\n".join(lines[1:index])
            parsed = yaml.safe_load(block)
            return parsed if isinstance(parsed, dict) else None
    return None


def _scalar(value: Any, key: str) -> Any:
    """Extract the representative scalar of an axis frontmatter value: the
    tagger nests `field`/`claim_type`/`theory_school` as `{primary, ...}` and
    `empirical_scope` as `{value, polity?}`; `role_in_argument` is already a
    flat scalar. A value that is already a scalar is returned as-is."""
    if isinstance(value, dict):
        return value.get(key)
    return value


def parse_note(path: Path) -> dict[str, Any] | None:
    """Parse one prose note into a flat gold record (the representative scalar
    of each axis), or None if the note is malformed or missing a chunk_id.
    `source` is derived from the chunk_id, not read from frontmatter (the
    vault note carries no top-level source key)."""
    frontmatter = _split_frontmatter(path.read_text(encoding="utf-8"))
    if frontmatter is None:
        return None
    chunk_id = frontmatter.get("chunk_id")
    if not isinstance(chunk_id, str) or not chunk_id.strip():
        return None
    return {
        "chunk_id": chunk_id,
        "source": source_id_of(chunk_id),
        "section": frontmatter.get("section", ""),
        "chunk_text": frontmatter.get("chunk_text", ""),
        "field": _scalar(frontmatter.get("field"), "primary"),
        "empirical_scope": _scalar(frontmatter.get("empirical_scope"), "value"),
        "polities_touched": frontmatter.get("polities_touched", []) or [],
        "role_in_argument": frontmatter.get("role_in_argument"),
        "claim_type": _scalar(frontmatter.get("claim_type"), "primary"),
        "theory_school": _scalar(frontmatter.get("theory_school"), "primary"),
    }


def _read_frame(prose_dir: Path) -> list[dict[str, Any]]:
    """Read every prose note under `prose_dir`, drop back-matter (by title)
    and non-substantive fragments (by body, #131), and return the
    substantive records sorted by chunk_id (a stable base order)."""
    records = []
    for path in sorted(prose_dir.glob("*.md")):
        record = parse_note(path)
        if record is None or _is_back_matter(record["section"]):
            continue
        if not _is_substantive(record["chunk_text"]):
            continue
        records.append(record)
    records.sort(key=lambda r: r["chunk_id"])
    return records


def load_source_types(sources_path: Path) -> dict[str, str]:
    """Load the operator-declared `source_id -> book|paper` manifest, or an
    empty mapping when it is absent (source-type balancing is then skipped)."""
    if not sources_path.is_file():
        return {}
    with sources_path.open("r", encoding="utf-8") as handle:
        document = yaml.safe_load(handle) or {}
    return {str(key): str(value) for key, value in document.items()}


def select_chunks(
    records: list[dict[str, Any]],
    source_types: dict[str, str],
    *,
    min_size: int,
    max_size: int,
    seed: int,
) -> list[dict[str, Any]]:
    """Choose a stratified, deterministic subset of `records`.

    First guarantee coverage: at least one chunk for every represented value
    of each of the three balancing axes (field × empirical_scope ×
    role_in_argument), then at least one chunk for every source-type present
    in the corpus (when a manifest is supplied). Then fill toward `min_size`
    in a stable, seeded order, never exceeding `max_size` or the number of
    available chunks. Selection is returned sorted by chunk_id."""
    pool = list(records)
    random.Random(seed).shuffle(pool)

    selected: list[dict[str, Any]] = []
    selected_ids: set[str] = set()

    def add(record: dict[str, Any]) -> None:
        if record["chunk_id"] not in selected_ids:
            selected.append(record)
            selected_ids.add(record["chunk_id"])

    # Axis coverage: one chunk per represented value of each balancing axis.
    for axis in STRATA_AXES:
        for value in sorted({r[axis] for r in pool if r[axis] is not None}, key=str):
            if any(r[axis] == value for r in selected):
                continue
            pick = next(
                (r for r in pool if r[axis] == value and r["chunk_id"] not in selected_ids), None
            )
            if pick is not None:
                add(pick)

    # Source-type coverage: one chunk per declared type present in the corpus.
    if source_types:
        present = sorted({source_types[r["source"]] for r in pool if r["source"] in source_types})
        for source_type in present:
            if any(source_types.get(r["source"]) == source_type for r in selected):
                continue
            pick = next(
                (
                    r
                    for r in pool
                    if source_types.get(r["source"]) == source_type
                    and r["chunk_id"] not in selected_ids
                ),
                None,
            )
            if pick is not None:
                add(pick)

    # Fill toward the band ceiling in the seeded pool order, clamped to what
    # is available -- a larger labeled set is more useful and still within the
    # band. Coverage alone may already exceed the ceiling on a corpus with
    # many axis values; it is mandatory and is never trimmed.
    target = min(max_size, len(pool))
    for record in pool:
        if len(selected) >= target:
            break
        add(record)

    selected.sort(key=lambda r: r["chunk_id"])
    return selected


def _clear_dir(directory: Path) -> None:
    """Remove `directory` and its contents so a re-run never accumulates
    stale records, then recreate it empty."""
    shutil.rmtree(directory, ignore_errors=True)
    directory.mkdir(parents=True, exist_ok=True)


def run_gold_sample(
    vault_dir: Path | None = None,
    gold_dir: Path | None = None,
    config_path: Path = DEFAULT_PIPELINE_CONFIG_PATH,
    min_size: int = DEFAULT_MIN_SIZE,
    max_size: int = DEFAULT_MAX_SIZE,
    seed: int = DEFAULT_SEED,
) -> list[Path]:
    """Run gold-set sampling: read the tagged prose vault, exclude back-matter,
    select a stratified set, and write one JSON chunk record per selection
    under `<gold_dir>/chunks/`. Returns the written record paths, sorted.

    Raises `EmptyFrameError` when no substantive prose note is available to
    sample. Logs to stderr when the sources manifest is absent (source-type
    balancing skipped) or when the band clamps to the available chunk count."""
    if vault_dir is None:
        vault_dir = _default_vault_dir(config_path)
    if gold_dir is None:
        gold_dir = _default_gold_dir(config_path)

    prose_dir = vault_dir / "prose"
    records = _read_frame(prose_dir) if prose_dir.is_dir() else []
    if not records:
        raise EmptyFrameError(prose_dir)

    sources_path = gold_dir / "sources.yaml"
    source_types = load_source_types(sources_path)
    if not source_types:
        print(
            f"note: no source-type manifest at {sources_path}; stratifying on "
            f"field × empirical_scope × role_in_argument only "
            f"(source-type coverage skipped)",
            file=sys.stderr,
        )

    selected = select_chunks(records, source_types, min_size=min_size, max_size=max_size, seed=seed)

    if len(selected) < min_size:
        print(
            f"note: selected {len(selected)} chunk(s); the {min_size}-{max_size} "
            f"band clamped to the {len(records)} available substantive chunk(s)",
            file=sys.stderr,
        )

    chunks_dir = gold_dir / "chunks"
    _clear_dir(chunks_dir)

    written = []
    for record in selected:
        ordered = {key: record.get(key) for key in RECORD_FIELDS}
        path = chunks_dir / f"{record['chunk_id']}.json"
        path.write_text(
            json.dumps(ordered, indent=2, sort_keys=True, ensure_ascii=False), encoding="utf-8"
        )
        written.append(path)

    return sorted(written)


def _load_gold_records(chunks_dir: Path) -> list[dict[str, Any]]:
    """Read every sampled chunk record under `chunks_dir`, sorted by chunk_id
    for a stable row order."""
    records = []
    for path in sorted(chunks_dir.glob("*.json")):
        records.append(json.loads(path.read_text(encoding="utf-8")))
    records.sort(key=lambda r: r.get("chunk_id", ""))
    return records


def _axis_vocabularies(domain_dir: str | Path) -> dict[str, list[str]]:
    """Load the codebook and return the sorted controlled vocabulary for each
    of the four axis columns -- the dropdown option lists."""
    try:
        codebook = load_codebook(domain_dir)
    except CodebookError as exc:
        raise CodebookLoadError(exc) from exc
    return {axis: sorted(codebook.axes.get(axis, {})) for axis in AXIS_COLUMNS}


def _write_vocab_sheet(workbook: Workbook, vocabularies: dict[str, list[str]]) -> dict[str, str]:
    """Write each axis's vocabulary into its own column on a hidden helper
    sheet and return, per axis, the absolute range reference the dropdowns
    point at. A helper sheet (rather than an inline list) is required because
    claim_type/theory_school vocabularies exceed Excel's ~255-char inline
    data-validation limit."""
    sheet = workbook.create_sheet(VOCAB_SHEET_NAME)
    sheet.sheet_state = "hidden"

    ranges: dict[str, str] = {}
    for index, axis in enumerate(AXIS_COLUMNS, start=1):
        letter = get_column_letter(index)
        sheet.cell(row=1, column=index, value=axis)
        values = vocabularies[axis]
        for offset, value in enumerate(values, start=2):
            sheet.cell(row=offset, column=index, value=value)
        last_row = 1 + len(values)
        ranges[axis] = f"{VOCAB_SHEET_NAME}!${letter}$2:${letter}${last_row}"
    return ranges


def build_workbook(records: list[dict[str, Any]], vocabularies: dict[str, list[str]]) -> Workbook:
    """Build the label-sheet workbook: the Appendix-I header row, one row per
    sampled chunk (provenance + pre-filled field/empirical_scope/
    polities_touched, blind claim_type/theory_school/notes), and a
    codebook-sourced dropdown on each of the four axis columns."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = LABEL_SHEET_NAME

    for col, name in enumerate(SHEET_COLUMNS, start=1):
        sheet.cell(row=1, column=col, value=name)

    for row_index, record in enumerate(records, start=2):
        for col, name in enumerate(SHEET_COLUMNS, start=1):
            if name in AXIS_COLUMNS and name not in PRELABELED_COLUMNS:
                continue  # blind columns arrive empty for the Academic
            if name == "notes":
                continue  # ships empty
            if name == "polities_touched":
                sheet.cell(row=row_index, column=col, value="; ".join(record.get(name) or []))
                continue
            sheet.cell(row=row_index, column=col, value=record.get(name))

    ranges = _write_vocab_sheet(workbook, vocabularies)
    last_data_row = 1 + len(records)
    for col, name in enumerate(SHEET_COLUMNS, start=1):
        if name not in AXIS_COLUMNS:
            continue
        validation = DataValidation(type="list", formula1=ranges[name], allow_blank=True)
        sheet.add_data_validation(validation)
        letter = get_column_letter(col)
        validation.add(f"{letter}2:{letter}{last_data_row}")

    return workbook


def run_gold_sheet(
    gold_dir: Path | None = None,
    domain_dir: str | Path = DEFAULT_DOMAIN_DIR,
    config_path: Path = DEFAULT_PIPELINE_CONFIG_PATH,
) -> Path:
    """Render the sampled gold set under `<gold_dir>/chunks/` into
    `<gold_dir>/label_sheet.xlsx` (Appendix I), overwriting any prior sheet in
    place. Returns the written path. Raises `MissingChunksError` when no
    sampled records exist and `CodebookLoadError` when the dropdown vocabulary
    cannot be loaded. Offline -- no LLM call."""
    if gold_dir is None:
        gold_dir = _default_gold_dir(config_path)

    chunks_dir = gold_dir / "chunks"
    records = _load_gold_records(chunks_dir) if chunks_dir.is_dir() else []
    if not records:
        raise MissingChunksError(chunks_dir)

    vocabularies = _axis_vocabularies(domain_dir)
    workbook = build_workbook(records, vocabularies)

    gold_dir.mkdir(parents=True, exist_ok=True)
    sheet_path = gold_dir / "label_sheet.xlsx"
    workbook.save(sheet_path)
    return sheet_path


def _sheet_chunk_count(sheet_path: Path) -> int:
    """Number of labelable data rows in the sheet (its rows minus the header)."""
    worksheet = load_workbook(sheet_path, read_only=True).worksheets[0]
    return max(worksheet.max_row - 1, 0)


def _academic_readme(chunk_count: int) -> str:
    """The Academic-facing labeling instructions shipped in the bundle (§9
    hybrid labeling). Names every axis and the return location so the human
    reviewer needs nothing but this folder."""
    return f"""# Gold label sheet -- for the Academic

This bundle is the gold-corpus labeling task for the Axial tagging eval. It
holds **{chunk_count} sampled chunks**, one per row in `label_sheet.xlsx`.
Your labels become the answer key the pipeline's tagging is scored against.

## What to fill in

Each row is one chunk of source prose (`chunk_text`), with its provenance
(`chunk_id`, `source`, `section`). Label the four axis columns using the
in-cell **dropdowns** only -- do not free-type a value.

- **Blind** (label from scratch, the cell arrives empty):
  - `claim_type`
  - `theory_school`
- **Pre-labeled** (the pipeline's guess is filled in; correct it where wrong,
  leave it where right):
  - `field`
  - `empirical_scope`
  - `polities_touched` -- **free text, no dropdown.** The cell is pre-filled
    with the tagger's guessed polities (semicolon-separated); correct it by
    typing where wrong, leave it where right. An empty cell means the tagger
    found no engaged polity -- leave it empty if that's correct.

Leave the provenance columns (`chunk_id`, `source`, `section`) untouched.
`notes` is optional free text for any chunk you want to flag.

## How to return it

Save the filled sheet and place it under `{LABELS_RETURN_DIR}` in the repo.
The same sheet, once returned, is read directly as the scoring answer key --
there is no other form to fill.

See `manifest.json` in this folder for the machine-readable summary.
"""


def run_gold_deliver(
    gold_dir: Path | None = None,
    config_path: Path = DEFAULT_PIPELINE_CONFIG_PATH,
    stamp: str | None = None,
) -> Path:
    """Package the generated `<gold_dir>/label_sheet.xlsx` into a dated,
    self-contained handoff bundle for the Academic under
    `<gold_dir>/delivery/<stamp>/`: a byte-identical copy of the sheet, a
    human-facing `README-for-academic.md`, and a machine-readable
    `manifest.json`. Returns the delivery directory. Raises
    `MissingSheetError` when the sheet has not been generated yet (run
    `axial gold sheet` first). The delivery folder is cleared before writing,
    so a re-run never leaves stale files. Local and offline -- no network, no
    Drive. `stamp` defaults to today's date (`YYYY-MM-DD`)."""
    if gold_dir is None:
        gold_dir = _default_gold_dir(config_path)
    if stamp is None:
        stamp = datetime.date.today().isoformat()

    sheet_path = gold_dir / "label_sheet.xlsx"
    if not sheet_path.is_file():
        raise MissingSheetError(sheet_path)

    chunk_count = _sheet_chunk_count(sheet_path)

    delivery_dir = gold_dir / "delivery" / stamp
    _clear_dir(delivery_dir)

    shutil.copyfile(sheet_path, delivery_dir / "label_sheet.xlsx")

    manifest = {
        "sheet": "label_sheet.xlsx",
        "delivered": stamp,
        "chunk_count": chunk_count,
        "columns": list(SHEET_COLUMNS),
        "axes": list(AXIS_COLUMNS),
        "blind_axes": list(BLIND_COLUMNS),
        "prelabeled_axes": list(PRELABELED_COLUMNS),
        "prelabeled_freetext": list(PRELABELED_FREETEXT_COLUMNS),
        "return_to": LABELS_RETURN_DIR,
    }
    (delivery_dir / "manifest.json").write_text(
        json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )

    (delivery_dir / "README-for-academic.md").write_text(
        _academic_readme(chunk_count), encoding="utf-8"
    )

    return delivery_dir
