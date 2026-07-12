"""Inner unit tests for gold-set sampling (src/axial/gold.py, issue #53).

These cover the logic the outer acceptance test (tests/test_gold_sample.py)
cannot exercise on a tiny fixture -- chiefly stratified DOWN-selection when
the available frame exceeds the band ceiling -- plus the classifier, the
source_id derivation, the frontmatter parse, and the run-level edge cases.
"""

from __future__ import annotations

import json

import pytest
import yaml

from openpyxl import load_workbook

from axial.gold import (
    AXIS_COLUMNS,
    SHEET_COLUMNS,
    EmptyFrameError,
    MissingChunksError,
    _axis_vocabularies,
    _is_back_matter,
    build_workbook,
    load_source_types,
    parse_note,
    run_gold_sample,
    run_gold_sheet,
    select_chunks,
    source_id_of,
)
from axial.tag import DEFAULT_DOMAIN_DIR


def _record(chunk_id, field, scope, role, source=None):
    return {
        "chunk_id": chunk_id,
        "source": source if source is not None else source_id_of(chunk_id),
        "section": "Body",
        "chunk_text": f"text {chunk_id}",
        "field": field,
        "empirical_scope": scope,
        "role_in_argument": role,
        "claim_type": "state-formation",
        "theory_school": "bellicist",
    }


class TestIsBackMatter:
    @pytest.mark.parametrize(
        "section",
        [
            "Bibliography",
            "BIBLIOGRAPHY",
            " References ",
            "Index",
            "Subject Index",
            "Endnotes",
            "Notes",
            "Preface",
            "Acknowledgements",
            "Appendix",
            "Appendix A",
            "Appendix II: Coding Rules",
            "Annex 1",
            "Table of Contents",
        ],
    )
    def test_excluded(self, section):
        assert _is_back_matter(section) is True

    @pytest.mark.parametrize(
        "section",
        [
            "Introduction",
            "Chapter One",
            "The Long March",
            "Comparative Cases",
            "Conclusion",
            "Findings and Analysis",
        ],
    )
    def test_kept(self, section):
        assert _is_back_matter(section) is False


class TestSourceIdOf:
    def test_strips_order_slug_index(self):
        cid = "agamben-state-of-exception-b22edc40e0fc_10_4-chapter-one_001"
        assert source_id_of(cid) == "agamben-state-of-exception-b22edc40e0fc"

    def test_short_id_returned_as_is(self):
        assert source_id_of("weird") == "weird"


class TestSelectCoversWhenDownsampling:
    def _big_frame(self):
        # 3 fields × 2 scopes × 2 roles = 12 strata cells, 20 chunks each ->
        # 240 records, far above the band ceiling.
        fields = ["state", "violence", "ideology"]
        scopes = ["scope:general", "scope:country-case"]
        roles = ["role:claim", "role:evidence"]
        records = []
        n = 0
        for f in fields:
            for s in scopes:
                for r in roles:
                    for _ in range(20):
                        n += 1
                        records.append(_record(f"src-{n:04d}_1_body_001", f, s, r))
        return records, fields, scopes, roles

    def test_size_clamped_to_ceiling_and_all_axis_values_covered(self):
        records, fields, scopes, roles = self._big_frame()
        selected = select_chunks(records, {}, min_size=100, max_size=120, seed=0)

        assert len(selected) == 120, "should fill to the band ceiling when the frame is large"
        assert {r["field"] for r in selected} == set(fields)
        assert {r["empirical_scope"] for r in selected} == set(scopes)
        assert {r["role_in_argument"] for r in selected} == set(roles)

    def test_deterministic_same_seed(self):
        records, *_ = self._big_frame()
        a = select_chunks(records, {}, min_size=100, max_size=120, seed=7)
        b = select_chunks(records, {}, min_size=100, max_size=120, seed=7)
        assert [r["chunk_id"] for r in a] == [r["chunk_id"] for r in b]

    def test_returned_sorted_by_chunk_id(self):
        records, *_ = self._big_frame()
        selected = select_chunks(records, {}, min_size=100, max_size=120, seed=3)
        ids = [r["chunk_id"] for r in selected]
        assert ids == sorted(ids)


class TestSelectCoversSourceTypes:
    def test_rare_source_type_is_guaranteed_a_chunk(self):
        # 40 book chunks, a single paper chunk. A naive top-N by chunk_id
        # could miss the lone paper; source-type coverage must include it.
        records = [
            _record(f"book-src-{i:04d}_1_body_001", "state", "scope:general", "role:claim")
            for i in range(40)
        ]
        records.append(_record("paper-src-0001_1_body_001", "state", "scope:general", "role:claim"))
        source_types = {source_id_of(r["chunk_id"]): "book" for r in records[:40]}
        source_types["paper-src-0001"] = "paper"

        selected = select_chunks(records, source_types, min_size=5, max_size=8, seed=0)
        selected_types = {source_types[r["source"]] for r in selected}
        assert selected_types == {"book", "paper"}


class TestParseNote:
    def test_extracts_nested_scalars(self, tmp_path):
        cid = "src-abc-000000000001_1_intro_001"
        note = {
            "chunk_id": cid,
            "section": "Introduction",
            "chunk_text": "Some prose.",
            "role_in_argument": "role:claim",
            "field": {"primary": "state", "secondary": ["violence"]},
            "claim_type": {"primary": "state-formation", "secondary": None, "subtags": []},
            "theory_school": {"primary": "bellicist", "status": "candidate"},
            "empirical_scope": {"value": "scope:country-case", "country": "Syria"},
        }
        path = tmp_path / f"{cid}.md"
        block = yaml.safe_dump(note, sort_keys=False, allow_unicode=True)
        path.write_text(f"---\n{block}---\n# Introduction\n\nSome prose.\n", encoding="utf-8")

        record = parse_note(path)
        assert record["chunk_id"] == cid
        assert record["source"] == "src-abc-000000000001"
        assert record["field"] == "state"
        assert record["empirical_scope"] == "scope:country-case"
        assert record["role_in_argument"] == "role:claim"
        assert record["claim_type"] == "state-formation"
        assert record["theory_school"] == "bellicist"

    def test_non_note_returns_none(self, tmp_path):
        path = tmp_path / "not-a-note.md"
        path.write_text("just some text, no frontmatter\n", encoding="utf-8")
        assert parse_note(path) is None


class TestLoadSourceTypes:
    def test_absent_returns_empty(self, tmp_path):
        assert load_source_types(tmp_path / "sources.yaml") == {}

    def test_reads_flat_mapping(self, tmp_path):
        path = tmp_path / "sources.yaml"
        path.write_text(yaml.safe_dump({"a-1": "book", "b-2": "paper"}), encoding="utf-8")
        assert load_source_types(path) == {"a-1": "book", "b-2": "paper"}


class TestRunGoldSample:
    def _seed(self, tmp_path, specs):
        prose = tmp_path / "vault" / "prose"
        prose.mkdir(parents=True)
        for cid, field, scope, role, section in specs:
            note = {
                "chunk_id": cid,
                "section": section,
                "chunk_text": f"text {cid}",
                "role_in_argument": role,
                "field": {"primary": field},
                "claim_type": {"primary": "state-formation"},
                "theory_school": {"primary": "bellicist"},
                "empirical_scope": {"value": scope},
            }
            block = yaml.safe_dump(note, sort_keys=False, allow_unicode=True)
            (prose / f"{cid}.md").write_text(
                f"---\n{block}---\n# {section}\n\ntext\n", encoding="utf-8"
            )
        return tmp_path / "vault", tmp_path / "gold"

    def test_empty_frame_raises(self, tmp_path):
        vault = tmp_path / "vault"
        (vault / "prose").mkdir(parents=True)
        with pytest.raises(EmptyFrameError):
            run_gold_sample(vault_dir=vault, gold_dir=tmp_path / "gold")

    def test_all_back_matter_raises(self, tmp_path):
        vault, gold = self._seed(
            tmp_path,
            [("s-1_1_bib_001", "state", "scope:general", "role:claim", "Bibliography")],
        )
        with pytest.raises(EmptyFrameError):
            run_gold_sample(vault_dir=vault, gold_dir=gold)

    def test_writes_records_and_skips_manifest(self, tmp_path):
        vault, gold = self._seed(
            tmp_path,
            [
                ("s-1_1_intro_001", "state", "scope:general", "role:claim", "Introduction"),
                ("s-1_2_body_001", "violence", "scope:regional", "role:evidence", "Chapter"),
            ],
        )
        # No sources.yaml -> source-type coverage skipped, still writes.
        written = run_gold_sample(vault_dir=vault, gold_dir=gold)
        assert len(written) == 2
        records = [json.loads(p.read_text(encoding="utf-8")) for p in written]
        assert {r["field"] for r in records} == {"state", "violence"}
        # record carries the full field set in a stable key order.
        for r in records:
            assert set(r) == {
                "chunk_id",
                "source",
                "section",
                "chunk_text",
                "field",
                "empirical_scope",
                "role_in_argument",
                "claim_type",
                "theory_school",
            }


def _gold_record(chunk_id, field, scope, claim, theory):
    return {
        "chunk_id": chunk_id,
        "source": source_id_of(chunk_id),
        "section": "Body",
        "chunk_text": f"text {chunk_id}",
        "field": field,
        "empirical_scope": scope,
        "role_in_argument": "role:claim",
        "claim_type": claim,
        "theory_school": theory,
    }


class TestAxisVocabularies:
    def test_loads_codebook_vocab(self):
        vocab = _axis_vocabularies(DEFAULT_DOMAIN_DIR)
        assert set(vocab) == set(AXIS_COLUMNS)
        assert set(vocab["field"]) == {"state", "violence", "ideology"}
        assert "scope:country-case" in vocab["empirical_scope"]
        # long vocabularies (the reason for a helper sheet, not inline lists)
        assert len(vocab["claim_type"]) > 10
        assert len(vocab["theory_school"]) > 10
        # sorted for determinism
        assert vocab["claim_type"] == sorted(vocab["claim_type"])


class TestBuildWorkbook:
    def _vocab(self):
        return {
            "field": ["ideology", "state", "violence"],
            "empirical_scope": ["scope:country-case", "scope:general"],
            "claim_type": ["a-claim", "b-claim"],
            "theory_school": ["a-school", "b-school"],
        }

    def _records(self):
        return [
            _gold_record("s-1_1_a_001", "state", "scope:general", "a-claim", "a-school"),
            _gold_record("s-1_2_b_001", "violence", "scope:country-case", "b-claim", "b-school"),
        ]

    def test_header_and_row_count(self):
        wb = build_workbook(self._records(), self._vocab())
        ws = wb.worksheets[0]
        header = [ws.cell(row=1, column=i + 1).value for i in range(len(SHEET_COLUMNS))]
        assert header == list(SHEET_COLUMNS)
        assert ws.max_row == 3  # header + 2 rows

    def test_prefilled_and_blind_cells(self):
        records = self._records()
        wb = build_workbook(records, self._vocab())
        ws = wb.worksheets[0]
        idx = {name: i + 1 for i, name in enumerate(SHEET_COLUMNS)}
        # row 2 corresponds to the first (chunk_id-sorted) record
        assert ws.cell(row=2, column=idx["field"]).value == "state"
        assert ws.cell(row=2, column=idx["empirical_scope"]).value == "scope:general"
        assert ws.cell(row=2, column=idx["claim_type"]).value in (None, "")
        assert ws.cell(row=2, column=idx["theory_school"]).value in (None, "")
        assert ws.cell(row=2, column=idx["notes"]).value in (None, "")

    def test_axis_dropdowns_reference_vocab(self):
        wb = build_workbook(self._records(), self._vocab())
        ws = wb.worksheets[0]
        idx = {name: i + 1 for i, name in enumerate(SHEET_COLUMNS)}
        for axis in AXIS_COLUMNS:
            from openpyxl.utils import get_column_letter

            coord = f"{get_column_letter(idx[axis])}2"
            dv = next((v for v in ws.data_validations.dataValidation if coord in v.sqref), None)
            assert dv is not None, f"no dropdown on axis column {axis}"
            assert dv.type == "list"
            # formula1 points at the hidden vocab sheet's column range
            assert "vocab!" in (dv.formula1 or "")


class TestRunGoldSheet:
    def _seed_chunks(self, tmp_path, records):
        chunks = tmp_path / "gold" / "chunks"
        chunks.mkdir(parents=True)
        for r in records:
            (chunks / f"{r['chunk_id']}.json").write_text(json.dumps(r), encoding="utf-8")
        return tmp_path / "gold"

    def test_missing_chunks_raises(self, tmp_path):
        (tmp_path / "gold" / "chunks").mkdir(parents=True)
        with pytest.raises(MissingChunksError):
            run_gold_sheet(gold_dir=tmp_path / "gold", domain_dir=DEFAULT_DOMAIN_DIR)

    def test_writes_and_overwrites_in_place(self, tmp_path):
        records = [
            _gold_record("s-1_1_a_001", "state", "scope:general", "state-formation", "bellicist"),
        ]
        gold = self._seed_chunks(tmp_path, records)
        path = run_gold_sheet(gold_dir=gold, domain_dir=DEFAULT_DOMAIN_DIR)
        assert path.is_file()
        first_rows = load_workbook(path).worksheets[0].max_row
        # re-run overwrites in place, same row count
        run_gold_sheet(gold_dir=gold, domain_dir=DEFAULT_DOMAIN_DIR)
        assert load_workbook(path).worksheets[0].max_row == first_rows == 2
