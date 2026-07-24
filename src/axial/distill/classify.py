"""Stage-5d TF-IDF classifier for the two blind tag axes -- `claim_type`
and `theory_school` (issues #351/#352, DEC-37/DEC-38, plan
`plans/phase-a-completion/README.md` stage 5d).

This module implements the ONE technique DEC-38 measured as beating dense
sentence embeddings specifically for these two axes: `TfidfVectorizer`
(word/bigram bag-of-words) + `LogisticRegression`, trained directly on the
corpus's own existing best-of-3 production tags (stage 4 already tagged all
17,824 chunks -- no fresh LLM relabel needed, DEC-37). The other stage-5d
axes (`field_primary`, `role_in_argument`, `empirical_scope_value`) are a
**different job**: DEC-39 found them better served by the embedding
classifier already exercised in the DEC-37/38 investigation, not this
module's TF-IDF pipeline. This module deliberately does not generalize to a
multi-feature/multi-technique abstraction those callers don't exist yet to
need (repo convention: no speculative generality) -- a future axis with its
own measured-best technique gets its own small module, the same way this
one does.

A predicted-probability confidence threshold stands in for the
automate/defer-to-LLM split (DEC-37): above threshold, the classifier's
prediction would replace the LLM call; below, the existing LLM tag pass
handles it, unchanged. **This module is a measurement/eval artifact only**
-- it is never wired into `axial.tag.run_tag` or any production tagging
path (mirrors 5b's `readiness.py`, which never touched production tagging
either). Whether to actually build the automate path is separate spec
drift for the founder to adjudicate (DEC-32).

Training set: every vault chunk carrying a non-empty value for the axis,
**excluding the gold-sampled chunk_ids** (leakage-free -- DEC-38's method),
and **excluding classes with fewer than `DEFAULT_MIN_CLASS_COUNT` (6)
training examples** (too few to learn a stable boundary; this exact
threshold is what DEC-37/38 measured against). Evaluated against the
independent gold sheet (`data/gold/labels/label_sheet.xlsx`, DEC-29/30's
simulated-academic gold set) -- never against the tagger's own labels (an
internal-only check is misleading, proven twice in the DEC-37 investigation).

Reuses `axial.query.reader`'s shared primitives (`_iter_chunk_frontmatter`,
`_require`) the same way `axial.distill.embed._load_chunk_records` does --
built independently here rather than importing that private function
directly, since this module reads a different projection (chunk_text plus
one axis's nested `primary`, not the full flattened metadata row).

`scikit-learn` (the `distill` dependency group, alongside 5a/5b's
`sentence-transformers`/`lancedb`/`hdbscan`) is imported lazily, inside
`_default_train_fn`, never at module level -- importing this module (e.g.
from `axial.cli`) never requires it; only running the pass does.
`openpyxl` (reading the gold sheet) is a base dependency already imported
at module level elsewhere in this codebase (`axial.gold`, `axial.eval`), so
it is imported at module level here too, matching that precedent.

DEC-23: the emitted manifest carries chunk_ids, tag values, counts, and
confidence scores only -- never `chunk_text`. Nothing in this module
persists a chunk's prose anywhere.
"""

from __future__ import annotations

import json
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from axial.distill.staleness import resolve_current_pin
from axial.eval import corpus_pin as _corpus_pin
from axial.paths import DEFAULT_PIPELINE_CONFIG_PATH, default_vault_dir
from axial.query.reader import MissingVaultDirError as _ReaderMissingVaultDirError
from axial.query.reader import QueryError as _ReaderQueryError
from axial.query.reader import _iter_chunk_frontmatter, _require

# The two axes this module is validated for (DEC-38) -- NOT every stage-5d
# candidate axis (see module docstring).
AXES = ("claim_type", "theory_school")

# TF-IDF + LogisticRegression config, pinned to the exact values DEC-38
# measured against the real corpus (see module docstring's citation).
DEFAULT_MAX_FEATURES = 20000
DEFAULT_NGRAM_RANGE = (1, 2)
DEFAULT_MIN_DF = 2
DEFAULT_STOP_WORDS = "english"
DEFAULT_MAX_ITER = 2000

# A class below this many training examples is dropped rather than trained
# on -- too few to learn a stable boundary (DEC-38's measured threshold).
DEFAULT_MIN_CLASS_COUNT = 6

# The confidence-threshold sweep the manifest reports a curve across,
# rather than a single operating point (per the issue: coverage and
# accuracy-on-covered both move with the threshold, so one number hides the
# tradeoff).
DEFAULT_THRESHOLDS = (0.5, 0.6, 0.7, 0.8)

# Derived artifact, alongside 5a/5b's manifests under data/distill/
# (gitignored, DEC-23) -- one manifest per axis.
DEFAULT_MANIFEST_DIR = Path("data/distill")

# The gold answer key (DEC-29/30's simulated-academic gold set) and the
# tagger's own gold-agreement report, if it has been generated -- both
# gitignored (DEC-23), read only, never written by this module.
DEFAULT_GOLD_SHEET_PATH = Path("data/gold/labels/label_sheet.xlsx")
DEFAULT_EVAL_REPORT_PATH = Path("data/gold/labels/eval_report.json")

# One (predicted_label, confidence) pair per input text, in input order.
PredictFn = Callable[[list[str]], list[tuple[str, float]]]
# Fits on (texts, labels) and returns the fitted PredictFn -- the seam this
# module's own inner unit tests use to exercise the training/evaluation/
# manifest logic without a real TF-IDF+LogisticRegression fit, mirroring
# `axial.distill.embed`'s `encoder` and `axial.distill.readiness`'s
# `cluster_fn` injection seams.
TrainFn = Callable[[list[str], list[str]], PredictFn]


def _manifest_path_for(axis: str) -> Path:
    return DEFAULT_MANIFEST_DIR / f"classify_{axis}_manifest.json"


class ClassifyError(Exception):
    """Base class for all stage-5d classifier errors."""


class UnknownAxisError(ClassifyError):
    """Raised when `axis` is not one of `AXES` -- this module implements
    the TF-IDF technique validated for `claim_type`/`theory_school` only
    (see module docstring); the other stage-5d axes are not its job."""

    def __init__(self, axis: str):
        self.axis = axis
        super().__init__(f"unknown axis {axis!r}; expected one of {AXES!r}")


class NoChunksToClassifyError(ClassifyError):
    """Raised when the vault's `prose/` directory is absent or holds no
    chunk notes -- mirrors `axial.distill.embed.NoChunksToEmbedError`."""

    def __init__(self, prose_dir: Path):
        self.prose_dir = prose_dir
        super().__init__(
            f"no prose chunks found under {prose_dir} to train on; run the "
            f"tagging pipeline (`axial vault write`) first"
        )


class CorpusPinRequiredError(ClassifyError):
    """Raised when no corpus-pin manifest can be resolved (DEC-35: every
    stage-5 artifact records the pin it was built from) -- mirrors
    `axial.distill.embed.CorpusPinRequiredError`."""

    def __init__(self, cause: _corpus_pin.CorpusPinError):
        self.cause = cause
        super().__init__(
            f"classify pass requires a corpus pin to record provenance against "
            f"({cause}); run `axial pin write <name>` first"
        )


class MalformedVaultNoteError(ClassifyError):
    """Wraps `axial.query.reader.QueryError` -- mirrors
    `axial.distill.embed.MalformedVaultNoteError`."""

    def __init__(self, cause: _ReaderQueryError):
        self.cause = cause
        super().__init__(f"malformed vault note encountered while classifying: {cause}")


class MissingGoldSheetError(ClassifyError):
    """Raised when no returned gold answer-key sheet exists -- this pass's
    evaluation is meaningless without an independent referee (DEC-37/38:
    checking a classifier only against the tagger's own labels is
    misleading, proven twice)."""

    def __init__(self, gold_sheet_path: Path):
        self.gold_sheet_path = gold_sheet_path
        super().__init__(
            f"no gold answer-key sheet found at {gold_sheet_path}; place the "
            f"returned label_sheet.xlsx under data/gold/labels/ (see `axial eval`) "
            f"before running `axial distill classify`"
        )


class NoGoldLabelsError(ClassifyError):
    """Raised when the gold sheet carries zero non-empty labels for `axis`
    -- there is nothing to evaluate against."""

    def __init__(self, axis: str, gold_sheet_path: Path):
        self.axis = axis
        self.gold_sheet_path = gold_sheet_path
        super().__init__(f"no gold labels for axis {axis!r} found in {gold_sheet_path}")


class NoTrainingDataError(ClassifyError):
    """Raised when, after excluding gold chunk_ids and dropping classes
    below `min_class_count`, zero training examples remain for `axis`."""

    def __init__(self, axis: str):
        self.axis = axis
        super().__init__(
            f"no training examples remain for axis {axis!r} after excluding gold "
            f"chunks and classes below the min-class-count floor"
        )


@dataclass(frozen=True)
class ThresholdMetrics:
    """One point on the confidence-threshold curve."""

    threshold: float
    coverage: float
    covered_count: int
    accuracy_on_covered: float | None


@dataclass(frozen=True)
class ClassifyResult:
    """The outcome of one `run_classify` call."""

    manifest_path: Path
    axis: str
    train_chunk_count: int
    dropped_classes: list[str]
    gold_chunk_count: int
    full_coverage_accuracy: float
    teacher_gold_agreement: float | None
    thresholds: list[ThresholdMetrics]
    corpus_pin_id: str
    vault_snapshot_hash: str


def _axis_value(frontmatter: dict[str, Any], axis: str) -> str:
    """The axis's representative scalar (the nested `primary`, matching
    every other tag axis's own frontmatter shape) -- `""` when the chunk
    carries no value for this axis, never `None` (mirrors
    `axial.distill.embed._flatten_metadata`'s own empty-string convention)."""
    return (frontmatter.get(axis) or {}).get("primary") or ""


def _load_vault_axis_records(vault_dir: Path, axis: str) -> list[tuple[str, str, str]]:
    """Every vault chunk's `(chunk_id, chunk_text, axis_value)`, sorted by
    `chunk_id` -- the same determinism convention
    `axial.distill.embed._load_chunk_records` uses."""
    records = []
    for path, frontmatter in _iter_chunk_frontmatter(vault_dir):
        chunk_id = _require(frontmatter, path, "chunk_id")
        chunk_text = _require(frontmatter, path, "chunk_text")
        records.append((chunk_id, chunk_text, _axis_value(frontmatter, axis)))
    records.sort(key=lambda record: record[0])
    return records


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text


def _load_gold_sheet(
    gold_sheet_path: Path, axis: str
) -> tuple[set[str], list[tuple[str, str, str]]]:
    """Read the returned gold answer-key sheet's data rows: every
    `chunk_id` present (for the training-exclusion set, regardless of
    whether this axis was labeled on that row) and the `(chunk_id,
    chunk_text, label)` triples carrying a non-empty label for `axis` (the
    evaluation set). Reads the header row by name rather than assuming a
    fixed column order, so this reader tolerates any column arrangement
    `axial.gold.build_workbook` or a hand-built fixture sheet happens to
    use."""
    worksheet = load_workbook(gold_sheet_path, read_only=True).worksheets[0]
    rows_iter = worksheet.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        return set(), []
    column_index = {str(name).strip(): index for index, name in enumerate(header) if name}
    if (
        "chunk_id" not in column_index
        or "chunk_text" not in column_index
        or axis not in column_index
    ):
        return set(), []

    chunk_ids: set[str] = set()
    eval_records: list[tuple[str, str, str]] = []
    for row in rows_iter:
        chunk_id = _normalize_cell(row[column_index["chunk_id"]])
        if not chunk_id:
            continue
        chunk_ids.add(chunk_id)
        label = _normalize_cell(row[column_index[axis]])
        if not label:
            continue
        chunk_text = _normalize_cell(row[column_index["chunk_text"]])
        eval_records.append((chunk_id, chunk_text, label))

    eval_records.sort(key=lambda record: record[0])
    return chunk_ids, eval_records


def _drop_rare_classes(
    records: list[tuple[str, str, str]], min_class_count: int
) -> tuple[list[tuple[str, str, str]], list[str]]:
    """Drop every record whose label occurs fewer than `min_class_count`
    times in `records` -- too few examples to learn a stable boundary
    (DEC-38's measured floor). Returns the filtered records and the sorted
    list of dropped class labels."""
    counts = Counter(label for _chunk_id, _chunk_text, label in records)
    dropped = sorted(label for label, count in counts.items() if count < min_class_count)
    dropped_set = set(dropped)
    filtered = [record for record in records if record[2] not in dropped_set]
    return filtered, dropped


def _full_coverage_accuracy(predictions: list[tuple[str, float]], labels: list[str]) -> float:
    if not labels:
        return 0.0
    correct = sum(
        1 for (predicted, _confidence), label in zip(predictions, labels) if predicted == label
    )
    return correct / len(labels)


def _threshold_metrics(
    predictions: list[tuple[str, float]], labels: list[str], threshold: float
) -> ThresholdMetrics:
    """Coverage (the fraction of gold chunks whose top prediction clears
    `threshold`) and accuracy-on-covered (accuracy restricted to that
    covered subset -- `None`, not `0.0`, when nothing clears the
    threshold, since accuracy is undefined over an empty set, not zero)."""
    total = len(labels)
    covered = [
        (predicted, label)
        for (predicted, confidence), label in zip(predictions, labels)
        if confidence >= threshold
    ]
    coverage = len(covered) / total if total else 0.0
    if covered:
        accuracy_on_covered = sum(1 for predicted, label in covered if predicted == label) / len(
            covered
        )
    else:
        accuracy_on_covered = None
    return ThresholdMetrics(
        threshold=threshold,
        coverage=coverage,
        covered_count=len(covered),
        accuracy_on_covered=accuracy_on_covered,
    )


def _load_teacher_gold_agreement(eval_report_path: Path, axis: str) -> float | None:
    """The tagger's own gold agreement for `axis`, from `eval_report_path`'s
    `per_axis_agreement` (`axial.eval.run_eval`'s report shape), or `None`
    when the report is absent/malformed/missing this axis -- reused as a
    comparison point, never required (the report may not exist yet)."""
    if not eval_report_path.is_file():
        return None
    try:
        report = json.loads(eval_report_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    agreement = report.get("per_axis_agreement", {})
    value = agreement.get(axis)
    return float(value) if isinstance(value, (int, float)) else None


def _default_train_fn(
    texts: list[str],
    labels: list[str],
    *,
    max_features: int = DEFAULT_MAX_FEATURES,
    ngram_range: tuple[int, int] = DEFAULT_NGRAM_RANGE,
    min_df: int = DEFAULT_MIN_DF,
    stop_words: str = DEFAULT_STOP_WORDS,
    max_iter: int = DEFAULT_MAX_ITER,
) -> PredictFn:
    """Fit the real TF-IDF + LogisticRegression pipeline (DEC-38's measured
    config; lazy `sklearn` import -- see module docstring) and return a
    closure predicting `(label, confidence)` pairs -- the top predicted
    class and its predicted probability."""
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.linear_model import LogisticRegression

    vectorizer = TfidfVectorizer(
        max_features=max_features,
        ngram_range=ngram_range,
        min_df=min_df,
        stop_words=stop_words,
    )
    features = vectorizer.fit_transform(texts)
    classifier = LogisticRegression(max_iter=max_iter)
    classifier.fit(features, labels)

    def predict(query_texts: list[str]) -> list[tuple[str, float]]:
        query_features = vectorizer.transform(query_texts)
        probabilities = classifier.predict_proba(query_features)
        classes = classifier.classes_
        results = []
        for row in probabilities:
            best_index = row.argmax()
            results.append((str(classes[best_index]), float(row[best_index])))
        return results

    return predict


def run_classify(
    axis: str,
    vault_dir: Path | None = None,
    gold_sheet_path: Path | None = None,
    eval_report_path: Path | None = None,
    manifest_path: Path | None = None,
    config_path: Path = DEFAULT_PIPELINE_CONFIG_PATH,
    evals_dir: Path | None = None,
    min_class_count: int = DEFAULT_MIN_CLASS_COUNT,
    thresholds: tuple[float, ...] = DEFAULT_THRESHOLDS,
    train_fn: TrainFn | None = None,
) -> ClassifyResult:
    """Train a TF-IDF + LogisticRegression classifier for `axis` (one of
    `AXES`) on every vault chunk's existing tag (excluding gold chunks and
    classes below `min_class_count`), evaluate it against the independent
    gold sheet at full coverage and at each of `thresholds`, and write a
    JSON manifest to `manifest_path` (default `_manifest_path_for(axis)`)
    recording the corpus-pin provenance, the pinned config, and the
    results. This pass is a measurement artifact only -- it never writes
    to the vault or touches `axial.tag.run_tag` (see module docstring).

    `train_fn`, when given, replaces the default TF-IDF+LogisticRegression
    fit (`_default_train_fn`) -- the seam this module's own inner unit
    tests use, mirroring `axial.distill.embed`'s `encoder` and
    `axial.distill.readiness`'s `cluster_fn` injection seams.

    Raises `UnknownAxisError` when `axis` is not in `AXES`,
    `NoChunksToClassifyError` when the vault holds no prose chunks,
    `CorpusPinRequiredError` when no corpus pin can be resolved,
    `MissingGoldSheetError` when no gold answer-key sheet exists,
    `NoGoldLabelsError` when the gold sheet has no labels for `axis`, and
    `NoTrainingDataError` when no training examples survive gold-exclusion
    and rare-class filtering -- all loud failures rather than a silently
    empty or misleading manifest."""
    if axis not in AXES:
        raise UnknownAxisError(axis)

    if vault_dir is None:
        vault_dir = default_vault_dir(config_path)
    vault_dir = Path(vault_dir)
    if gold_sheet_path is None:
        gold_sheet_path = DEFAULT_GOLD_SHEET_PATH
    gold_sheet_path = Path(gold_sheet_path)
    if eval_report_path is None:
        eval_report_path = DEFAULT_EVAL_REPORT_PATH
    eval_report_path = Path(eval_report_path)
    if manifest_path is None:
        manifest_path = _manifest_path_for(axis)
    manifest_path = Path(manifest_path)

    try:
        pin = resolve_current_pin(evals_dir)
    except _corpus_pin.CorpusPinError as exc:
        raise CorpusPinRequiredError(exc) from exc

    if not gold_sheet_path.is_file():
        raise MissingGoldSheetError(gold_sheet_path)
    gold_chunk_ids, gold_eval_records = _load_gold_sheet(gold_sheet_path, axis)
    if not gold_eval_records:
        raise NoGoldLabelsError(axis, gold_sheet_path)

    try:
        vault_records = _load_vault_axis_records(vault_dir, axis)
    except _ReaderMissingVaultDirError as exc:
        raise NoChunksToClassifyError(vault_dir / "prose") from exc
    except _ReaderQueryError as exc:
        raise MalformedVaultNoteError(exc) from exc
    if not vault_records:
        raise NoChunksToClassifyError(vault_dir / "prose")

    training_records = [
        record for record in vault_records if record[0] not in gold_chunk_ids and record[2]
    ]
    training_records, dropped_classes = _drop_rare_classes(training_records, min_class_count)
    if not training_records:
        raise NoTrainingDataError(axis)

    if train_fn is None:
        train_fn = _default_train_fn
    predict_fn = train_fn(
        [chunk_text for _chunk_id, chunk_text, _label in training_records],
        [label for _chunk_id, _chunk_text, label in training_records],
    )

    gold_texts = [chunk_text for _chunk_id, chunk_text, _label in gold_eval_records]
    gold_labels = [label for _chunk_id, _chunk_text, label in gold_eval_records]
    predictions = predict_fn(gold_texts)

    full_coverage_accuracy = _full_coverage_accuracy(predictions, gold_labels)
    threshold_metrics = [
        _threshold_metrics(predictions, gold_labels, threshold) for threshold in thresholds
    ]
    teacher_gold_agreement = _load_teacher_gold_agreement(eval_report_path, axis)

    gold_predictions = [
        {
            "chunk_id": chunk_id,
            "true": label,
            "predicted": predicted,
            "confidence": confidence,
            "correct": predicted == label,
        }
        for (chunk_id, _chunk_text, label), (predicted, confidence) in zip(
            gold_eval_records, predictions
        )
    ]

    manifest = {
        "axis": axis,
        "corpus_pin_id": pin["corpus_pin_id"],
        "vault_snapshot_hash": pin["vault_snapshot_hash"],
        "config": {
            "max_features": DEFAULT_MAX_FEATURES,
            "ngram_range": list(DEFAULT_NGRAM_RANGE),
            "min_df": DEFAULT_MIN_DF,
            "stop_words": DEFAULT_STOP_WORDS,
            "max_iter": DEFAULT_MAX_ITER,
            "min_class_count": min_class_count,
            "thresholds": list(thresholds),
        },
        "train_chunk_count": len(training_records),
        "dropped_classes": dropped_classes,
        "gold_chunk_count": len(gold_eval_records),
        "full_coverage_accuracy": full_coverage_accuracy,
        "teacher_gold_agreement": teacher_gold_agreement,
        "thresholds": [
            {
                "threshold": metric.threshold,
                "coverage": metric.coverage,
                "covered_count": metric.covered_count,
                "accuracy_on_covered": metric.accuracy_on_covered,
            }
            for metric in threshold_metrics
        ],
        "gold_predictions": gold_predictions,
    }
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(
        json.dumps(manifest, indent=2, sort_keys=True, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    return ClassifyResult(
        manifest_path=manifest_path,
        axis=axis,
        train_chunk_count=len(training_records),
        dropped_classes=dropped_classes,
        gold_chunk_count=len(gold_eval_records),
        full_coverage_accuracy=full_coverage_accuracy,
        teacher_gold_agreement=teacher_gold_agreement,
        thresholds=threshold_metrics,
        corpus_pin_id=pin["corpus_pin_id"],
        vault_snapshot_hash=pin["vault_snapshot_hash"],
    )
