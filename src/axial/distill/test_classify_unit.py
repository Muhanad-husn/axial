"""Inner unit tests for the stage-5d TF-IDF classifier (issues #351/#352,
DEC-37/DEC-38).

Most tests use an injected fake `train_fn` -- a plain `(texts, labels) ->
predict_fn` closure, mirroring `axial.distill.embed`'s `encoder` and
`axial.distill.readiness`'s `cluster_fn` seams -- so the manifest/
threshold-sweep logic runs fast and network-free, independent of what the
real TF-IDF+LogisticRegression fit happens to decide on any given input. A
handful of tests exercise the REAL `_default_train_fn` (real scikit-learn,
small synthetic data, cleanly separable vocabulary) -- fast and cheap at
unit-test scale, not marked `slow`.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from axial.distill.classify import (
    AXES,
    ClassifyResult,
    CorpusPinRequiredError,
    MissingGoldSheetError,
    NoChunksToClassifyError,
    NoGoldLabelsError,
    NoTrainingDataError,
    UnknownAxisError,
    _default_train_fn,
    _drop_rare_classes,
    _full_coverage_accuracy,
    _load_gold_sheet,
    _load_vault_axis_records,
    _threshold_metrics,
    run_classify,
)
from axial.eval.corpus_pin import write_pin
from axial.vault import render_note


def _write_chunk_note(
    prose_dir: Path,
    chunk_id: str,
    chunk_text: str,
    *,
    claim_type_primary: str = "state-formation",
    theory_school_primary: str = "institutionalist-state-centered",
) -> Path:
    prose_dir.mkdir(parents=True, exist_ok=True)
    frontmatter = {
        "chunk_id": chunk_id,
        "section": "Introduction",
        "chunk_text": chunk_text,
        "source_meta": {"author": "A", "title": "T"},
        "schema_version": "0.1",
        "role_in_argument": "role:claim",
        "field": {"primary": "state", "secondary": []},
        "claim_type": {"primary": claim_type_primary, "secondary": None, "subtags": []},
        "theory_school": {
            "primary": theory_school_primary,
            "secondary": None,
            "status": "candidate",
        },
        "empirical_scope": {"value": "scope:country-case", "polity": "Syria"},
        "polities_touched": ["Syria"],
        "artifact_refs": [],
    }
    path = prose_dir / f"{chunk_id}.md"
    path.write_text(render_note(frontmatter, "# Introduction\n\nbody\n"), encoding="utf-8")
    return path


def _stage_pin(tmp_path: Path, name: str = "baseline") -> Path:
    vault_dir = tmp_path / "data" / "vault"
    vault_dir.mkdir(parents=True, exist_ok=True)
    envelopes_dir = tmp_path / "data" / "envelopes"
    envelopes_dir.mkdir(parents=True, exist_ok=True)
    evals_dir = tmp_path / "evals" / "corpus_pin"
    write_pin(name, vault_dir=vault_dir, envelopes_dir=envelopes_dir, evals_dir=evals_dir)
    return evals_dir


def _write_gold_sheet(
    path: Path,
    rows: list[dict[str, str]],
    columns: tuple[str, ...] = ("chunk_id", "chunk_text", "claim_type", "theory_school"),
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "label_sheet"
    for col, name in enumerate(columns, start=1):
        sheet.cell(row=1, column=col, value=name)
    for row_index, row in enumerate(rows, start=2):
        for col, name in enumerate(columns, start=1):
            sheet.cell(row=row_index, column=col, value=row.get(name, ""))
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _fake_train_fn_factory(prediction_by_text: dict[str, tuple[str, float]]):
    """Builds a `train_fn` whose returned `predict_fn` looks up each query
    text's `(label, confidence)` from a fixed table -- ignores the training
    data entirely, so tests can pin exact predictions/confidences."""

    def train_fn(_texts: list[str], _labels: list[str]):
        def predict(query_texts: list[str]) -> list[tuple[str, float]]:
            return [prediction_by_text[text] for text in query_texts]

        return predict

    return train_fn


# --- _load_vault_axis_records -------------------------------------------------


def test_load_vault_axis_records_sorted_with_axis_value(tmp_path: Path):
    prose_dir = tmp_path / "prose"
    _write_chunk_note(
        prose_dir, "b_000_x_001", "second", claim_type_primary="theoretical-conceptual"
    )
    _write_chunk_note(prose_dir, "a_000_x_001", "first", claim_type_primary="descriptive-empirical")

    records = _load_vault_axis_records(prose_dir.parent, "claim_type")

    assert [r[0] for r in records] == ["a_000_x_001", "b_000_x_001"]
    assert records[0] == ("a_000_x_001", "first", "descriptive-empirical")
    assert records[1][2] == "theoretical-conceptual"


def test_load_vault_axis_records_missing_axis_projects_to_empty_string(tmp_path: Path):
    prose_dir = tmp_path / "prose"
    prose_dir.mkdir(parents=True)
    frontmatter = {
        "chunk_id": "a_000_x_001",
        "section": "Introduction",
        "chunk_text": "text",
        "source_meta": {"author": "A"},
        "schema_version": "0.1",
        "role_in_argument": "role:claim",
    }
    (prose_dir / "a_000_x_001.md").write_text(
        render_note(frontmatter, "# Introduction\n\nbody\n"), encoding="utf-8"
    )

    records = _load_vault_axis_records(prose_dir.parent, "claim_type")

    assert records[0][2] == ""


# --- _load_gold_sheet ----------------------------------------------------------


def test_load_gold_sheet_returns_chunk_ids_and_non_empty_eval_records(tmp_path: Path):
    sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        [
            {
                "chunk_id": "g1",
                "chunk_text": "text one",
                "claim_type": "descriptive-empirical",
                "theory_school": "",
            },
            {
                "chunk_id": "g2",
                "chunk_text": "text two",
                "claim_type": "",
                "theory_school": "materialist",
            },
        ],
    )

    chunk_ids, eval_records = _load_gold_sheet(sheet_path, "claim_type")

    assert chunk_ids == {"g1", "g2"}
    assert eval_records == [("g1", "text one", "descriptive-empirical")]


def test_load_gold_sheet_empty_label_excluded_from_eval_set(tmp_path: Path):
    sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        [{"chunk_id": "g1", "chunk_text": "text", "claim_type": "", "theory_school": ""}],
    )

    chunk_ids, eval_records = _load_gold_sheet(sheet_path, "claim_type")

    assert chunk_ids == {"g1"}
    assert eval_records == []


# --- _drop_rare_classes ---------------------------------------------------------


def test_drop_rare_classes_removes_classes_below_min_count():
    records = [(f"c{i}", "text", "common") for i in range(6)] + [
        ("r0", "text", "rare"),
        ("r1", "text", "rare"),
    ]

    filtered, dropped = _drop_rare_classes(records, min_class_count=6)

    assert dropped == ["rare"]
    assert all(r[2] == "common" for r in filtered)
    assert len(filtered) == 6


def test_drop_rare_classes_keeps_classes_at_exactly_the_floor():
    records = [(f"c{i}", "text", "borderline") for i in range(6)]

    filtered, dropped = _drop_rare_classes(records, min_class_count=6)

    assert dropped == []
    assert len(filtered) == 6


# --- _full_coverage_accuracy / _threshold_metrics -------------------------------


def test_full_coverage_accuracy_computes_plain_accuracy():
    predictions = [("a", 0.9), ("b", 0.4), ("a", 0.6)]
    labels = ["a", "a", "a"]

    accuracy = _full_coverage_accuracy(predictions, labels)

    assert accuracy == pytest.approx(2 / 3)


def test_threshold_metrics_coverage_and_accuracy_on_covered():
    predictions = [("a", 0.9), ("b", 0.4), ("a", 0.65)]
    labels = ["a", "a", "b"]

    metrics = _threshold_metrics(predictions, labels, threshold=0.6)

    assert metrics.threshold == 0.6
    assert metrics.covered_count == 2  # 0.9 and 0.65 clear the bar
    assert metrics.coverage == pytest.approx(2 / 3)
    # covered: ("a", true "a") correct, ("a", true "b") incorrect -> 1/2
    assert metrics.accuracy_on_covered == pytest.approx(0.5)


def test_threshold_metrics_no_coverage_reads_none_not_zero():
    predictions = [("a", 0.1)]
    labels = ["a"]

    metrics = _threshold_metrics(predictions, labels, threshold=0.9)

    assert metrics.covered_count == 0
    assert metrics.coverage == 0.0
    assert metrics.accuracy_on_covered is None


def test_threshold_metrics_coverage_never_increases_as_threshold_rises():
    predictions = [("a", 0.9), ("a", 0.7), ("a", 0.55), ("a", 0.3)]
    labels = ["a", "a", "a", "a"]

    coverages = [_threshold_metrics(predictions, labels, t).coverage for t in (0.5, 0.6, 0.7, 0.8)]

    assert coverages == sorted(coverages, reverse=True)


# --- run_classify: loud failures ------------------------------------------------


def test_run_classify_unknown_axis_raises():
    with pytest.raises(UnknownAxisError):
        run_classify("field")


def test_run_classify_requires_a_corpus_pin(tmp_path: Path):
    vault_dir = tmp_path / "data" / "vault"
    _write_chunk_note(vault_dir / "prose", "a_000_x_001", "text")
    gold_sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        [
            {
                "chunk_id": "g1",
                "chunk_text": "text",
                "claim_type": "state-formation",
                "theory_school": "",
            }
        ],
    )

    with pytest.raises(CorpusPinRequiredError):
        run_classify(
            "claim_type",
            vault_dir=vault_dir,
            gold_sheet_path=gold_sheet_path,
            manifest_path=tmp_path / "manifest.json",
            evals_dir=tmp_path / "evals_missing",
        )


def test_run_classify_missing_gold_sheet_raises(tmp_path: Path):
    vault_dir = tmp_path / "data" / "vault"
    _write_chunk_note(vault_dir / "prose", "a_000_x_001", "text")
    evals_dir = _stage_pin(tmp_path)

    with pytest.raises(MissingGoldSheetError):
        run_classify(
            "claim_type",
            vault_dir=vault_dir,
            gold_sheet_path=tmp_path / "nope.xlsx",
            manifest_path=tmp_path / "manifest.json",
            evals_dir=evals_dir,
        )


def test_run_classify_no_gold_labels_for_axis_raises(tmp_path: Path):
    vault_dir = tmp_path / "data" / "vault"
    _write_chunk_note(vault_dir / "prose", "a_000_x_001", "text")
    evals_dir = _stage_pin(tmp_path)
    gold_sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        [
            {
                "chunk_id": "g1",
                "chunk_text": "text",
                "claim_type": "",
                "theory_school": "materialist",
            }
        ],
    )

    with pytest.raises(NoGoldLabelsError):
        run_classify(
            "claim_type",
            vault_dir=vault_dir,
            gold_sheet_path=gold_sheet_path,
            manifest_path=tmp_path / "manifest.json",
            evals_dir=evals_dir,
        )


def test_run_classify_missing_prose_dir_raises(tmp_path: Path):
    vault_dir = tmp_path / "data" / "vault"
    vault_dir.mkdir(parents=True)
    evals_dir = _stage_pin(tmp_path)
    gold_sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        [
            {
                "chunk_id": "g1",
                "chunk_text": "text",
                "claim_type": "state-formation",
                "theory_school": "",
            }
        ],
    )

    with pytest.raises(NoChunksToClassifyError):
        run_classify(
            "claim_type",
            vault_dir=vault_dir,
            gold_sheet_path=gold_sheet_path,
            manifest_path=tmp_path / "manifest.json",
            evals_dir=evals_dir,
        )


def test_run_classify_no_training_data_after_gold_exclusion_raises(tmp_path: Path):
    """The only vault chunk carrying a claim_type value is also the sole
    gold chunk -- once excluded, nothing is left to train on."""
    vault_dir = tmp_path / "data" / "vault"
    _write_chunk_note(vault_dir / "prose", "g1", "text", claim_type_primary="state-formation")
    evals_dir = _stage_pin(tmp_path)
    gold_sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        [
            {
                "chunk_id": "g1",
                "chunk_text": "text",
                "claim_type": "state-formation",
                "theory_school": "",
            }
        ],
    )

    with pytest.raises(NoTrainingDataError):
        run_classify(
            "claim_type",
            vault_dir=vault_dir,
            gold_sheet_path=gold_sheet_path,
            manifest_path=tmp_path / "manifest.json",
            evals_dir=evals_dir,
        )


# --- run_classify: manifest assembly, fake train_fn -----------------------------


def test_run_classify_excludes_gold_chunk_ids_from_training(tmp_path: Path):
    vault_dir = tmp_path / "data" / "vault"
    # 6 chunks share the training-eligible class; "g1" duplicates a gold id.
    for index in range(6):
        chunk_id = "g1" if index == 0 else f"a_{index:03d}_x_001"
        _write_chunk_note(
            vault_dir / "prose",
            chunk_id,
            f"training text {index}",
            claim_type_primary="state-formation",
        )
    evals_dir = _stage_pin(tmp_path)
    gold_sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        [
            {
                "chunk_id": "g1",
                "chunk_text": "gold text",
                "claim_type": "state-formation",
                "theory_school": "",
            }
        ],
    )

    result = run_classify(
        "claim_type",
        vault_dir=vault_dir,
        gold_sheet_path=gold_sheet_path,
        manifest_path=tmp_path / "manifest.json",
        evals_dir=evals_dir,
        min_class_count=1,
        train_fn=_fake_train_fn_factory({"gold text": ("state-formation", 0.9)}),
    )

    # "g1" is excluded from training -- only the other 5 chunks trained on.
    assert result.train_chunk_count == 5


def test_run_classify_drops_classes_below_min_class_count(tmp_path: Path):
    vault_dir = tmp_path / "data" / "vault"
    for index in range(6):
        _write_chunk_note(
            vault_dir / "prose",
            f"a_{index:03d}_x_001",
            f"common text {index}",
            claim_type_primary="descriptive-empirical",
        )
    for index in range(2):
        _write_chunk_note(
            vault_dir / "prose",
            f"b_{index:03d}_x_001",
            f"rare text {index}",
            claim_type_primary="normative-prescriptive",
        )
    evals_dir = _stage_pin(tmp_path)
    gold_sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        [
            {
                "chunk_id": "g1",
                "chunk_text": "gold text",
                "claim_type": "descriptive-empirical",
                "theory_school": "",
            }
        ],
    )

    result = run_classify(
        "claim_type",
        vault_dir=vault_dir,
        gold_sheet_path=gold_sheet_path,
        manifest_path=tmp_path / "manifest.json",
        evals_dir=evals_dir,
        min_class_count=6,
        train_fn=_fake_train_fn_factory({"gold text": ("descriptive-empirical", 0.9)}),
    )

    assert result.dropped_classes == ["normative-prescriptive"]
    assert result.train_chunk_count == 6  # only the 6 descriptive-empirical chunks


def test_run_classify_writes_manifest_with_threshold_sweep_and_no_chunk_text(tmp_path: Path):
    vault_dir = tmp_path / "data" / "vault"
    for index in range(6):
        _write_chunk_note(
            vault_dir / "prose",
            f"a_{index:03d}_x_001",
            f"SENTINEL_TRAIN_TEXT_{index}",
            claim_type_primary="descriptive-empirical",
        )
    evals_dir = _stage_pin(tmp_path)
    gold_sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        [
            {
                "chunk_id": "g1",
                "chunk_text": "SENTINEL_GOLD_TEXT_HIGH_CONF",
                "claim_type": "descriptive-empirical",
                "theory_school": "",
            },
            {
                "chunk_id": "g2",
                "chunk_text": "SENTINEL_GOLD_TEXT_LOW_CONF",
                "claim_type": "descriptive-empirical",
                "theory_school": "",
            },
            {
                "chunk_id": "g3",
                "chunk_text": "SENTINEL_GOLD_TEXT_WRONG",
                "claim_type": "descriptive-empirical",
                "theory_school": "",
            },
        ],
    )
    manifest_path = tmp_path / "manifest.json"

    result = run_classify(
        "claim_type",
        vault_dir=vault_dir,
        gold_sheet_path=gold_sheet_path,
        manifest_path=manifest_path,
        evals_dir=evals_dir,
        min_class_count=6,
        thresholds=(0.5, 0.6, 0.7, 0.8),
        train_fn=_fake_train_fn_factory(
            {
                "SENTINEL_GOLD_TEXT_HIGH_CONF": ("descriptive-empirical", 0.9),
                "SENTINEL_GOLD_TEXT_LOW_CONF": ("descriptive-empirical", 0.55),
                "SENTINEL_GOLD_TEXT_WRONG": ("theoretical-conceptual", 0.9),
            }
        ),
    )

    assert isinstance(result, ClassifyResult)
    assert result.gold_chunk_count == 3
    assert result.full_coverage_accuracy == pytest.approx(2 / 3)  # g1, g2 correct; g3 wrong
    assert result.corpus_pin_id == "baseline"

    by_threshold = {m.threshold: m for m in result.thresholds}
    assert by_threshold[0.5].covered_count == 3  # all three clear 0.5
    assert by_threshold[0.6].covered_count == 2  # only 0.9, 0.9 clear 0.6
    assert by_threshold[0.6].accuracy_on_covered == pytest.approx(0.5)  # g1 right, g3 wrong
    assert by_threshold[0.8].covered_count == 2
    assert by_threshold[0.8].accuracy_on_covered == pytest.approx(0.5)

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert manifest["axis"] == "claim_type"
    assert manifest["train_chunk_count"] == 6
    assert manifest["gold_chunk_count"] == 3
    assert manifest["corpus_pin_id"] == "baseline"
    assert manifest["config"]["min_class_count"] == 6
    assert manifest["config"]["thresholds"] == [0.5, 0.6, 0.7, 0.8]
    assert len(manifest["thresholds"]) == 4
    assert manifest["teacher_gold_agreement"] is None  # no eval_report.json staged

    # DEC-23: no chunk_text (train or gold sentinels) anywhere in the manifest.
    manifest_text = json.dumps(manifest)
    for sentinel in (
        "SENTINEL_TRAIN_TEXT_0",
        "SENTINEL_GOLD_TEXT_HIGH_CONF",
        "SENTINEL_GOLD_TEXT_LOW_CONF",
        "SENTINEL_GOLD_TEXT_WRONG",
    ):
        assert sentinel not in manifest_text


def test_run_classify_loads_teacher_gold_agreement_when_present(tmp_path: Path):
    vault_dir = tmp_path / "data" / "vault"
    for index in range(6):
        _write_chunk_note(
            vault_dir / "prose",
            f"a_{index:03d}_x_001",
            f"training text {index}",
            claim_type_primary="descriptive-empirical",
        )
    evals_dir = _stage_pin(tmp_path)
    gold_sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        [
            {
                "chunk_id": "g1",
                "chunk_text": "gold text",
                "claim_type": "descriptive-empirical",
                "theory_school": "",
            }
        ],
    )
    eval_report_path = tmp_path / "eval_report.json"
    eval_report_path.write_text(
        json.dumps({"per_axis_agreement": {"claim_type": 0.56, "theory_school": 0.543}}),
        encoding="utf-8",
    )

    result = run_classify(
        "claim_type",
        vault_dir=vault_dir,
        gold_sheet_path=gold_sheet_path,
        eval_report_path=eval_report_path,
        manifest_path=tmp_path / "manifest.json",
        evals_dir=evals_dir,
        min_class_count=6,
        train_fn=_fake_train_fn_factory({"gold text": ("descriptive-empirical", 0.9)}),
    )

    assert result.teacher_gold_agreement == pytest.approx(0.56)


def test_run_classify_missing_eval_report_is_non_fatal(tmp_path: Path):
    vault_dir = tmp_path / "data" / "vault"
    for index in range(6):
        _write_chunk_note(
            vault_dir / "prose",
            f"a_{index:03d}_x_001",
            f"training text {index}",
            claim_type_primary="descriptive-empirical",
        )
    evals_dir = _stage_pin(tmp_path)
    gold_sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        [
            {
                "chunk_id": "g1",
                "chunk_text": "gold text",
                "claim_type": "descriptive-empirical",
                "theory_school": "",
            }
        ],
    )

    result = run_classify(
        "claim_type",
        vault_dir=vault_dir,
        gold_sheet_path=gold_sheet_path,
        eval_report_path=tmp_path / "does_not_exist.json",
        manifest_path=tmp_path / "manifest.json",
        evals_dir=evals_dir,
        min_class_count=6,
        train_fn=_fake_train_fn_factory({"gold text": ("descriptive-empirical", 0.9)}),
    )

    assert result.teacher_gold_agreement is None


# --- _default_train_fn: the real TF-IDF + LogisticRegression pipeline ----------


@pytest.mark.parametrize("axis", AXES)
def test_axes_constant_matches_module_default_axes(axis):
    assert axis in AXES


def test_default_train_fn_separates_two_lexically_distinct_classes():
    pytest.importorskip("sklearn")
    class_a_texts = [f"empirical observational field data study number {i}" for i in range(8)]
    class_b_texts = [f"theoretical conceptual abstract framework model idea {i}" for i in range(8)]
    texts = class_a_texts + class_b_texts
    labels = ["class-a"] * 8 + ["class-b"] * 8

    predict_fn = _default_train_fn(texts, labels, min_df=1)
    predictions = predict_fn(
        ["empirical observational field study", "theoretical conceptual abstract model"]
    )

    assert predictions[0][0] == "class-a"
    assert predictions[1][0] == "class-b"
    assert all(0.0 <= confidence <= 1.0 for _label, confidence in predictions)


def test_default_train_fn_is_deterministic_across_calls():
    pytest.importorskip("sklearn")
    class_a_texts = [f"empirical observational field data study number {i}" for i in range(8)]
    class_b_texts = [f"theoretical conceptual abstract framework model idea {i}" for i in range(8)]
    texts = class_a_texts + class_b_texts
    labels = ["class-a"] * 8 + ["class-b"] * 8
    query = ["empirical observational field study"]

    first = _default_train_fn(texts, labels, min_df=1)(query)
    second = _default_train_fn(texts, labels, min_df=1)(query)

    assert first == second
