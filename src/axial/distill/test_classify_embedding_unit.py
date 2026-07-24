"""Inner unit tests for the stage-5d dense-embedding classifier (issue #350,
DEC-39).

Most tests use an injected fake `train_fn` -- a plain `(vectors, labels) ->
predict_fn` closure, mirroring `axial.distill.classify`'s own `train_fn` seam
-- so the manifest/threshold-sweep logic runs fast and sklearn-free,
independent of what a real `LogisticRegression` fit happens to decide. A
handful of tests exercise the REAL `_default_train_fn` (real scikit-learn,
tiny cleanly-separable synthetic vectors) -- fast and cheap at unit-test
scale, not marked `slow`. The embeddings store itself is always real (a tiny
LanceDB table written directly, `pytest.importorskip("lancedb")` at module
level) -- this module never mocks its own read path.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

lancedb = pytest.importorskip("lancedb")

from axial.distill.classify_embedding import (  # noqa: E402
    AXES,
    ClassifyResult,
    CorpusPinRequiredError,
    MissingGoldEmbeddingError,
    MissingGoldSheetError,
    NoEmbeddingsToClassifyError,
    NoGoldLabelsError,
    NoTrainingDataError,
    UnknownAxisError,
    _drop_rare_classes,
    _load_gold_labels,
    run_classify_embedding,
)
from axial.eval.corpus_pin import write_pin  # noqa: E402


def _write_embeddings(embeddings_dir: Path, rows: list[dict]) -> None:
    db = lancedb.connect(embeddings_dir)
    db.create_table("chunks", data=rows, mode="overwrite")


def _row(chunk_id: str, vector: list[float], field_primary: str) -> dict:
    return {"chunk_id": chunk_id, "vector": vector, "field_primary": field_primary}


def _stage_pin(tmp_path: Path, name: str = "baseline") -> Path:
    vault_dir = tmp_path / "data" / "vault"
    vault_dir.mkdir(parents=True, exist_ok=True)
    envelopes_dir = tmp_path / "data" / "envelopes"
    envelopes_dir.mkdir(parents=True, exist_ok=True)
    evals_dir = tmp_path / "evals" / "corpus_pin"
    write_pin(name, vault_dir=vault_dir, envelopes_dir=envelopes_dir, evals_dir=evals_dir)
    return evals_dir


def _write_gold_sheet(path: Path, columns: tuple[str, ...], rows: list[dict[str, str]]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "label_sheet"
    for col, name in enumerate(columns, start=1):
        sheet.cell(row=1, column=col, value=name)
    for row_index, row in enumerate(rows, start=2):
        for col, name in enumerate(columns, start=1):
            sheet.cell(row=row_index, column=col, value=row.get(name, ""))
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _fake_train_fn_factory(prediction_by_id: dict[str, tuple[str, float]], vectors_to_ids: dict):
    """Builds a `train_fn` whose returned `predict_fn` looks up each query
    vector's `(label, confidence)` by identity in `vectors_to_ids` -- ignores
    the training data entirely, so tests can pin exact predictions."""

    def train_fn(_vectors: list[list[float]], _labels: list[str]):
        def predict(query_vectors: list[list[float]]) -> list[tuple[str, float]]:
            return [prediction_by_id[vectors_to_ids[tuple(v)]] for v in query_vectors]

        return predict

    return train_fn


# --- AXES ------------------------------------------------------------------


def test_axes_is_field_only():
    assert AXES == ("field",)


# --- _load_gold_labels -------------------------------------------------------


def test_load_gold_labels_reads_gold_column_not_bare_axis_column(tmp_path: Path):
    """DEC-39's own wrinkle: the eval set must come from `field_gold`, and
    the sheet's own `field` column (tagger pre-fill) must never leak into
    `eval_records` as if it were the independent judgment."""
    sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        ("chunk_id", "field", "field_gold"),
        [
            {"chunk_id": "g1", "field": "state", "field_gold": "violence"},
            {"chunk_id": "g2", "field": "state", "field_gold": ""},
        ],
    )

    chunk_ids, eval_records, teacher_gold_agreement = _load_gold_labels(sheet_path, "field")

    assert chunk_ids == {"g1", "g2"}  # g2 stays in the training-exclusion set
    assert eval_records == [("g1", "violence")]  # only g1 carries a field_gold label
    assert teacher_gold_agreement == pytest.approx(0.0)  # field="state" vs field_gold="violence"


def test_load_gold_labels_teacher_agreement_averages_over_gold_labeled_rows(tmp_path: Path):
    sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        ("chunk_id", "field", "field_gold"),
        [
            {"chunk_id": "g1", "field": "state", "field_gold": "state"},  # agree
            {"chunk_id": "g2", "field": "violence", "field_gold": "state"},  # disagree
            {"chunk_id": "g3", "field": "ideology", "field_gold": "ideology"},  # agree
        ],
    )

    _chunk_ids, eval_records, teacher_gold_agreement = _load_gold_labels(sheet_path, "field")

    assert len(eval_records) == 3
    assert teacher_gold_agreement == pytest.approx(2 / 3)


def test_load_gold_labels_no_tagger_prefill_column_gives_none_agreement(tmp_path: Path):
    """No `field` column at all (a hand-built sheet carrying only the gold
    judgment) -- `teacher_gold_agreement` is `None`, never a bogus 0.0."""
    sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        ("chunk_id", "field_gold"),
        [{"chunk_id": "g1", "field_gold": "state"}],
    )

    _chunk_ids, eval_records, teacher_gold_agreement = _load_gold_labels(sheet_path, "field")

    assert eval_records == [("g1", "state")]
    assert teacher_gold_agreement is None


def test_load_gold_labels_missing_gold_column_returns_empty(tmp_path: Path):
    sheet_path = _write_gold_sheet(
        tmp_path / "label_sheet.xlsx",
        ("chunk_id", "field"),
        [{"chunk_id": "g1", "field": "state"}],
    )

    chunk_ids, eval_records, teacher_gold_agreement = _load_gold_labels(sheet_path, "field")

    assert chunk_ids == set()
    assert eval_records == []
    assert teacher_gold_agreement is None


# --- _drop_rare_classes -------------------------------------------------------


def test_drop_rare_classes_removes_classes_below_min_count():
    records = [(f"c{i}", [0.0], "common") for i in range(6)] + [
        ("r0", [0.0], "rare"),
        ("r1", [0.0], "rare"),
    ]

    filtered, dropped = _drop_rare_classes(records, min_class_count=6)

    assert dropped == ["rare"]
    assert all(r[2] == "common" for r in filtered)
    assert len(filtered) == 6


def test_drop_rare_classes_keeps_classes_at_exactly_the_floor():
    records = [(f"c{i}", [0.0], "borderline") for i in range(6)]

    filtered, dropped = _drop_rare_classes(records, min_class_count=6)

    assert dropped == []
    assert len(filtered) == 6


# --- run_classify_embedding: error paths --------------------------------------


def test_run_classify_embedding_unknown_axis_raises(tmp_path: Path):
    with pytest.raises(UnknownAxisError):
        run_classify_embedding("not-a-real-axis", embeddings_dir=tmp_path / "emb.lance")


def test_run_classify_embedding_requires_corpus_pin(tmp_path: Path):
    with pytest.raises(CorpusPinRequiredError):
        run_classify_embedding(
            "field",
            embeddings_dir=tmp_path / "emb.lance",
            gold_sheet_path=tmp_path / "gold.xlsx",
            evals_dir=tmp_path / "evals" / "corpus_pin",
        )


def test_run_classify_embedding_missing_gold_sheet_raises(tmp_path: Path):
    evals_dir = _stage_pin(tmp_path)

    with pytest.raises(MissingGoldSheetError):
        run_classify_embedding(
            "field",
            embeddings_dir=tmp_path / "emb.lance",
            gold_sheet_path=tmp_path / "does-not-exist.xlsx",
            evals_dir=evals_dir,
        )


def test_run_classify_embedding_no_gold_labels_raises(tmp_path: Path):
    evals_dir = _stage_pin(tmp_path)
    gold_sheet_path = _write_gold_sheet(
        tmp_path / "gold.xlsx",
        ("chunk_id", "field", "field_gold"),
        [{"chunk_id": "g1", "field": "state", "field_gold": ""}],
    )

    with pytest.raises(NoGoldLabelsError):
        run_classify_embedding(
            "field",
            embeddings_dir=tmp_path / "emb.lance",
            gold_sheet_path=gold_sheet_path,
            evals_dir=evals_dir,
        )


def test_run_classify_embedding_missing_embeddings_store_raises(tmp_path: Path):
    evals_dir = _stage_pin(tmp_path)
    gold_sheet_path = _write_gold_sheet(
        tmp_path / "gold.xlsx",
        ("chunk_id", "field", "field_gold"),
        [{"chunk_id": "g1", "field": "state", "field_gold": "state"}],
    )

    with pytest.raises(NoEmbeddingsToClassifyError):
        run_classify_embedding(
            "field",
            embeddings_dir=tmp_path / "does-not-exist.lance",
            gold_sheet_path=gold_sheet_path,
            evals_dir=evals_dir,
        )


def test_run_classify_embedding_missing_gold_embedding_raises(tmp_path: Path):
    evals_dir = _stage_pin(tmp_path)
    embeddings_dir = tmp_path / "emb.lance"
    _write_embeddings(
        embeddings_dir,
        [_row("c1", [1.0, 0.0], "state"), _row("c2", [0.0, 1.0], "state")],
    )
    gold_sheet_path = _write_gold_sheet(
        tmp_path / "gold.xlsx",
        ("chunk_id", "field", "field_gold"),
        # "ghost" carries a gold label but was never embedded.
        [{"chunk_id": "ghost", "field": "state", "field_gold": "state"}],
    )

    with pytest.raises(MissingGoldEmbeddingError) as excinfo:
        run_classify_embedding(
            "field",
            embeddings_dir=embeddings_dir,
            gold_sheet_path=gold_sheet_path,
            evals_dir=evals_dir,
        )
    assert excinfo.value.missing_chunk_ids == ["ghost"]


# --- run_classify_embedding: happy path ---------------------------------------


def test_run_classify_embedding_excludes_gold_drops_rare_class_writes_manifest(tmp_path: Path):
    evals_dir = _stage_pin(tmp_path)
    embeddings_dir = tmp_path / "emb.lance"

    # 7 "state" chunks (one, s0, is also a gold row -- leakage exclusion) and
    # 2 "ideology" chunks (below the min-class-count floor of 6, dropped).
    rows = [_row(f"s{i}", [float(i), 0.0], "state") for i in range(7)]
    rows += [_row(f"i{i}", [0.0, float(i)], "ideology") for i in range(2)]
    rows.append(_row("gold_only", [9.0, 9.0], "state"))  # gold-labeled, excluded from training
    _write_embeddings(embeddings_dir, rows)

    gold_sheet_path = _write_gold_sheet(
        tmp_path / "gold.xlsx",
        ("chunk_id", "field", "field_gold"),
        [
            {"chunk_id": "s0", "field": "state", "field_gold": "state"},
            {"chunk_id": "gold_only", "field": "violence", "field_gold": "state"},  # prefill wrong
        ],
    )

    vectors_to_ids = {(0.0, 0.0): "s0", (9.0, 9.0): "gold_only"}
    train_fn = _fake_train_fn_factory(
        {"s0": ("state", 0.95), "gold_only": ("state", 0.55)}, vectors_to_ids
    )

    result = run_classify_embedding(
        "field",
        embeddings_dir=embeddings_dir,
        gold_sheet_path=gold_sheet_path,
        manifest_path=tmp_path / "manifest.json",
        evals_dir=evals_dir,
        min_class_count=6,
        thresholds=(0.5, 0.6, 0.7, 0.8),
        train_fn=train_fn,
    )

    assert isinstance(result, ClassifyResult)
    assert result.axis == "field"
    assert result.dropped_classes == ["ideology"]
    assert result.train_chunk_count == 6  # 7 state minus the gold-excluded s0
    assert result.gold_chunk_count == 2
    assert result.full_coverage_accuracy == pytest.approx(1.0)  # both predictions correct
    # teacher pre-fill: s0 agrees (state=state), gold_only disagrees (violence!=state) -> 0.5
    assert result.teacher_gold_agreement == pytest.approx(0.5)
    assert result.corpus_pin_id == "baseline"

    by_threshold = {m.threshold: m for m in result.thresholds}
    assert by_threshold[0.5].covered_count == 2
    assert by_threshold[0.6].covered_count == 1  # only s0's 0.95 clears 0.6
    assert by_threshold[0.6].accuracy_on_covered == pytest.approx(1.0)

    import json

    manifest = json.loads((tmp_path / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["axis"] == "field"
    assert manifest["config"]["technique"] == "logistic_regression_on_dense_embeddings"
    assert manifest["teacher_gold_agreement"] == pytest.approx(0.5)
    manifest_text = json.dumps(manifest)
    assert "chunk_text" not in manifest_text


def test_run_classify_embedding_no_training_data_raises_when_only_rare_classes(tmp_path: Path):
    evals_dir = _stage_pin(tmp_path)
    embeddings_dir = tmp_path / "emb.lance"
    _write_embeddings(embeddings_dir, [_row("s0", [1.0], "state"), _row("s1", [2.0], "state")])
    gold_sheet_path = _write_gold_sheet(
        tmp_path / "gold.xlsx",
        ("chunk_id", "field", "field_gold"),
        [{"chunk_id": "s0", "field": "state", "field_gold": "state"}],
    )

    with pytest.raises(NoTrainingDataError):
        run_classify_embedding(
            "field",
            embeddings_dir=embeddings_dir,
            gold_sheet_path=gold_sheet_path,
            evals_dir=evals_dir,
            min_class_count=6,  # only 1 non-gold "state" example remains (s1) -- below floor
        )


# --- run_classify_embedding: real LogisticRegression fit ----------------------


def test_run_classify_embedding_real_train_fn_separates_clean_clusters(tmp_path: Path):
    pytest.importorskip("sklearn")
    evals_dir = _stage_pin(tmp_path)
    embeddings_dir = tmp_path / "emb.lance"

    rows = [_row(f"s{i}", [10.0 + i * 0.01, 0.0], "state") for i in range(8)]
    rows += [_row(f"v{i}", [-10.0 - i * 0.01, 0.0], "violence") for i in range(8)]
    rows.append(_row("gs", [10.0, 0.0], "state"))  # gold row, held out
    rows.append(_row("gv", [-10.0, 0.0], "violence"))  # gold row, held out
    _write_embeddings(embeddings_dir, rows)

    gold_sheet_path = _write_gold_sheet(
        tmp_path / "gold.xlsx",
        ("chunk_id", "field", "field_gold"),
        [
            {"chunk_id": "gs", "field": "state", "field_gold": "state"},
            {"chunk_id": "gv", "field": "violence", "field_gold": "violence"},
        ],
    )

    result = run_classify_embedding(
        "field",
        embeddings_dir=embeddings_dir,
        gold_sheet_path=gold_sheet_path,
        manifest_path=tmp_path / "manifest.json",
        evals_dir=evals_dir,
    )

    assert result.full_coverage_accuracy == pytest.approx(1.0)
    assert result.teacher_gold_agreement == pytest.approx(1.0)
