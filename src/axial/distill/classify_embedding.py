"""Stage-5d dense-embedding classifier for `field_primary` (issue #350,
DEC-39, plan `plans/phase-a-completion/README.md` stage 5d).

DEC-37/38 found TF-IDF beats dense sentence embeddings for the two *blind*
axes (`claim_type`/`theory_school`, `axial.distill.classify`). DEC-39 found
the opposite for `field_primary` (this module) and, on a weaker margin,
`role_in_argument`: a plain multinomial `LogisticRegression` trained
directly on the dense vectors 5a already persisted (`axial.distill.embed`'s
LanceDB store) is the measured-best technique. This module never re-embeds
text -- it reads vectors straight out of that store, keyed by `chunk_id`,
the same reuse `axial.distill.readiness` already established for its own
clustering input. `sentence-transformers` is never imported here; if the
store is missing, this fails loudly (`NoEmbeddingsToClassifyError`) rather
than silently re-embedding (a different, uncontrolled input to what DEC-39
actually measured).

Only `field` is validated (DEC-39's gold check). `role_in_argument` (#348)
and `empirical_scope` (#349, which DEC-39 found does NOT clear the bar) are
each their own graduation call -- `AXIS_METADATA_COLUMNS`/`AXIS_GOLD_COLUMNS`
below are a two-entry-ready dict specifically so a future axis on this same
measured technique adds one entry, not a copy-pasted module (mirroring
`axial.distill.classify`'s own "a future axis... gets its own small module"
precedent, one level down: same technique, different axis, stays one
module; a genuinely different technique still gets its own module).

Gold-column wrinkle (DEC-37's citation, DEC-39's fix) -- why this module
does NOT reuse `axial.distill.classify`'s gold-reading convention
-----------------------------------------------------------------------
For `claim_type`/`theory_school`, the gold sheet's own `claim_type`/
`theory_school` columns ARE the independent gold judgment (DEC-29/30's
blind coding method from the start). For `field` (and `role_in_argument`/
`empirical_scope`), the ORIGINAL gold sheet's `field` column was a rubber-
stamped copy of the tagger's own pre-filled production tag, never
independently judged -- trivially "1.0 agreement" against itself (DEC-37).
DEC-39 re-labeled these three axes blind (no pre-fill shown) into new
`field_gold`/`role_in_argument_gold`/`empirical_scope_gold` columns. This
module reads `{axis}_gold` for the independent judgment, and computes the
teacher's own gold agreement FRESH, from the sheet's own `{axis}` (tagger
pre-fill) column against `{axis}_gold` -- NOT from
`data/gold/labels/eval_report.json`'s `per_axis_agreement`, which for this
axis is still the old rubber-stamped 1.0 (verified directly against the
real gold sheet while building this module: comparing `field` to
`field_gold` over the real 120-row sheet reproduces DEC-39's cited 76.7%
exactly; `eval_report.json["per_axis_agreement"]["field"]` reads 1.0, the
stale number). Reusing that report here would silently misreport the one
comparison number this module's whole acceptance bar hinges on.

`scikit-learn`/`lancedb` (the `distill` dependency group) are imported
lazily, inside `_default_train_fn`/`_load_embedding_lookup`, never at
module level -- mirrors `axial.distill.classify`'s own lazy-sklearn
precedent, so importing this module (e.g. from `axial.cli`) never requires
either package.

DEC-23: this module never reads `chunk_text` at all (not even transiently
-- unlike `axial.distill.classify`, which reads it to feed TF-IDF, this
module's whole input is vectors). The manifest carries chunk_ids, tag
values, counts, and confidence scores only.
"""

from __future__ import annotations

import json
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from axial.distill.classify import (
    DEFAULT_GOLD_SHEET_PATH,
    DEFAULT_MIN_CLASS_COUNT,
    DEFAULT_THRESHOLDS,
    ThresholdMetrics,
    _full_coverage_accuracy,
    _manifest_path_for,
    _normalize_cell,
    _threshold_metrics,
)
from axial.distill.embed import DEFAULT_EMBEDDINGS_DIR, TABLE_NAME
from axial.distill.staleness import resolve_current_pin
from axial.eval import corpus_pin as _corpus_pin

# axis -> its flattened metadata column in the 5a embeddings store
# (`axial.distill.embed._flatten_metadata`), and axis -> the gold sheet's
# independent-judgment column (DEC-39's `_gold` suffix -- see module
# docstring). Only `field` is validated (DEC-39); adding a second axis on
# this same technique is a one-line addition to each dict, not a new module.
AXIS_METADATA_COLUMNS = {"field": "field_primary"}
AXIS_GOLD_COLUMNS = {axis: f"{axis}_gold" for axis in AXIS_METADATA_COLUMNS}
AXES = tuple(AXIS_METADATA_COLUMNS)

# sklearn's own convergence ceiling for LogisticRegression on this axis's
# ~14.6k x 384 training matrix -- matches `axial.distill.classify`'s same
# constant/precedent, not a re-tuned value.
DEFAULT_MAX_ITER = 2000

# One (predicted_label, confidence) pair per input vector, in input order.
PredictFn = Callable[[list[list[float]]], list[tuple[str, float]]]
# Fits on (vectors, labels) and returns the fitted PredictFn -- the seam
# this module's own inner unit tests use, mirroring
# `axial.distill.classify`'s own `train_fn` injection seam.
TrainFn = Callable[[list[list[float]], list[str]], PredictFn]


class ClassifyEmbeddingError(Exception):
    """Base class for all stage-5d embedding-classifier errors."""


class UnknownAxisError(ClassifyEmbeddingError):
    """Raised when `axis` is not one of `AXES` -- this module implements
    the dense-embedding technique DEC-39 validated for `field` only (see
    module docstring)."""

    def __init__(self, axis: str):
        self.axis = axis
        super().__init__(f"unknown axis {axis!r}; expected one of {AXES!r}")


class NoEmbeddingsToClassifyError(ClassifyEmbeddingError):
    """Raised when `embeddings_dir` holds no persisted vector table -- this
    pass depends on 5a; mirrors `axial.distill.readiness`'s own
    `NoEmbeddingsToClusterError`."""

    def __init__(self, embeddings_dir: Path):
        self.embeddings_dir = embeddings_dir
        super().__init__(
            f"no persisted embeddings found at {embeddings_dir}; run `axial distill embed` first"
        )


class CorpusPinRequiredError(ClassifyEmbeddingError):
    """Raised when no corpus-pin manifest can be resolved (DEC-35: every
    stage-5 artifact records the pin it was built from)."""

    def __init__(self, cause: _corpus_pin.CorpusPinError):
        self.cause = cause
        super().__init__(
            f"classify pass requires a corpus pin to record provenance against "
            f"({cause}); run `axial pin write <name>` first"
        )


class MissingGoldSheetError(ClassifyEmbeddingError):
    """Raised when no returned gold answer-key sheet exists."""

    def __init__(self, gold_sheet_path: Path):
        self.gold_sheet_path = gold_sheet_path
        super().__init__(
            f"no gold answer-key sheet found at {gold_sheet_path}; place the "
            f"returned label_sheet.xlsx under data/gold/labels/ (see `axial eval`) "
            f"before running `axial distill classify`"
        )


class NoGoldLabelsError(ClassifyEmbeddingError):
    """Raised when the gold sheet carries zero non-empty independent
    judgments for `axis` (the `{axis}_gold` column -- see module docstring)
    -- there is nothing to evaluate against."""

    def __init__(self, axis: str, gold_sheet_path: Path):
        self.axis = axis
        self.gold_sheet_path = gold_sheet_path
        super().__init__(
            f"no gold labels for axis {axis!r} (column {AXIS_GOLD_COLUMNS.get(axis)!r}) "
            f"found in {gold_sheet_path}"
        )


class NoTrainingDataError(ClassifyEmbeddingError):
    """Raised when, after excluding gold chunk_ids and dropping classes
    below `min_class_count`, zero training examples remain for `axis`."""

    def __init__(self, axis: str):
        self.axis = axis
        super().__init__(
            f"no training examples remain for axis {axis!r} after excluding gold "
            f"chunks and classes below the min-class-count floor"
        )


class MissingGoldEmbeddingError(ClassifyEmbeddingError):
    """Raised when a gold-labeled chunk_id has no persisted vector in the
    embeddings store -- a stale store relative to the gold sheet (the vault
    changed since the last `axial distill embed`), not a normal skip: this
    module never silently drops a gold row it cannot vectorize."""

    def __init__(self, axis: str, embeddings_dir: Path, missing_chunk_ids: list[str]):
        self.axis = axis
        self.embeddings_dir = embeddings_dir
        self.missing_chunk_ids = missing_chunk_ids
        super().__init__(
            f"{len(missing_chunk_ids)} gold chunk(s) for axis {axis!r} have no "
            f"persisted vector in {embeddings_dir}; re-run `axial distill embed` "
            f"against the current vault (first missing: {missing_chunk_ids[:5]!r})"
        )


@dataclass(frozen=True)
class ClassifyResult:
    """The outcome of one `run_classify_embedding` call."""

    manifest_path: Path
    axis: str
    train_chunk_count: int
    dropped_classes: list[str]
    gold_chunk_count: int
    full_coverage_accuracy: float
    teacher_gold_agreement: float | None
    thresholds: list[ThresholdMetrics]
    corpus_pin_id: str
    vault_snapshot_hash: str


def _load_embedding_lookup(
    embeddings_dir: Path, axis_column: str
) -> dict[str, tuple[list[float], str]]:
    """Every persisted row's `chunk_id -> (vector, axis_value)`, read once
    from the 5a LanceDB store -- never `chunk_text` (DEC-23; the store never
    persisted it in the first place)."""
    embeddings_dir = Path(embeddings_dir)
    if not embeddings_dir.exists():
        raise NoEmbeddingsToClassifyError(embeddings_dir)

    import lancedb

    db = lancedb.connect(embeddings_dir)
    if TABLE_NAME not in db.list_tables().tables:
        raise NoEmbeddingsToClassifyError(embeddings_dir)
    rows = db.open_table(TABLE_NAME).to_arrow().to_pylist()
    if not rows:
        raise NoEmbeddingsToClassifyError(embeddings_dir)
    return {row["chunk_id"]: (row["vector"], row.get(axis_column) or "") for row in rows}


def _load_gold_labels(
    gold_sheet_path: Path, axis: str
) -> tuple[set[str], list[tuple[str, str]], float | None]:
    """Read the returned gold answer-key sheet: every `chunk_id` present
    (the training-exclusion set, regardless of whether this axis was
    independently gold-judged on that row), the `(chunk_id, gold_label)`
    pairs carrying a non-empty `{axis}_gold` independent judgment (the
    evaluation set -- NOT the bare `axis` column; see module docstring), and
    the teacher's own gold agreement computed fresh over those same rows
    from the sheet's own `axis` (tagger pre-fill) column against
    `{axis}_gold` -- `None` when that pre-fill column is absent, never
    silently substituted from `eval_report.json` (see module docstring)."""
    worksheet = load_workbook(gold_sheet_path, read_only=True).worksheets[0]
    rows_iter = worksheet.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        return set(), [], None
    column_index = {str(name).strip(): index for index, name in enumerate(header) if name}
    gold_column = AXIS_GOLD_COLUMNS[axis]
    if "chunk_id" not in column_index or gold_column not in column_index:
        return set(), [], None
    tagger_column_index = column_index.get(axis)

    chunk_ids: set[str] = set()
    eval_records: list[tuple[str, str]] = []
    teacher_correct = 0
    teacher_total = 0
    for row in rows_iter:
        chunk_id = _normalize_cell(row[column_index["chunk_id"]])
        if not chunk_id:
            continue
        chunk_ids.add(chunk_id)
        gold_label = _normalize_cell(row[column_index[gold_column]])
        if not gold_label:
            continue
        eval_records.append((chunk_id, gold_label))
        if tagger_column_index is not None:
            teacher_total += 1
            if _normalize_cell(row[tagger_column_index]) == gold_label:
                teacher_correct += 1

    eval_records.sort(key=lambda record: record[0])
    teacher_gold_agreement = teacher_correct / teacher_total if teacher_total else None
    return chunk_ids, eval_records, teacher_gold_agreement


def _drop_rare_classes(
    records: list[tuple[str, list[float], str]], min_class_count: int
) -> tuple[list[tuple[str, list[float], str]], list[str]]:
    """Drop every record whose label occurs fewer than `min_class_count`
    times -- too few examples to learn a stable boundary, same floor
    `axial.distill.classify` uses."""
    counts = Counter(label for _chunk_id, _vector, label in records)
    dropped = sorted(label for label, count in counts.items() if count < min_class_count)
    dropped_set = set(dropped)
    filtered = [record for record in records if record[2] not in dropped_set]
    return filtered, dropped


def _default_train_fn(
    vectors: list[list[float]],
    labels: list[str],
    *,
    max_iter: int = DEFAULT_MAX_ITER,
) -> PredictFn:
    """Fit a plain multinomial `LogisticRegression` directly on the 5a
    vectors (DEC-39's measured technique; lazy `sklearn` import -- see
    module docstring) and return a closure predicting `(label, confidence)`
    pairs -- the top predicted class and its predicted probability."""
    import numpy as np
    from sklearn.linear_model import LogisticRegression

    classifier = LogisticRegression(max_iter=max_iter)
    classifier.fit(np.asarray(vectors, dtype=np.float64), labels)

    def predict(query_vectors: list[list[float]]) -> list[tuple[str, float]]:
        probabilities = classifier.predict_proba(np.asarray(query_vectors, dtype=np.float64))
        classes = classifier.classes_
        results = []
        for row in probabilities:
            best_index = row.argmax()
            results.append((str(classes[best_index]), float(row[best_index])))
        return results

    return predict


def run_classify_embedding(
    axis: str,
    embeddings_dir: Path | None = None,
    gold_sheet_path: Path | None = None,
    manifest_path: Path | None = None,
    evals_dir: Path | None = None,
    min_class_count: int = DEFAULT_MIN_CLASS_COUNT,
    thresholds: tuple[float, ...] = DEFAULT_THRESHOLDS,
    train_fn: TrainFn | None = None,
) -> ClassifyResult:
    """Train a `LogisticRegression` classifier for `axis` (one of `AXES`) on
    the 5a embeddings store's own vectors (excluding gold chunks and classes
    below `min_class_count`), evaluate it against the independent
    `{axis}_gold` judgments at full coverage and at each of `thresholds`,
    and write a JSON manifest to `manifest_path` (default
    `_manifest_path_for(axis)`, the same `data/distill/classify_<axis>_
    manifest.json` convention `axial.distill.classify` uses -- the two
    modules' `AXES` are disjoint, so there is no collision). This pass is a
    measurement artifact only -- it never writes to the vault.

    `train_fn`, when given, replaces the default `LogisticRegression` fit
    (`_default_train_fn`) -- the seam this module's own inner unit tests use.

    Raises `UnknownAxisError` when `axis` is not in `AXES`,
    `NoEmbeddingsToClassifyError` when no persisted embedding table exists,
    `CorpusPinRequiredError` when no corpus pin can be resolved,
    `MissingGoldSheetError` when no gold answer-key sheet exists,
    `NoGoldLabelsError` when the gold sheet has no `{axis}_gold` judgments,
    `MissingGoldEmbeddingError` when a gold-judged chunk has no persisted
    vector, and `NoTrainingDataError` when no training examples survive
    gold-exclusion and rare-class filtering -- all loud failures rather than
    a silently empty or misleading manifest.
    """
    if axis not in AXES:
        raise UnknownAxisError(axis)
    axis_column = AXIS_METADATA_COLUMNS[axis]

    if embeddings_dir is None:
        embeddings_dir = DEFAULT_EMBEDDINGS_DIR
    embeddings_dir = Path(embeddings_dir)
    if gold_sheet_path is None:
        gold_sheet_path = DEFAULT_GOLD_SHEET_PATH
    gold_sheet_path = Path(gold_sheet_path)
    if manifest_path is None:
        manifest_path = _manifest_path_for(axis)
    manifest_path = Path(manifest_path)

    try:
        pin = resolve_current_pin(evals_dir)
    except _corpus_pin.CorpusPinError as exc:
        raise CorpusPinRequiredError(exc) from exc

    if not gold_sheet_path.is_file():
        raise MissingGoldSheetError(gold_sheet_path)
    gold_chunk_ids, gold_eval_records, teacher_gold_agreement = _load_gold_labels(
        gold_sheet_path, axis
    )
    if not gold_eval_records:
        raise NoGoldLabelsError(axis, gold_sheet_path)

    lookup = _load_embedding_lookup(embeddings_dir, axis_column)

    missing_gold_ids = [
        chunk_id for chunk_id, _label in gold_eval_records if chunk_id not in lookup
    ]
    if missing_gold_ids:
        raise MissingGoldEmbeddingError(axis, embeddings_dir, missing_gold_ids)

    training_records = [
        (chunk_id, vector, label)
        for chunk_id, (vector, label) in lookup.items()
        if chunk_id not in gold_chunk_ids and label
    ]
    training_records.sort(key=lambda record: record[0])
    training_records, dropped_classes = _drop_rare_classes(training_records, min_class_count)
    if not training_records:
        raise NoTrainingDataError(axis)

    if train_fn is None:
        train_fn = _default_train_fn
    predict_fn = train_fn(
        [vector for _chunk_id, vector, _label in training_records],
        [label for _chunk_id, _vector, label in training_records],
    )

    gold_vectors = [lookup[chunk_id][0] for chunk_id, _label in gold_eval_records]
    gold_labels = [label for _chunk_id, label in gold_eval_records]
    predictions = predict_fn(gold_vectors)

    full_coverage_accuracy = _full_coverage_accuracy(predictions, gold_labels)
    threshold_metrics = [
        _threshold_metrics(predictions, gold_labels, threshold) for threshold in thresholds
    ]

    gold_predictions = [
        {
            "chunk_id": chunk_id,
            "true": label,
            "predicted": predicted,
            "confidence": confidence,
            "correct": predicted == label,
        }
        for (chunk_id, label), (predicted, confidence) in zip(gold_eval_records, predictions)
    ]

    manifest = {
        "axis": axis,
        "corpus_pin_id": pin["corpus_pin_id"],
        "vault_snapshot_hash": pin["vault_snapshot_hash"],
        "config": {
            "technique": "logistic_regression_on_dense_embeddings",
            "max_iter": DEFAULT_MAX_ITER,
            "min_class_count": min_class_count,
            "thresholds": list(thresholds),
        },
        "train_chunk_count": len(training_records),
        "dropped_classes": dropped_classes,
        "gold_chunk_count": len(gold_eval_records),
        "full_coverage_accuracy": full_coverage_accuracy,
        "teacher_gold_agreement": teacher_gold_agreement,
        "thresholds": [
            {
                "threshold": metric.threshold,
                "coverage": metric.coverage,
                "covered_count": metric.covered_count,
                "accuracy_on_covered": metric.accuracy_on_covered,
            }
            for metric in threshold_metrics
        ],
        "gold_predictions": gold_predictions,
    }
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(
        json.dumps(manifest, indent=2, sort_keys=True, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    return ClassifyResult(
        manifest_path=manifest_path,
        axis=axis,
        train_chunk_count=len(training_records),
        dropped_classes=dropped_classes,
        gold_chunk_count=len(gold_eval_records),
        full_coverage_accuracy=full_coverage_accuracy,
        teacher_gold_agreement=teacher_gold_agreement,
        thresholds=threshold_metrics,
        corpus_pin_id=pin["corpus_pin_id"],
        vault_snapshot_hash=pin["vault_snapshot_hash"],
    )
