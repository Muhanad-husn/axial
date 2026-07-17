"""Outer acceptance test for issue #54, slice 02 (gold: render the label
sheet label_sheet.xlsx).

Locked behavioral contract (DEC-1) -- do not edit once committed red.

Given a sampled gold set under data/gold/chunks/ and the Syria codebook
When  the user runs `axial gold sheet`
Then  data/gold/label_sheet.xlsx exists with one header row and one row per
      sampled chunk
And   the columns are chunk_id, source, section, chunk_text, field,
      empirical_scope, polities_touched, claim_type, theory_school, notes in
      that order
And   the field and empirical_scope cells are pre-filled from each chunk's tags
And   the polities_touched cell is pre-filled from the chunk's tagged
      polities_touched list, "; "-joined, and carries NO dropdown validation
      (it is a context column, not a labeling axis)
And   the claim_type and theory_school cells are empty (blind)
And   the field, empirical_scope, claim_type and theory_school columns carry
      dropdown validation whose options are the codebook's vocabulary for
      that axis
And   re-running overwrites the sheet in place (no duplicate rows, no stale
      sheet)

See specs/PRODUCT.md §7.5 (label sheet), Appendix I (columns), §8 (P0-9).
Plan: plans/gold/02-label-sheet.md.

Arrange -- seed the sampled records directly
-----------------------------------------------------------------------
Slice 02 consumes slice 01's chunk records. This test seeds a handful of
records directly under `<root>/data/gold/chunks/` in the exact flat shape
`axial gold sample` writes (chunk_id, source, section, chunk_text, field,
empirical_scope, polities_touched, role_in_argument, claim_type,
theory_school -- axis values
as representative scalars), rather than running the sampler, so slice 02 is
tested independently of slice 01's selection logic. No LLM, no network.

Isolation -- the isolated staging root (issue #68)
-----------------------------------------------------------------------
Runs `axial gold sheet` with `cwd` set to `isolated_vault_root`, which also
copies `config/domains/syria/{schema,codebook}.yaml` into the root so the
default domain dir (the dropdown vocabulary source) resolves under the
staging cwd. The real `data/` tree is never touched.

Dropdown options -- resolved from whatever seam the sheet uses
-----------------------------------------------------------------------
The codebook vocabularies for claim_type/theory_school are far longer than
Excel's ~255-char inline-list limit, so the sheet may back its dropdowns
with a helper range rather than an inline list. This test does not lock the
mechanism: it resolves each axis column's DataValidation options whether
they are inline (`"a,b,c"`) or a range reference to another sheet, and
compares the resolved option SET to the codebook's vocabulary for that axis
(read from the real codebook.yaml, never hardcoded).
"""

from __future__ import annotations

import json
import os
import subprocess
from pathlib import Path

import yaml
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
CODEBOOK_PATH = REPO_ROOT / "config" / "domains" / "syria" / "codebook.yaml"

ARGPARSE_FALLBACK_MARKERS = ("invalid choice", "unrecognized arguments")

EXPECTED_COLUMNS = [
    "chunk_id",
    "source",
    "section",
    "chunk_text",
    "field",
    "empirical_scope",
    "polities_touched",
    "claim_type",
    "theory_school",
    "notes",
]

# 1-indexed column numbers of the four DROPDOWN axis columns (for dropdown
# checks). `polities_touched` (column 7) is deliberately excluded: it is a
# pre-filled context column, not one of the four labeling axes, and must
# carry no dropdown (Appendix I).
AXIS_COLUMNS = {
    "field": 5,
    "empirical_scope": 6,
    "claim_type": 8,
    "theory_school": 9,
}

# 1-indexed column number of the polities_touched context column.
POLITIES_TOUCHED_COLUMN = 7

# Seeded sampled records (slice 01's flat output shape).
SEEDED_RECORDS = [
    {
        "chunk_id": "alpha-history-000000000001_1_introduction_001",
        "source": "alpha-history-000000000001",
        "section": "Introduction",
        "chunk_text": "First substantive prose chunk.",
        "field": "state",
        "empirical_scope": "scope:general",
        "polities_touched": ["Syria", "Iraq"],
        "role_in_argument": "role:setup",
        "claim_type": "state-formation",
        "theory_school": "bellicist",
    },
    {
        "chunk_id": "alpha-history-000000000001_2_chapter-one_001",
        "source": "alpha-history-000000000001",
        "section": "Chapter One",
        "chunk_text": "Second substantive prose chunk.",
        "field": "violence",
        "empirical_scope": "scope:country-case",
        "polities_touched": ["Lebanon"],
        "role_in_argument": "role:claim",
        "claim_type": "civilian-targeting",
        "theory_school": "micro-sociological",
    },
    {
        "chunk_id": "beta-analysis-000000000002_1_analysis_001",
        "source": "beta-analysis-000000000002",
        "section": "Analysis",
        "chunk_text": "Third substantive prose chunk.",
        "field": "ideology",
        "empirical_scope": "scope:comparative",
        "polities_touched": [],
        "role_in_argument": "role:evidence",
        "claim_type": "ideology-as-system",
        "theory_school": "discursive",
    },
]


def _chunks_dir(root: Path) -> Path:
    return root / "data" / "gold" / "chunks"


def _sheet_path(root: Path) -> Path:
    return root / "data" / "gold" / "label_sheet.xlsx"


def _seed_records(root: Path) -> list[dict]:
    chunks_dir = _chunks_dir(root)
    chunks_dir.mkdir(parents=True, exist_ok=True)
    for record in SEEDED_RECORDS:
        (chunks_dir / f"{record['chunk_id']}.json").write_text(
            json.dumps(record, indent=2, sort_keys=True), encoding="utf-8"
        )
    return SEEDED_RECORDS


def _codebook_vocab(axis: str) -> set[str]:
    document = yaml.safe_load(CODEBOOK_PATH.read_text(encoding="utf-8"))
    return set(document["axes"][axis].keys())


def _run_gold_sheet(root: Path, *args: str) -> subprocess.CompletedProcess:
    env = dict(os.environ)
    env["AXIAL_LLM_PROVIDER"] = "explode"  # any LLM call is a bug
    return subprocess.run(
        ["uv", "run", "--project", str(REPO_ROOT), "axial", "gold", "sheet", *args],
        cwd=root,
        capture_output=True,
        text=True,
        env=env,
    )


def _assert_not_argparse_fallback(result: subprocess.CompletedProcess) -> None:
    combined = result.stdout + result.stderr
    for marker in ARGPARSE_FALLBACK_MARKERS:
        assert marker not in combined, (
            f"expected a real `gold sheet` behavior path, not an argparse "
            f"fallback (found {marker!r}) -- the subcommand does not exist "
            f"yet or was never reached:\nstdout: {result.stdout!r}\n"
            f"stderr: {result.stderr!r}"
        )


def _column_letter(col: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(col)


def _validation_options(wb, dv) -> set[str]:
    """Resolve a DataValidation's option set, tolerating an inline list
    (`"a,b,c"`) or a range reference to a helper sheet (`vocab!$A$2:$A$4`)."""
    formula = (dv.formula1 or "").strip()
    if not formula:
        return set()
    if formula.startswith('"') and formula.endswith('"'):
        return {opt for opt in formula.strip('"').split(",") if opt}

    ref = formula.lstrip("=")
    sheet_name, _, cell_range = ref.partition("!")
    sheet_name = sheet_name.strip("'")
    target = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    options: set[str] = set()
    for row in target[cell_range.replace("$", "")]:
        cells = row if isinstance(row, tuple) else (row,)
        for cell in cells:
            if cell.value is not None:
                options.add(str(cell.value))
    return options


def _validation_for_column(wb, ws, col: int):
    """Return the DataValidation applied to that column's first data cell
    (row 2), or None."""
    coord = f"{_column_letter(col)}2"
    for dv in ws.data_validations.dataValidation:
        if coord in dv.sqref:
            return dv
    return None


def _main_sheet(wb):
    """The label sheet is the first worksheet; a helper vocab sheet (if any)
    follows it."""
    return wb.worksheets[0]


def test_gold_sheet_renders_label_sheet(isolated_vault_root):
    root = isolated_vault_root
    records = _seed_records(root)

    result = _run_gold_sheet(root)
    _assert_not_argparse_fallback(result)
    assert result.returncode == 0, (
        f"expected exit code 0 for `axial gold sheet` on a seeded gold set, "
        f"got {result.returncode}\nstdout: {result.stdout!r}\n"
        f"stderr: {result.stderr!r}"
    )

    sheet_path = _sheet_path(root)
    assert sheet_path.is_file(), (
        f"expected `axial gold sheet` to write {sheet_path}, but it does not exist"
    )

    wb = load_workbook(sheet_path)
    ws = _main_sheet(wb)

    # Header row: exactly Appendix I's columns in order.
    header = [ws.cell(row=1, column=i + 1).value for i in range(len(EXPECTED_COLUMNS))]
    assert header == EXPECTED_COLUMNS, (
        f"expected the header row to be Appendix I's columns in order "
        f"{EXPECTED_COLUMNS}, got {header}"
    )

    # One data row per sampled chunk.
    assert ws.max_row == 1 + len(records), (
        f"expected one header row plus one row per sampled chunk "
        f"({1 + len(records)} rows), got max_row={ws.max_row}"
    )

    # Rows keyed by chunk_id, so column order in the file is what we assert on.
    by_chunk = {}
    for row in range(2, ws.max_row + 1):
        values = {
            col: ws.cell(row=row, column=idx + 1).value for idx, col in enumerate(EXPECTED_COLUMNS)
        }
        by_chunk[values["chunk_id"]] = values

    assert set(by_chunk) == {r["chunk_id"] for r in records}, (
        f"expected exactly one row per seeded chunk_id, got rows for {sorted(by_chunk)}"
    )

    for record in records:
        row = by_chunk[record["chunk_id"]]
        # Verbatim provenance columns.
        for col in ("chunk_id", "source", "section", "chunk_text"):
            assert row[col] == record[col], (
                f"expected column {col!r} of chunk {record['chunk_id']!r} to be "
                f"{record[col]!r}, got {row[col]!r}"
            )
        # Pre-labeled columns filled from the chunk's own tags.
        assert row["field"] == record["field"], (
            f"expected the field cell to be pre-filled with the chunk's tag "
            f"{record['field']!r}, got {row['field']!r}"
        )
        assert row["empirical_scope"] == record["empirical_scope"], (
            f"expected the empirical_scope cell to be pre-filled with the "
            f"chunk's tag {record['empirical_scope']!r}, got "
            f"{row['empirical_scope']!r}"
        )
        # polities_touched: pre-filled context column, "; "-joined from the
        # chunk's tagged polities list -- not an axis, never blind.
        expected_polities = "; ".join(record.get("polities_touched") or [])
        if expected_polities:
            assert row["polities_touched"] == expected_polities, (
                f"expected the polities_touched cell to be the chunk's tagged "
                f"polities joined with '; ' ({expected_polities!r}), got "
                f"{row['polities_touched']!r}"
            )
        else:
            assert row["polities_touched"] in (None, ""), (
                f"expected an empty polities_touched cell for a chunk with no "
                f"tagged polities, got {row['polities_touched']!r}"
            )
        # Blind columns empty for the Academic.
        assert row["claim_type"] in (None, ""), (
            f"expected the claim_type cell to be blind (empty), got {row['claim_type']!r}"
        )
        assert row["theory_school"] in (None, ""), (
            f"expected the theory_school cell to be blind (empty), got {row['theory_school']!r}"
        )
        assert row["notes"] in (None, ""), (
            f"expected the notes cell to ship empty, got {row['notes']!r}"
        )

    # Each axis column carries a dropdown whose options are the codebook vocab.
    for axis, col in AXIS_COLUMNS.items():
        dv = _validation_for_column(wb, ws, col)
        assert dv is not None, (
            f"expected the {axis!r} column (column {col}) to carry a dropdown "
            f"data-validation on its data cells, found none"
        )
        assert (dv.type or "").lower() == "list", (
            f"expected the {axis!r} column's validation to be a 'list' "
            f"(dropdown), got type {dv.type!r}"
        )
        options = _validation_options(wb, dv)
        expected_vocab = _codebook_vocab(axis)
        assert options == expected_vocab, (
            f"expected the {axis!r} dropdown options to equal the codebook's "
            f"vocabulary for that axis ({sorted(expected_vocab)}), got "
            f"{sorted(options)}"
        )

    # polities_touched is a context column, not an axis: header present, but
    # no dropdown/data-validation on its data cells.
    header_at_7 = ws.cell(row=1, column=POLITIES_TOUCHED_COLUMN).value
    assert header_at_7 == "polities_touched", (
        f"expected column {POLITIES_TOUCHED_COLUMN} to be the polities_touched "
        f"header, got {header_at_7!r}"
    )
    polities_dv = _validation_for_column(wb, ws, POLITIES_TOUCHED_COLUMN)
    assert polities_dv is None, (
        f"expected the polities_touched column (column {POLITIES_TOUCHED_COLUMN}) "
        f"to carry NO dropdown/data-validation -- it is a pre-filled context "
        f"column, not one of the four labeling axes -- but found {polities_dv!r}"
    )


def test_gold_sheet_overwrites_in_place(isolated_vault_root):
    root = isolated_vault_root
    records = _seed_records(root)

    first = _run_gold_sheet(root)
    _assert_not_argparse_fallback(first)
    assert first.returncode == 0, (
        f"expected exit code 0 for the first `axial gold sheet` run, got "
        f"{first.returncode}\nstderr: {first.stderr!r}"
    )
    first_rows = load_workbook(_sheet_path(root)).worksheets[0].max_row

    second = _run_gold_sheet(root)
    _assert_not_argparse_fallback(second)
    assert second.returncode == 0, (
        f"expected exit code 0 for the second `axial gold sheet` run, got "
        f"{second.returncode}\nstderr: {second.stderr!r}"
    )
    second_rows = load_workbook(_sheet_path(root)).worksheets[0].max_row

    assert first_rows == second_rows == 1 + len(records), (
        f"expected re-running `axial gold sheet` to overwrite the sheet in "
        f"place with one header + one row per chunk ({1 + len(records)} rows), "
        f"got {first_rows} then {second_rows}"
    )
