"""Outer acceptance test for `axial gold deliver` -- package the generated
label sheet into a self-contained handoff bundle for the Academic.

Locked behavioral contract -- do not edit once committed red.

Given a generated `data/gold/label_sheet.xlsx` (the P0-9 deliverable)
When  the user runs `axial gold deliver`
Then  a dated delivery folder `data/gold/delivery/<YYYY-MM-DD>/` exists
And   it contains exactly three files: label_sheet.xlsx, a byte-identical
      copy of the generated sheet; README-for-academic.md, the human-facing
      labeling instructions; and manifest.json, the machine-readable summary
And   manifest.json records the chunk_count, Appendix I's full column list
      (including the pre-filled polities_touched context column), the four
      axis columns, which axes are labeled blind (claim_type, theory_school)
      vs pre-labeled (field, empirical_scope), the sheet filename, the
      delivery date, and where the Academic returns the filled sheet
      (data/gold/labels/)
And   README-for-academic.md names the four axes, the blind/correct split,
      and the return location
And   re-running overwrites the same dated folder in place (no stale files)
And   running deliver with no generated sheet fails with a clear error that
      tells the operator to run `axial gold sheet` first.

This step has no spec section of its own yet (specs/PRODUCT.md §7.5 specs
*producing* the sheet; the delivery handoff between build step 4 and the
step-5 Academic-labeling pause is unspecified). Delivery is deliberately
LOCAL and offline: no Drive, no email, no network -- a reviewable bundle on
disk. The columns/axis contract it echoes is Appendix I (§7.5, §9).

Isolation mirrors test_gold_sheet.py: run under `isolated_vault_root` so the
real `data/` tree is never touched, and arrange the precondition by running
the real `axial gold sheet` upstream rather than hand-crafting an xlsx.
"""

from __future__ import annotations

import json
import os
import re
import subprocess
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent.parent

ARGPARSE_FALLBACK_MARKERS = ("invalid choice", "unrecognized arguments")

DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

EXPECTED_AXES = ["field", "empirical_scope", "claim_type", "theory_school"]
EXPECTED_BLIND = ["claim_type", "theory_school"]
EXPECTED_PRELABELED = ["field", "empirical_scope"]

# Appendix I's full label-sheet column order, including the pre-filled
# `polities_touched` context column (not an axis) between empirical_scope
# and claim_type. The delivery manifest's `columns` field must echo this.
EXPECTED_COLUMNS = [
    "chunk_id",
    "source",
    "section",
    "chunk_text",
    "field",
    "empirical_scope",
    "polities_touched",
    "claim_type",
    "theory_school",
    "notes",
]

# Seeded sampled records, reusing slice 01's flat output shape (see
# test_gold_sheet.py). `gold sheet` renders these; `gold deliver` packages them.
SEEDED_RECORDS = [
    {
        "chunk_id": "alpha-history-000000000001_1_introduction_001",
        "source": "alpha-history-000000000001",
        "section": "Introduction",
        "chunk_text": "First substantive prose chunk.",
        "field": "state",
        "empirical_scope": "scope:general",
        "polities_touched": ["Syria", "Iraq"],
        "role_in_argument": "role:setup",
        "claim_type": "state-formation",
        "theory_school": "bellicist",
    },
    {
        "chunk_id": "alpha-history-000000000001_2_chapter-one_001",
        "source": "alpha-history-000000000001",
        "section": "Chapter One",
        "chunk_text": "Second substantive prose chunk.",
        "field": "violence",
        "empirical_scope": "scope:country-case",
        "polities_touched": ["Lebanon"],
        "role_in_argument": "role:claim",
        "claim_type": "civilian-targeting",
        "theory_school": "micro-sociological",
    },
    {
        "chunk_id": "beta-analysis-000000000002_1_analysis_001",
        "source": "beta-analysis-000000000002",
        "section": "Analysis",
        "chunk_text": "Third substantive prose chunk.",
        "field": "ideology",
        "empirical_scope": "scope:comparative",
        "polities_touched": [],
        "role_in_argument": "role:evidence",
        "claim_type": "ideology-as-system",
        "theory_school": "discursive",
    },
]


def _gold_dir(root: Path) -> Path:
    return root / "data" / "gold"


def _chunks_dir(root: Path) -> Path:
    return _gold_dir(root) / "chunks"


def _sheet_path(root: Path) -> Path:
    return _gold_dir(root) / "label_sheet.xlsx"


def _delivery_root(root: Path) -> Path:
    return _gold_dir(root) / "delivery"


def _seed_records(root: Path) -> list[dict]:
    chunks_dir = _chunks_dir(root)
    chunks_dir.mkdir(parents=True, exist_ok=True)
    for record in SEEDED_RECORDS:
        (chunks_dir / f"{record['chunk_id']}.json").write_text(
            json.dumps(record, indent=2, sort_keys=True), encoding="utf-8"
        )
    return SEEDED_RECORDS


def _run_gold(root: Path, *args: str) -> subprocess.CompletedProcess:
    env = dict(os.environ)
    env["AXIAL_LLM_PROVIDER"] = "explode"  # any LLM call is a bug
    return subprocess.run(
        ["uv", "run", "--project", str(REPO_ROOT), "axial", "gold", *args],
        cwd=root,
        capture_output=True,
        text=True,
        env=env,
    )


def _assert_not_argparse_fallback(result: subprocess.CompletedProcess) -> None:
    combined = result.stdout + result.stderr
    for marker in ARGPARSE_FALLBACK_MARKERS:
        assert marker not in combined, (
            f"expected a real `gold deliver` behavior path, not an argparse "
            f"fallback (found {marker!r}) -- the subcommand does not exist yet "
            f"or was never reached:\nstdout: {result.stdout!r}\n"
            f"stderr: {result.stderr!r}"
        )


def _generate_sheet(root: Path) -> None:
    """Arrange the precondition: seed records and render the real label sheet."""
    _seed_records(root)
    sheet = _run_gold(root, "sheet")
    _assert_not_argparse_fallback(sheet)
    assert sheet.returncode == 0, (
        f"arrange failed: `axial gold sheet` did not succeed\n"
        f"stdout: {sheet.stdout!r}\nstderr: {sheet.stderr!r}"
    )
    assert _sheet_path(root).is_file(), "arrange failed: no label_sheet.xlsx produced"


def _sole_delivery_dir(root: Path) -> Path:
    delivery_root = _delivery_root(root)
    assert delivery_root.is_dir(), (
        f"expected a delivery root at {delivery_root}, but it does not exist"
    )
    subdirs = sorted(p for p in delivery_root.iterdir() if p.is_dir())
    assert len(subdirs) == 1, (
        f"expected exactly one dated delivery folder under {delivery_root}, "
        f"got {[p.name for p in subdirs]}"
    )
    folder = subdirs[0]
    assert DATE_RE.match(folder.name), (
        f"expected the delivery folder to be date-stamped (YYYY-MM-DD), got {folder.name!r}"
    )
    return folder


def test_gold_deliver_packages_sheet_for_academic(isolated_vault_root):
    root = isolated_vault_root
    _generate_sheet(root)

    result = _run_gold(root, "deliver")
    _assert_not_argparse_fallback(result)
    assert result.returncode == 0, (
        f"expected exit code 0 for `axial gold deliver` on a generated sheet, "
        f"got {result.returncode}\nstdout: {result.stdout!r}\n"
        f"stderr: {result.stderr!r}"
    )

    delivery_dir = _sole_delivery_dir(root)

    # Exactly the three handoff artifacts, nothing else.
    names = sorted(p.name for p in delivery_dir.iterdir())
    assert names == ["README-for-academic.md", "label_sheet.xlsx", "manifest.json"], (
        f"expected the delivery folder to hold exactly the three handoff files, got {names}"
    )

    # The delivered sheet is a byte-identical copy of the generated one.
    delivered_sheet = delivery_dir / "label_sheet.xlsx"
    assert delivered_sheet.read_bytes() == _sheet_path(root).read_bytes(), (
        "expected the delivered label_sheet.xlsx to be a byte-identical copy "
        "of the generated data/gold/label_sheet.xlsx"
    )
    # And it is a real, loadable workbook with the seeded rows.
    delivered_rows = load_workbook(delivered_sheet).worksheets[0].max_row
    assert delivered_rows == 1 + len(SEEDED_RECORDS), (
        f"expected the delivered sheet to carry one header + one row per chunk "
        f"({1 + len(SEEDED_RECORDS)} rows), got {delivered_rows}"
    )

    # The manifest describes the handoff for the eval harness.
    manifest = json.loads((delivery_dir / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["chunk_count"] == len(SEEDED_RECORDS), (
        f"expected manifest chunk_count={len(SEEDED_RECORDS)}, got {manifest.get('chunk_count')!r}"
    )
    assert manifest["columns"] == EXPECTED_COLUMNS, (
        f"expected manifest columns to equal Appendix I's label-sheet columns "
        f"in order ({EXPECTED_COLUMNS}, including the polities_touched context "
        f"column), got {manifest.get('columns')!r}"
    )
    assert manifest["axes"] == EXPECTED_AXES, (
        f"expected manifest axes={EXPECTED_AXES}, got {manifest.get('axes')!r}"
    )
    assert manifest["blind_axes"] == EXPECTED_BLIND, (
        f"expected manifest blind_axes={EXPECTED_BLIND}, got {manifest.get('blind_axes')!r}"
    )
    assert manifest["prelabeled_axes"] == EXPECTED_PRELABELED, (
        f"expected manifest prelabeled_axes={EXPECTED_PRELABELED}, got "
        f"{manifest.get('prelabeled_axes')!r}"
    )
    assert manifest["prelabeled_freetext"] == ["polities_touched"], (
        f"expected manifest prelabeled_freetext=['polities_touched'] to flag "
        f"the pre-filled, non-axis context column, got "
        f"{manifest.get('prelabeled_freetext')!r}"
    )
    assert manifest["sheet"] == "label_sheet.xlsx", (
        f"expected manifest sheet='label_sheet.xlsx', got {manifest.get('sheet')!r}"
    )
    assert DATE_RE.match(str(manifest.get("delivered", ""))), (
        f"expected manifest delivered to be a YYYY-MM-DD date, got {manifest.get('delivered')!r}"
    )
    assert manifest["delivered"] == delivery_dir.name, (
        f"expected manifest delivered ({manifest.get('delivered')!r}) to match "
        f"the delivery folder name ({delivery_dir.name!r})"
    )
    assert "labels" in str(manifest.get("return_to", "")), (
        f"expected manifest return_to to point at the labels inbox, got "
        f"{manifest.get('return_to')!r}"
    )

    # The README tells the Academic what to do.
    readme = (delivery_dir / "README-for-academic.md").read_text(encoding="utf-8")
    assert readme.strip(), "expected README-for-academic.md to be non-empty"
    for axis in EXPECTED_AXES:
        assert axis in readme, f"expected the README to name the {axis!r} axis, but it does not"
    assert "labels" in readme, (
        "expected the README to tell the Academic where to return the sheet (data/gold/labels/)"
    )
    assert "polities_touched" in readme, (
        "expected the README to mention the pre-filled polities_touched context "
        "column so the Academic knows it is not a blind labeling axis"
    )


def test_gold_deliver_overwrites_delivery_in_place(isolated_vault_root):
    root = isolated_vault_root
    _generate_sheet(root)

    first = _run_gold(root, "deliver")
    _assert_not_argparse_fallback(first)
    assert first.returncode == 0, (
        f"expected exit code 0 for the first `axial gold deliver`, got "
        f"{first.returncode}\nstderr: {first.stderr!r}"
    )
    first_dir = _sole_delivery_dir(root)
    # Drop a stale file that a clean re-run must remove.
    (first_dir / "stale.txt").write_text("stale", encoding="utf-8")

    second = _run_gold(root, "deliver")
    _assert_not_argparse_fallback(second)
    assert second.returncode == 0, (
        f"expected exit code 0 for the second `axial gold deliver`, got "
        f"{second.returncode}\nstderr: {second.stderr!r}"
    )
    second_dir = _sole_delivery_dir(root)

    names = sorted(p.name for p in second_dir.iterdir())
    assert names == ["README-for-academic.md", "label_sheet.xlsx", "manifest.json"], (
        f"expected re-running deliver to overwrite the folder in place with "
        f"exactly the three handoff files (no stale.txt), got {names}"
    )


def test_gold_deliver_errors_without_sheet(isolated_vault_root):
    root = isolated_vault_root
    # No `gold sheet` run -- there is nothing to deliver.
    result = _run_gold(root, "deliver")
    _assert_not_argparse_fallback(result)
    assert result.returncode != 0, (
        "expected a non-zero exit when there is no generated label_sheet.xlsx "
        f"to deliver, got {result.returncode}\nstdout: {result.stdout!r}"
    )
    combined = (result.stdout + result.stderr).lower()
    assert "gold sheet" in combined, (
        "expected the error to tell the operator to run `axial gold sheet` "
        f"first, got:\nstdout: {result.stdout!r}\nstderr: {result.stderr!r}"
    )
    assert not _delivery_root(root).exists(), (
        "expected no delivery folder to be created when there is no sheet to deliver"
    )
