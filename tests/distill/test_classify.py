"""Outer acceptance test for issues #351/#352 (stage-5d: TF-IDF classifier
for `claim_type`/`theory_school`, DEC-37/DEC-38).

Given a vault of tagged chunks and an independent gold answer-key sheet
When  `axial distill classify <axis>` trains and evaluates the classifier
Then  it runs with zero LLM calls
And   the gold-sampled chunk is excluded from training (leakage-free)
And   a class below the min-training-example floor is dropped, not trained on
And   the confidence-threshold split behaves correctly across multiple
      thresholds (coverage never increases as the threshold rises)
And   no source chunk_text appears in the emitted manifest (DEC-23)

See `plans/phase-a-completion/README.md` stage 5d,
`docs/DECISIONS.md` DEC-37/DEC-38, `docs/eval/02-hybrid-tagging-distillation.md`.

Isolation -- the isolated staging root (issue #68), same seam as
`tests/distill/test_embedding_pass.py` / `tests/distill/test_readiness_map.py`
-----------------------------------------------------------------------
`axial distill classify` resolves `data/vault/`, `data/gold/labels/
label_sheet.xlsx`, `data/distill/classify_<axis>_manifest.json`, and
`evals/corpus_pin/` as plain paths relative to the process's cwd. This test
runs the CLI as a subprocess with `cwd` set to `isolated_vault_root`, never
touching the real, shared `data/` tree.

Fixture design -- lexically distinct, fabricated vocabulary per class
-----------------------------------------------------------------------
Three vault chunk groups, each combining a distinct `claim_type` vocabulary
with a distinct `theory_school` vocabulary, so TF-IDF cleanly separates
every class:

  - group A (10 chunks): claim_type=descriptive-empirical,
    theory_school=materialist
  - group B (10 chunks): claim_type=theoretical-conceptual,
    theory_school=institutionalist
  - group C (3 chunks, deliberately BELOW the min-class-count floor of 6):
    claim_type=normative-prescriptive, theory_school=institutionalist (so
    it pools into group B's theory_school class -- 13 total, not dropped --
    while its own claim_type class stays rare and IS dropped; this
    exercises both the "dropped" and "not dropped" paths in one fixture)

The gold sheet's first row reuses group A's own `a_000` chunk_id (with a
fresh chunk_text variant) -- the leakage-exclusion check: `run_classify`
must exclude it from training by chunk_id even though its label/vocabulary
would otherwise train cleanly. Four more gold rows use chunk_ids absent
from the vault entirely (a real gold sheet's chunk_text is its own copy,
independent of the vault note). Every prediction/threshold value asserted
below was verified directly against the real `_default_train_fn` pipeline
before being hardcoded (see the PR body for the derivation), the same
"verify then hardcode" precedent `tests/distill/test_readiness_map.py` sets.
"""

from __future__ import annotations

import json
import os
import subprocess
from pathlib import Path

import pytest
from openpyxl import Workbook

pytest.importorskip("sklearn")

from axial.vault import render_note  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent.parent

PROVIDER_ENV_VAR = "AXIAL_LLM_PROVIDER"

CLAIM_DESC = "empirical observational fieldwork survey data collection descriptive documented pattern measurement"
CLAIM_THEORY = (
    "theoretical conceptual abstract framework model paradigm philosophical logic deductive"
)
CLAIM_NORM = "normative prescriptive policy recommendation ought reform proposal advocate should"
THEORY_MATERIALIST = (
    "material economic resources production wealth distribution capital labor market"
)
THEORY_INST = (
    "institutions rules organizations governance formal structures bureaucracy procedure legal"
)


def _run_axial(root: Path, *args: str) -> subprocess.CompletedProcess:
    env = dict(os.environ)
    env[PROVIDER_ENV_VAR] = "explode"  # poison: any text-gen LLM call crashes the run
    return subprocess.run(
        ["uv", "run", "--project", str(REPO_ROOT), "axial", *args],
        cwd=root,
        capture_output=True,
        text=True,
        env=env,
    )


def _chunk_text(claim_words: str, theory_words: str, variant: int) -> str:
    return f"{claim_words} {theory_words} chunk variant {variant}"


def _write_chunk_note(
    prose_dir: Path,
    chunk_id: str,
    chunk_text: str,
    *,
    claim_type_primary: str,
    theory_school_primary: str,
) -> None:
    prose_dir.mkdir(parents=True, exist_ok=True)
    frontmatter = {
        "chunk_id": chunk_id,
        "section": "Introduction",
        "chunk_text": chunk_text,
        "source_meta": {"author": "A", "title": "T"},
        "schema_version": "0.1",
        "role_in_argument": "role:claim",
        "field": {"primary": "state", "secondary": []},
        "claim_type": {"primary": claim_type_primary, "secondary": None, "subtags": []},
        "theory_school": {
            "primary": theory_school_primary,
            "secondary": None,
            "status": "candidate",
        },
        "empirical_scope": {"value": "scope:country-case", "polity": "Syria"},
        "polities_touched": ["Syria"],
        "artifact_refs": [],
    }
    (prose_dir / f"{chunk_id}.md").write_text(
        render_note(frontmatter, "# Introduction\n\nbody\n"), encoding="utf-8"
    )


# group A: 10 chunks, group B: 10 chunks, group C: 3 chunks (below the
# min-class-count floor of 6 for its OWN claim_type value).
GROUP_A_IDS = [f"a_{i:03d}_x_001" for i in range(10)]
GROUP_B_IDS = [f"b_{i:03d}_x_001" for i in range(10)]
GROUP_C_IDS = [f"c_{i:03d}_x_001" for i in range(3)]

GOLD_ROWS = [
    # reuses group A's own chunk_id -- the leakage-exclusion case.
    (
        "a_000_x_001",
        _chunk_text(CLAIM_DESC, THEORY_MATERIALIST, 100),
        "descriptive-empirical",
        "materialist",
    ),
    (
        "gold_001",
        _chunk_text(CLAIM_DESC, THEORY_MATERIALIST, 101),
        "descriptive-empirical",
        "materialist",
    ),
    (
        "gold_002",
        _chunk_text(CLAIM_THEORY, THEORY_INST, 102),
        "theoretical-conceptual",
        "institutionalist",
    ),
    (
        "gold_003",
        _chunk_text(CLAIM_DESC, THEORY_MATERIALIST, 103),
        "descriptive-empirical",
        "materialist",
    ),
    (
        "gold_004",
        _chunk_text(CLAIM_THEORY, THEORY_INST, 104),
        "theoretical-conceptual",
        "institutionalist",
    ),
]


def _build_fixture_vault(root: Path) -> None:
    prose_dir = root / "data" / "vault" / "prose"
    for index, chunk_id in enumerate(GROUP_A_IDS):
        _write_chunk_note(
            prose_dir,
            chunk_id,
            _chunk_text(CLAIM_DESC, THEORY_MATERIALIST, index),
            claim_type_primary="descriptive-empirical",
            theory_school_primary="materialist",
        )
    for index, chunk_id in enumerate(GROUP_B_IDS):
        _write_chunk_note(
            prose_dir,
            chunk_id,
            _chunk_text(CLAIM_THEORY, THEORY_INST, index),
            claim_type_primary="theoretical-conceptual",
            theory_school_primary="institutionalist",
        )
    for index, chunk_id in enumerate(GROUP_C_IDS):
        _write_chunk_note(
            prose_dir,
            chunk_id,
            _chunk_text(CLAIM_NORM, THEORY_INST, index),
            claim_type_primary="normative-prescriptive",
            theory_school_primary="institutionalist",
        )

    envelopes_dir = root / "data" / "envelopes"
    envelopes_dir.mkdir(parents=True, exist_ok=True)
    result = _run_axial(root, "pin", "write", "baseline")
    assert result.returncode == 0, (
        f"fixture setup: `axial pin write baseline` failed\n"
        f"stdout: {result.stdout!r}\nstderr: {result.stderr!r}"
    )


def _write_gold_sheet(root: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "label_sheet"
    columns = ("chunk_id", "chunk_text", "claim_type", "theory_school")
    for col, name in enumerate(columns, start=1):
        sheet.cell(row=1, column=col, value=name)
    for row_index, row in enumerate(GOLD_ROWS, start=2):
        for col, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col, value=value)
    labels_dir = root / "data" / "gold" / "labels"
    labels_dir.mkdir(parents=True, exist_ok=True)
    workbook.save(labels_dir / "label_sheet.xlsx")


def _assert_ran_the_real_subcommand(result: subprocess.CompletedProcess) -> None:
    combined_output = result.stdout + result.stderr
    assert (
        "invalid choice" not in combined_output and "unrecognized arguments" not in combined_output
    ), (
        "expected a real 'axial distill classify' run, not an argparse fallback -- "
        "this means the CLI subcommand does not exist yet:\n"
        f"stdout: {result.stdout!r}\nstderr: {result.stderr!r}"
    )


def test_classify_claim_type_excludes_gold_drops_rare_class_zero_llm_no_chunk_text(
    isolated_vault_root,
):
    root = isolated_vault_root
    _build_fixture_vault(root)
    _write_gold_sheet(root)

    result = _run_axial(root, "distill", "classify", "claim_type")
    _assert_ran_the_real_subcommand(result)
    assert result.returncode == 0, (
        f"expected exit 0, got {result.returncode}\nstdout: {result.stdout!r}\nstderr: {result.stderr!r}"
    )

    manifest_path = root / "data" / "distill" / "classify_claim_type_manifest.json"
    assert manifest_path.is_file(), f"expected a classify manifest at {manifest_path}"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    assert manifest["axis"] == "claim_type"
    assert manifest["corpus_pin_id"] == "baseline"
    assert isinstance(manifest["vault_snapshot_hash"], str) and manifest["vault_snapshot_hash"]

    # -- zero LLM calls: AXIAL_LLM_PROVIDER=explode above would have crashed
    # the run had any text-generating call been attempted; exit 0 already
    # proves it, this just documents the intent.

    # -- leakage exclusion: 23 vault chunks carry a claim_type value; "a_000"
    # is excluded (it is also a gold chunk_id), and the 3-chunk
    # normative-prescriptive class is dropped (below the min-class-count
    # floor of 6) -- 9 (group A minus a_000) + 10 (group B) = 19.
    assert manifest["train_chunk_count"] == 19
    assert manifest["dropped_classes"] == ["normative-prescriptive"]

    # -- gold evaluation: all 5 gold rows carry a claim_type label.
    assert manifest["gold_chunk_count"] == 5
    assert manifest["full_coverage_accuracy"] == pytest.approx(1.0)

    # -- the threshold sweep: coverage never increases as the threshold
    # rises (verified directly against the real pipeline: stays 1.0 through
    # 0.5/0.6/0.7/0.8 here, since the fabricated vocabulary is cleanly
    # separable and every gold prediction lands above 0.8 confidence).
    by_threshold = {row["threshold"]: row for row in manifest["thresholds"]}
    assert [row["threshold"] for row in manifest["thresholds"]] == [0.5, 0.6, 0.7, 0.8]
    coverages = [by_threshold[t]["coverage"] for t in (0.5, 0.6, 0.7, 0.8)]
    assert coverages == sorted(coverages, reverse=True), (
        "coverage must never rise with the threshold"
    )
    assert by_threshold[0.8]["coverage"] == pytest.approx(1.0)
    assert by_threshold[0.8]["covered_count"] == 5
    assert by_threshold[0.8]["accuracy_on_covered"] == pytest.approx(1.0)

    # -- no eval_report.json staged in this fixture -- non-fatal, not an error.
    assert manifest["teacher_gold_agreement"] is None

    # -- DEC-23: no fixture chunk_text anywhere in the emitted manifest.
    manifest_text = json.dumps(manifest)
    for words in (CLAIM_DESC, CLAIM_THEORY, CLAIM_NORM, THEORY_MATERIALIST, THEORY_INST):
        assert words not in manifest_text, (
            "DEC-23 violation: fixture chunk_text leaked into the manifest"
        )


def test_classify_theory_school_pools_group_c_into_institutionalist_no_class_dropped(
    isolated_vault_root,
):
    """theory_school's own class shape differs from claim_type's on the
    SAME fixture: group C's 3 chunks carry `theory_school=institutionalist`
    (pooling with group B's 10 into 13), so nothing is dropped here even
    though `claim_type`'s parallel run (above) drops group C's rare
    `claim_type` class -- this is the "not dropped" half of the fixture."""
    root = isolated_vault_root
    _build_fixture_vault(root)
    _write_gold_sheet(root)

    result = _run_axial(root, "distill", "classify", "theory_school")
    _assert_ran_the_real_subcommand(result)
    assert result.returncode == 0, (
        f"expected exit 0, got {result.returncode}\nstdout: {result.stdout!r}\nstderr: {result.stderr!r}"
    )

    manifest_path = root / "data" / "distill" / "classify_theory_school_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    assert manifest["axis"] == "theory_school"
    # 9 (group A minus gold-excluded a_000) + 10 (group B) + 3 (group C) = 22.
    assert manifest["train_chunk_count"] == 22
    assert manifest["dropped_classes"] == []

    assert manifest["gold_chunk_count"] == 5
    assert manifest["full_coverage_accuracy"] == pytest.approx(1.0)

    by_threshold = {row["threshold"]: row for row in manifest["thresholds"]}
    coverages = [by_threshold[t]["coverage"] for t in (0.5, 0.6, 0.7, 0.8)]
    assert coverages == sorted(coverages, reverse=True), (
        "coverage must never rise with the threshold"
    )
    # verified directly against the real pipeline: the materialist rows'
    # confidence (~0.79) clears every threshold through 0.7 but not 0.8, so
    # coverage genuinely drops at the top of the sweep on this axis --
    # demonstrating the split moves, not just a flat 1.0 everywhere.
    assert by_threshold[0.7]["coverage"] == pytest.approx(1.0)
    assert by_threshold[0.8]["coverage"] == pytest.approx(0.4)
    assert by_threshold[0.8]["covered_count"] == 2
    assert by_threshold[0.8]["accuracy_on_covered"] == pytest.approx(1.0)

    manifest_text = json.dumps(manifest)
    for words in (CLAIM_DESC, CLAIM_THEORY, CLAIM_NORM, THEORY_MATERIALIST, THEORY_INST):
        assert words not in manifest_text, (
            "DEC-23 violation: fixture chunk_text leaked into the manifest"
        )


def test_classify_fails_loudly_without_a_gold_sheet(isolated_vault_root):
    root = isolated_vault_root
    _build_fixture_vault(root)
    # deliberately never write data/gold/labels/label_sheet.xlsx

    result = _run_axial(root, "distill", "classify", "claim_type")
    _assert_ran_the_real_subcommand(result)

    assert result.returncode != 0, (
        f"expected a non-zero exit with no gold sheet, got 0\nstdout: {result.stdout!r}\nstderr: {result.stderr!r}"
    )
    assert not (root / "data" / "distill" / "classify_claim_type_manifest.json").is_file()
    assert "gold" in result.stderr.lower(), (
        f"expected the failure to mention the missing gold sheet, got stderr: {result.stderr!r}"
    )


def test_classify_unknown_axis_rejected_by_the_cli(isolated_vault_root):
    root = isolated_vault_root
    (root / "data").mkdir(parents=True, exist_ok=True)

    # "field" used to be the unknown-axis case here, but #350 gave it a real
    # classifier (`axial.distill.classify_embedding`) -- use a string that is
    # not, and never will be, a real tag axis.
    result = _run_axial(root, "distill", "classify", "not-a-real-axis")

    assert result.returncode != 0
    assert "invalid choice" in (result.stdout + result.stderr)
