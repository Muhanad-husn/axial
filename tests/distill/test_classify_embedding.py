"""Outer acceptance tests for the stage-5d dense-embedding classifier --
`field` (issue #350) and `role_in_argument` (issue #348), DEC-39.

Given the persisted 5a embedding store and an independent gold answer-key
sheet
When  `axial distill classify <axis>` trains and evaluates the classifier
Then  it runs with zero LLM calls, and never re-embeds text (only sklearn
      touches the vectors)
And   the gold-sampled chunk is excluded from training (leakage-free)
And   a class below the min-training-example floor is dropped, not trained on
And   `teacher_gold_agreement` is computed from the gold sheet's own
      `<axis>`-vs-`<axis>_gold` columns (DEC-39's independent re-judgment),
      NOT from `eval_report.json`'s stale/absent number for this axis
And   the confidence-threshold split behaves correctly across multiple
      thresholds (coverage never increases as the threshold rises)
And   no chunk_text appears anywhere in the emitted manifest (DEC-23)

See `plans/phase-a-completion/README.md` stage 5d, `docs/DECISIONS.md`
DEC-37/DEC-38/DEC-39, `docs/eval/02-hybrid-tagging-distillation.md`.

Isolation -- the isolated staging root (issue #68), same seam as
`tests/distill/test_readiness_map.py`
-----------------------------------------------------------------------
`axial distill classify <axis>` resolves `data/distill/embeddings.lance`,
`data/gold/labels/label_sheet.xlsx`, `data/distill/classify_<axis>_manifest.json`,
and `evals/corpus_pin/` as plain paths relative to the process's cwd. This
test runs the CLI as a subprocess with `cwd` set to `isolated_vault_root`,
never touching the real, shared `data/` tree.

Seam decision -- the 5a embedding store fixture is built directly, not
through the real sentence-transformer
-----------------------------------------------------------------------
Same reasoning as `tests/distill/test_readiness_map.py`: this module's own
subject matter is the classifier, not the embedding pass (5a has its own
coverage). The `field` test below goes through `run_embed(..., encoder=
<deterministic fake>)` -- the real `axial.distill.embed` injection seam --
to also exercise the vault-note-to-metadata-column path; the
`role_in_argument` test writes the LanceDB table directly (same "write the
fixture directly, real behavior only where it's the thing under test" seam)
since it does not need to re-prove that path a second time. Both stage a
real corpus pin via `axial pin write` (the pin-resolution machinery is
shared, axis-generic code -- see `classify_embedding.py`'s own docstring on
why this module resolves a *current* pin rather than reading one back out
of the embedding manifest).

Gold-sheet fixture shape -- `<axis>` (tagger pre-fill) vs `<axis>_gold`
(independent judgment) deliberately disagree on one row
-----------------------------------------------------------------------
Real gold sheets, per DEC-39, carry BOTH the tagger's own pre-filled `<axis>`
value and the independently re-judged `<axis>_gold` value in separate
columns. Each fixture's gold rows mostly agree (tagger correct) but at
least one row's tagger pre-fill is deliberately wrong -- proving
`teacher_gold_agreement` is computed from that comparison, not copied from
some other source, and is a genuinely different number from
`full_coverage_accuracy` (which the classifier predicts from the vector,
not the tagger pre-fill).
"""

from __future__ import annotations

import json
import os
import subprocess
from pathlib import Path

import numpy as np
import pytest
from openpyxl import Workbook

lancedb = pytest.importorskip("lancedb")
pytest.importorskip("sklearn")

from axial.distill.embed import run_embed  # noqa: E402
from axial.vault import render_note  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent.parent

PROVIDER_ENV_VAR = "AXIAL_LLM_PROVIDER"

VECTOR_DIM = 5


def _run_axial(root: Path, *args: str) -> subprocess.CompletedProcess:
    env = dict(os.environ)
    env[PROVIDER_ENV_VAR] = "explode"  # poison: any text-gen LLM call crashes the run
    return subprocess.run(
        ["uv", "run", "--project", str(REPO_ROOT), "axial", *args],
        cwd=root,
        capture_output=True,
        text=True,
        env=env,
    )


def _fake_encoder(texts: list[str]) -> list[list[float]]:
    """Parses each chunk's `chunk_text` (a comma-separated vector, written
    by `_write_chunk_note`) straight back into floats -- the same fake
    encoder `tests/distill/test_readiness_map.py` uses, letting this test
    place each fixture chunk at an exact, known coordinate."""
    return [[float(value) for value in text.split(",")] for text in texts]


def _vector_text(vector: list[float]) -> str:
    return ",".join(str(value) for value in vector)


def _write_chunk_note(prose_dir: Path, chunk_id: str, vector: list[float], *, field: str) -> None:
    prose_dir.mkdir(parents=True, exist_ok=True)
    frontmatter = {
        "chunk_id": chunk_id,
        "section": "Introduction",
        "chunk_text": _vector_text(vector),
        "source_meta": {"author": "A", "title": "T"},
        "schema_version": "0.1",
        "role_in_argument": "role:claim",
        "field": {"primary": field, "secondary": []},
        "claim_type": {"primary": "state-formation", "secondary": None, "subtags": []},
        "theory_school": {
            "primary": "institutionalist-state-centered",
            "secondary": None,
            "status": "candidate",
        },
        "empirical_scope": {"value": "scope:country-case", "polity": "Syria"},
        "polities_touched": ["Syria"],
        "artifact_refs": [],
    }
    (prose_dir / f"{chunk_id}.md").write_text(
        render_note(frontmatter, "# Introduction\n\nbody\n"), encoding="utf-8"
    )


def _blob_vectors(n: int, center: float, std: float, seed: int) -> list[list[float]]:
    return np.random.RandomState(seed).normal(loc=center, scale=std, size=(n, VECTOR_DIM)).tolist()


# Two well-separated, tight groups (10 "state" + 10 "violence") give
# LogisticRegression a cleanly separable problem; a third, 3-chunk
# "ideology" group stays below DEFAULT_MIN_CLASS_COUNT (6) and exercises
# the dropped-class path -- same fixture shape `tests/distill/test_classify.py`
# uses for its own group C.
GROUP_STATE_VECTORS = _blob_vectors(10, center=5.0, std=0.3, seed=1)
GROUP_VIOLENCE_VECTORS = _blob_vectors(10, center=-5.0, std=0.3, seed=2)
GROUP_IDEOLOGY_VECTORS = _blob_vectors(3, center=0.0, std=0.3, seed=3)

GROUP_STATE_IDS = [f"g1_{i:03d}_x_001" for i in range(10)]
GROUP_VIOLENCE_IDS = [f"g2_{i:03d}_x_001" for i in range(10)]
GROUP_IDEOLOGY_IDS = [f"g3_{i:03d}_x_001" for i in range(3)]

# (chunk_id, field [tagger pre-fill], field_gold [independent judgment]) --
# column order matches `_write_gold_sheet`'s own `("chunk_id", "field",
# "field_gold")` header. gold_003's tagger pre-fill ("ideology")
# deliberately disagrees with the independent judgment ("violence"), so
# teacher_gold_agreement lands at 4/5 = 0.8, distinct from
# full_coverage_accuracy (the classifier predicts off the vector, not the
# pre-fill, so it still gets gold_003 right).
GOLD_ROWS = [
    ("g1_000_x_001", "state", "state"),  # reused vault chunk_id -- leakage-exclusion case
    ("g1_005_x_001", "state", "state"),
    ("g2_005_x_001", "ideology", "violence"),  # tagger pre-fill wrong
    ("g2_006_x_001", "violence", "violence"),
    ("g1_006_x_001", "state", "state"),
]


def _write_fixture_vault(root: Path) -> None:
    prose_dir = root / "data" / "vault" / "prose"
    for chunk_id, vector in zip(GROUP_STATE_IDS, GROUP_STATE_VECTORS):
        _write_chunk_note(prose_dir, chunk_id, vector, field="state")
    for chunk_id, vector in zip(GROUP_VIOLENCE_IDS, GROUP_VIOLENCE_VECTORS):
        _write_chunk_note(prose_dir, chunk_id, vector, field="violence")
    for chunk_id, vector in zip(GROUP_IDEOLOGY_IDS, GROUP_IDEOLOGY_VECTORS):
        _write_chunk_note(prose_dir, chunk_id, vector, field="ideology")


def _write_fixture_pin(root: Path) -> None:
    envelopes_dir = root / "data" / "envelopes"
    envelopes_dir.mkdir(parents=True, exist_ok=True)
    result = _run_axial(root, "pin", "write", "baseline")
    assert result.returncode == 0, (
        f"fixture setup: `axial pin write baseline` failed\n"
        f"stdout: {result.stdout!r}\nstderr: {result.stderr!r}"
    )


def _build_fixture_embeddings(root: Path) -> None:
    _write_fixture_vault(root)
    _write_fixture_pin(root)
    run_embed(
        vault_dir=root / "data" / "vault",
        embeddings_dir=root / "data" / "distill" / "embeddings.lance",
        manifest_path=root / "data" / "distill" / "embedding_manifest.json",
        evals_dir=root / "evals" / "corpus_pin",
        encoder=_fake_encoder,
    )


def _write_gold_sheet(root: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "label_sheet"
    columns = ("chunk_id", "field", "field_gold")
    for col, name in enumerate(columns, start=1):
        sheet.cell(row=1, column=col, value=name)
    for row_index, row in enumerate(GOLD_ROWS, start=2):
        for col, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col, value=value)
    labels_dir = root / "data" / "gold" / "labels"
    labels_dir.mkdir(parents=True, exist_ok=True)
    workbook.save(labels_dir / "label_sheet.xlsx")


def _assert_ran_the_real_subcommand(result: subprocess.CompletedProcess) -> None:
    combined_output = result.stdout + result.stderr
    assert (
        "invalid choice" not in combined_output and "unrecognized arguments" not in combined_output
    ), (
        "expected a real 'axial distill classify field' run, not an argparse "
        f"fallback:\nstdout: {result.stdout!r}\nstderr: {result.stderr!r}"
    )


def test_classify_field_excludes_gold_drops_rare_class_teacher_agreement_from_gold_columns(
    isolated_vault_root,
):
    root = isolated_vault_root
    _build_fixture_embeddings(root)
    _write_gold_sheet(root)

    result = _run_axial(root, "distill", "classify", "field")
    _assert_ran_the_real_subcommand(result)
    assert result.returncode == 0, (
        f"expected exit 0, got {result.returncode}\nstdout: {result.stdout!r}\nstderr: {result.stderr!r}"
    )

    manifest_path = root / "data" / "distill" / "classify_field_manifest.json"
    assert manifest_path.is_file(), f"expected a classify manifest at {manifest_path}"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    assert manifest["axis"] == "field"
    assert manifest["corpus_pin_id"] == "baseline"
    assert isinstance(manifest["vault_snapshot_hash"], str) and manifest["vault_snapshot_hash"]

    # -- zero LLM calls: AXIAL_LLM_PROVIDER=explode above would have crashed
    # the run had any text-generating call been attempted; exit 0 already
    # proves it.

    # -- leakage exclusion + rare-class drop: 3 of group state's 10 chunks
    # and 2 of group violence's 10 are also gold chunk_ids -> 7 + 8 = 15;
    # group ideology's 3 chunks stay below the min-class-count floor (6).
    assert manifest["train_chunk_count"] == 15
    assert manifest["dropped_classes"] == ["ideology"]

    assert manifest["gold_chunk_count"] == 5

    # -- the classifier predicts off the vector, not the tagger pre-fill, so
    # it gets every gold row right even though one pre-fill was wrong.
    assert manifest["full_coverage_accuracy"] == pytest.approx(1.0)

    # -- teacher_gold_agreement: computed fresh from `field` vs `field_gold`
    # over the gold sheet's own rows (4 of 5 agree) -- NOT read from
    # eval_report.json (this fixture never stages one, so a wrong
    # implementation reading that file would return `None` here instead).
    assert manifest["teacher_gold_agreement"] == pytest.approx(0.8)

    by_threshold = {row["threshold"]: row for row in manifest["thresholds"]}
    assert [row["threshold"] for row in manifest["thresholds"]] == [0.5, 0.6, 0.7, 0.8]
    coverages = [by_threshold[t]["coverage"] for t in (0.5, 0.6, 0.7, 0.8)]
    assert coverages == sorted(coverages, reverse=True), (
        "coverage must never rise with the threshold"
    )
    # the two groups are far apart (centers +5/-5) and tight (std 0.3), so
    # every gold prediction clears even the top of the sweep.
    assert by_threshold[0.8]["coverage"] == pytest.approx(1.0)
    assert by_threshold[0.8]["covered_count"] == 5
    assert by_threshold[0.8]["accuracy_on_covered"] == pytest.approx(1.0)

    # -- DEC-23: this module never reads chunk_text at all.
    manifest_text = json.dumps(manifest)
    assert "chunk_text" not in manifest_text


def test_classify_field_fails_loudly_without_a_gold_sheet(isolated_vault_root):
    root = isolated_vault_root
    _build_fixture_embeddings(root)
    # deliberately never write data/gold/labels/label_sheet.xlsx

    result = _run_axial(root, "distill", "classify", "field")
    _assert_ran_the_real_subcommand(result)

    assert result.returncode != 0, (
        f"expected a non-zero exit with no gold sheet, got 0\nstdout: {result.stdout!r}\nstderr: {result.stderr!r}"
    )
    assert not (root / "data" / "distill" / "classify_field_manifest.json").is_file()
    assert "gold" in result.stderr.lower(), (
        f"expected the failure to mention the missing gold sheet, got stderr: {result.stderr!r}"
    )


def test_classify_field_fails_loudly_without_embeddings(isolated_vault_root):
    root = isolated_vault_root
    _write_fixture_vault(root)
    _write_fixture_pin(root)
    _write_gold_sheet(root)
    # deliberately never run `axial distill embed` -- no persisted vector store

    result = _run_axial(root, "distill", "classify", "field")
    _assert_ran_the_real_subcommand(result)

    assert result.returncode != 0, (
        f"expected a non-zero exit with no embeddings store, got 0\n"
        f"stdout: {result.stdout!r}\nstderr: {result.stderr!r}"
    )
    assert not (root / "data" / "distill" / "classify_field_manifest.json").is_file()
    assert "embed" in result.stderr.lower(), (
        f"expected the failure to mention the missing embeddings store, got stderr: {result.stderr!r}"
    )


# --- role_in_argument (issue #348) ---------------------------------------------
#
# Fixture design -- two well-separated clusters, one rare (dropped) class,
# predictions/confidences verified against the real pipeline before being
# hardcoded (same "verify then hardcode" precedent as the `field` fixture
# above and `tests/distill/test_classify.py`).
#
# Cluster A (`role:claim`, 8 vectors near `(0, 0)`), cluster B
# (`role:evidence`, 8 vectors near `(10, 10)`), cluster C (`role:digression`,
# 3 vectors -- below the min-class-count floor of 6, dropped). Four gold
# rows:
#   - `a_000` reuses cluster A's own chunk_id (leakage-exclusion check); its
#     tagger pre-fill is deliberately wrong (`role:evidence`) while
#     `role_in_argument_gold` is correct (`role:claim`), so
#     `teacher_gold_agreement` is a genuinely different number from
#     `full_coverage_accuracy`.
#   - `gold_b1`, at cluster B's own centroid -- correct, high confidence.
#   - `gold_a2`, near the midpoint but closer to A -- correct, lower
#     confidence (the point that makes the threshold curve move).
#   - `gold_wrong`, at cluster B's centroid but gold-labelled `role:claim` --
#     a genuine classifier miss, so `full_coverage_accuracy` is not a
#     vacuous 1.0.


def _write_role_embeddings_store(root: Path) -> None:
    def row(chunk_id: str, vector: list[float], role_in_argument: str) -> dict:
        return {"chunk_id": chunk_id, "vector": vector, "role_in_argument": role_in_argument}

    rows = [row(f"a_{i:03d}", [0.0 + 0.05 * i, 0.0], "role:claim") for i in range(8)]
    rows += [row(f"b_{i:03d}", [10.0 + 0.05 * i, 10.0], "role:evidence") for i in range(8)]
    rows += [row(f"c_{i:03d}", [5.0 + 0.05 * i, -5.0], "role:digression") for i in range(3)]
    rows.append(row("gold_b1", [10.0, 10.0], "role:evidence"))
    rows.append(row("gold_a2", [4.5, 4.5], "role:claim"))
    rows.append(row("gold_wrong", [10.0, 10.0], "role:claim"))

    distill_dir = root / "data" / "distill"
    distill_dir.mkdir(parents=True, exist_ok=True)
    db = lancedb.connect(distill_dir / "embeddings.lance")
    db.create_table("chunks", data=rows, mode="overwrite")
    # `axial pin write` requires a `data/vault` directory to exist -- this
    # fixture never writes real vault notes (the store is written directly,
    # see module docstring), so an empty directory is enough.
    (root / "data" / "vault").mkdir(parents=True, exist_ok=True)


# (chunk_id, role_in_argument [tagger pre-fill], role_in_argument_gold
# [independent judgment]) -- a_000's pre-fill deliberately disagrees.
ROLE_GOLD_ROWS = [
    ("a_000", "role:evidence", "role:claim"),
    ("gold_b1", "role:evidence", "role:evidence"),
    ("gold_a2", "role:claim", "role:claim"),
    ("gold_wrong", "role:claim", "role:claim"),
]


def _write_role_gold_sheet(root: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "label_sheet"
    columns = ("chunk_id", "role_in_argument", "role_in_argument_gold")
    for col, name in enumerate(columns, start=1):
        sheet.cell(row=1, column=col, value=name)
    for row_index, row in enumerate(ROLE_GOLD_ROWS, start=2):
        for col, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col, value=value)
    labels_dir = root / "data" / "gold" / "labels"
    labels_dir.mkdir(parents=True, exist_ok=True)
    workbook.save(labels_dir / "label_sheet.xlsx")


def test_classify_role_in_argument_excludes_gold_drops_rare_class_teacher_agreement_from_gold_columns(
    isolated_vault_root,
):
    root = isolated_vault_root
    _write_role_embeddings_store(root)
    _write_fixture_pin(root)
    _write_role_gold_sheet(root)

    result = _run_axial(root, "distill", "classify", "role_in_argument")
    _assert_ran_the_real_subcommand(result)
    assert result.returncode == 0, (
        f"expected exit 0, got {result.returncode}\nstdout: {result.stdout!r}\nstderr: {result.stderr!r}"
    )

    manifest_path = root / "data" / "distill" / "classify_role_in_argument_manifest.json"
    assert manifest_path.is_file(), f"expected a classify manifest at {manifest_path}"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    assert manifest["axis"] == "role_in_argument"
    assert manifest["corpus_pin_id"] == "baseline"
    assert isinstance(manifest["vault_snapshot_hash"], str) and manifest["vault_snapshot_hash"]

    # -- leakage exclusion + rare-class drop: cluster A's a_000 is also a
    # gold chunk_id -> 7 (cluster A) + 8 (cluster B) = 15; cluster C's 3
    # chunks stay below the min-class-count floor (6).
    assert manifest["train_chunk_count"] == 15
    assert manifest["dropped_classes"] == ["role:digression"]

    assert manifest["gold_chunk_count"] == 4

    # -- the classifier predicts off the vector, not the tagger pre-fill:
    # a_000/gold_b1/gold_a2 correct, gold_wrong (vector at B's centroid,
    # gold-labelled role:claim) wrong.
    assert manifest["full_coverage_accuracy"] == pytest.approx(0.75)

    # -- teacher_gold_agreement: computed fresh from role_in_argument vs
    # role_in_argument_gold over the gold sheet's own rows (3 of 4 agree) --
    # NOT read from eval_report.json (this fixture never stages one).
    assert manifest["teacher_gold_agreement"] == pytest.approx(0.75)

    by_threshold = {row["threshold"]: row for row in manifest["thresholds"]}
    assert [row["threshold"] for row in manifest["thresholds"]] == [0.5, 0.6, 0.7, 0.8]
    coverages = [by_threshold[t]["coverage"] for t in (0.5, 0.6, 0.7, 0.8)]
    assert coverages == sorted(coverages, reverse=True), (
        "coverage must never rise with the threshold"
    )

    # -- DEC-23: this module never reads chunk_text at all.
    manifest_text = json.dumps(manifest)
    assert "chunk_text" not in manifest_text


def test_classify_role_in_argument_fails_loudly_without_a_gold_sheet(isolated_vault_root):
    root = isolated_vault_root
    _write_role_embeddings_store(root)
    _write_fixture_pin(root)
    # deliberately never write data/gold/labels/label_sheet.xlsx

    result = _run_axial(root, "distill", "classify", "role_in_argument")
    _assert_ran_the_real_subcommand(result)

    assert result.returncode != 0, (
        f"expected a non-zero exit with no gold sheet, got 0\nstdout: {result.stdout!r}\nstderr: {result.stderr!r}"
    )
    assert not (root / "data" / "distill" / "classify_role_in_argument_manifest.json").is_file()
    assert "gold" in result.stderr.lower(), (
        f"expected the failure to mention the missing gold sheet, got stderr: {result.stderr!r}"
    )


def test_classify_unknown_axis_rejected_by_the_cli(isolated_vault_root):
    root = isolated_vault_root
    (root / "data").mkdir(parents=True, exist_ok=True)

    result = _run_axial(root, "distill", "classify", "bogus_axis")

    assert result.returncode != 0
    assert "invalid choice" in (result.stdout + result.stderr)
