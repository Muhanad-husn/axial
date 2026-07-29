"""Outer acceptance test for issue #478 (score Gather's disagreement entries
on grounding, calibrated against the founder -- D15/D17, DEC-54, spec §7.18).

Locked behavioural contract, read off issue #478 and DEC-54
(`docs/DECISIONS.md`):

Given `data/names/disagreements.jsonl` (Gather's own output, #412) and the
      per-note answer records + source metadata behind each entry's
      `chunk_ids`
When  the operator runs `axial gather-eval score`
Then  every entry with a non-null `disagreement` is judged from a packet
      rebuilt from ITS OWN stored `chunk_ids` -- reusing
      `axial.gather.build_packets`/`render_packet` verbatim, never
      re-running Gather -- on two questions: is each attributed position
      actually present in a cited note (attribution), and do the attributed
      positions actually oppose each other (conflict)
And   one record is appended per judgment, keyed by a content hash of what
      was shown (D17): an unchanged entry is never re-asked, and a Gather
      re-run that drew different text at the same packets IS re-asked
And   a seeded sample of null entries is re-asked at Gather's own settings,
      bypassing the disagreements.jsonl checkpoint entirely -- this is
      pinned directly (the single easiest way to ship a broken instrument,
      per the issue's own warning)
And   `axial gather-eval sheet` emits a real seeded, stratified sample
      across the member-count bands, in the shape `axial.gold` already uses
      for founder marking, and `score` reports judge-vs-founder agreement
      once a marked copy is returned

Seam decisions
-----------------------------------------------------------------------
1. Fixtures are written directly in slice 02/07's own documented on-disk
   shapes (answer records, source metadata, disagreement records) -- this
   instrument's subject is what it does with them, not how they got there.
2. A `FakeJudgeClient`, injected through `run_gather_eval_score`/
   `_recheck_nulls`'s own `client` seam, dispatches by `pass_name` (the
   judge pass vs. Gather's own null re-ask pass) so every prompt is
   captured verbatim and asserted against.
3. `workers=1` everywhere so call order is deterministic where it matters.
"""

from __future__ import annotations

import json
from pathlib import Path

from axial.gather_eval import (
    eval_key,
    load_calibration_marks,
    member_count_band,
    run_gather_eval_score,
    run_gather_eval_sheet,
    stratified_sample,
)
from axial.llm import GATHER_EVAL_PASS_NAME, GATHER_PASS_NAME


class FakeJudgeClient:
    """Dispatches a canned response by `pass_name`: the judge pass
    (`GATHER_EVAL_PASS_NAME`) gets `judge_responses` in order, Gather's own
    null re-ask pass (`GATHER_PASS_NAME`) gets `gather_responses` in order.
    Both lists repeat their last entry once exhausted."""

    def __init__(self, judge_responses=None, gather_responses=None):
        self._judge = list(judge_responses or [_judge_response(True, True)])
        self._gather = list(gather_responses or ['{"disagreement": null, "names": []}'])
        self.judge_prompts: list[str] = []
        self.gather_prompts: list[str] = []

    def complete(self, prompt: str, pass_name: str | None = None) -> str:
        if pass_name == GATHER_EVAL_PASS_NAME:
            self.judge_prompts.append(prompt)
            return self._judge[min(len(self.judge_prompts) - 1, len(self._judge) - 1)]
        assert pass_name == GATHER_PASS_NAME, f"unexpected pass_name {pass_name!r}"
        self.gather_prompts.append(prompt)
        return self._gather[min(len(self.gather_prompts) - 1, len(self._gather) - 1)]

    def model_for_pass(self, pass_name: str | None = None) -> str:
        return "fake"


def _judge_response(
    attribution: bool, conflict: bool, explanation: str = "Reasoned from the passages."
) -> str:
    return json.dumps(
        {
            "attribution_grounded": attribution,
            "conflict_grounded": conflict,
            "explanation": explanation,
        }
    )


def _write_json(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data), encoding="utf-8")


def _write_jsonl(path: Path, records: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record) + "\n")


def _answer_record(
    chunk_id: str, source_id: str, *, claim: str, position_of: str, arguing_against: list[str]
) -> dict:
    return {
        "chunk_id": chunk_id,
        "source_id": source_id,
        "section": "Introduction",
        "pass": "note_interrogate",
        "model": "stub",
        "frame_version": "0.1",
        "answered_at": "2026-01-01T00:00:00Z",
        "answers": {
            "claim": claim,
            "position_of": position_of,
            "arguing_against": arguing_against,
        },
    }


def _write_source(
    root: Path, source_id: str, author: str, year: int, chunk_id: str, **claim_kwargs
) -> None:
    _write_jsonl(
        root / "data" / "answers" / f"{source_id}.jsonl",
        [_answer_record(chunk_id, source_id, **claim_kwargs)],
    )
    _write_json(
        root / "data" / "source_meta" / f"{source_id}.json",
        {
            "author": {"value": author, "provenance": "title page"},
            "title": {"value": f"Book {source_id}", "provenance": "embedded metadata"},
            "date": {"value": year, "provenance": "embedded metadata"},
        },
    )


def _dirs(root: Path) -> dict:
    return {
        "disagreements_path": root / "data" / "names" / "disagreements.jsonl",
        "judgments_path": root / "data" / "names" / "gather_eval.jsonl",
        "calibration_dir": root / "data" / "gather_eval",
        "answers_dir": root / "data" / "answers",
        "source_meta_dir": root / "data" / "source_meta",
    }


def _build_fixture(root: Path, *, member_count: int = 10) -> dict:
    """One "war making" disagreement entry, its two-source packet, plus
    `member_count - 2` filler member notes off a third source so the entry
    clears whatever `member_count` band a test wants, and one null entry
    ("border disputes") with the same two real sources for the null re-ask
    tests."""
    _write_source(
        root,
        "tilly-1990",
        "Charles Tilly",
        1990,
        "tilly-1990_000_intro_001",
        claim="War made the state and the state made war.",
        position_of="bellicist historical sociology",
        arguing_against=["modernization theory"],
    )
    _write_source(
        root,
        "centeno-2002",
        "Miguel Centeno",
        2002,
        "centeno-2002_000_intro_001",
        claim="Limited war produced limited states in Latin America.",
        position_of="comparative historical sociology",
        arguing_against=["Charles Tilly"],
    )
    chunk_ids = ["tilly-1990_000_intro_001", "centeno-2002_000_intro_001"]
    for index in range(member_count - 2):
        source_id = f"filler-{index:03d}-1999"
        chunk_id = f"{source_id}_000_intro_001"
        _write_source(
            root,
            source_id,
            f"Filler Author {index}",
            1999,
            chunk_id,
            claim="A filler claim about war making.",
            position_of="filler school",
            arguing_against=[],
        )
        chunk_ids.append(chunk_id)

    disagreement_record = {
        "name_key": "war-making-key-1",
        "canonical": "war making",
        "chunk_ids": chunk_ids,
        "batches": [],
        "merged": False,
        "disagreement": (
            "Tilly reads war making as the engine of state capacity; Centeno "
            "reads Latin America's limited wars as producing limited states."
        ),
        "names": [],
        "pass": "gather",
        "model": "stub",
        "gathered_at": "2026-01-01T00:00:00Z",
    }
    null_record = {
        "name_key": "border-disputes-key-1",
        "canonical": "border disputes",
        "chunk_ids": ["tilly-1990_000_intro_001", "centeno-2002_000_intro_001"],
        "batches": [],
        "merged": False,
        "disagreement": None,
        "names": [],
        "pass": "gather",
        "model": "stub",
        "gathered_at": "2026-01-01T00:00:00Z",
    }
    _write_jsonl(_dirs(root)["disagreements_path"], [disagreement_record, null_record])
    return {"disagreement": disagreement_record, "null": null_record}


def _score(root: Path, client, **overrides):
    kwargs = {**_dirs(root), "client": client, "workers": 1, "null_sample_size": 0, "seed": 0}
    kwargs.update(overrides)
    return run_gather_eval_score(**kwargs)


def _judgments(root: Path) -> list[dict]:
    path = _dirs(root)["judgments_path"]
    if not path.is_file():
        return []
    return [
        json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()
    ]


# -- the judge sees exactly the packet Gather saw, plus the disagreement -----


def test_the_judge_prompt_carries_the_rebuilt_packet_and_the_disagreement_text(tmp_path):
    fixture = _build_fixture(tmp_path)
    client = FakeJudgeClient(judge_responses=[_judge_response(True, True)])

    result = _score(tmp_path, client)

    assert len(client.judge_prompts) == 1
    prompt = client.judge_prompts[0]
    assert "Charles Tilly" in prompt
    assert "1990" in prompt
    assert "War made the state and the state made war." in prompt
    assert "bellicist historical sociology" in prompt
    assert "Miguel Centeno" in prompt
    assert fixture["disagreement"]["disagreement"] in prompt

    assert result["scoreable"] == 1
    assert result["scored"] == 1
    assert result["asked"] == 1
    assert result["grounding_rate"] == 1.0
    assert result["conflict_rate"] == 1.0
    assert result["grounding_rate_by_band"]["10-20"] == 1.0

    report_path = Path(result["report_path"])
    assert report_path.is_file()
    persisted = json.loads(report_path.read_text(encoding="utf-8"))
    assert persisted["grounding_rate"] == 1.0


def test_a_not_grounded_judgment_is_reported_in_failures(tmp_path):
    _build_fixture(tmp_path)
    client = FakeJudgeClient(
        judge_responses=[_judge_response(False, True, "Centeno's position is invented.")]
    )

    result = _score(tmp_path, client)

    assert result["grounding_rate"] == 0.0
    assert result["conflict_rate"] == 1.0
    (failure,) = result["failures"]
    assert failure["canonical"] == "war making"
    assert failure["attribution_grounded"] is False
    assert "invented" in failure["explanation"]


# -- D17: one record per judgment, content-keyed the way GatherJob.key is ----


def test_a_judgment_is_persisted_and_reused_without_a_second_call(tmp_path):
    fixture = _build_fixture(tmp_path)
    first = FakeJudgeClient(judge_responses=[_judge_response(True, True)])
    first_result = _score(tmp_path, first)
    assert first_result["asked"] == 1

    second = FakeJudgeClient(judge_responses=[_judge_response(False, False)])
    second_result = _score(tmp_path, second)

    assert second.judge_prompts == [], "an unchanged entry must never be re-asked"
    assert second_result["asked"] == 0
    assert second_result["reused"] == 1
    assert second_result["grounding_rate"] == 1.0, (
        "the cached first judgment must still be reported"
    )

    (judgment,) = _judgments(tmp_path)
    assert judgment["name_key"] == "war-making-key-1"
    assert judgment["eval_key"] == eval_key(fixture["disagreement"], second)


def test_the_checkpoint_re_asks_when_gather_redraws_different_text_over_unchanged_packets(tmp_path):
    """DEC-54: `France` flipped null<->finding on identical packets at
    temperature 1 -- a Gather re-run can change `disagreement` without any
    upstream packet changing. `eval_key` must key on the text, not just
    `name_key`, so this case is re-asked, not silently skipped."""
    root = tmp_path
    _build_fixture(root)
    first = FakeJudgeClient(judge_responses=[_judge_response(True, True)])
    _score(root, first)

    disagreements_path = _dirs(root)["disagreements_path"]
    records = [
        json.loads(line) for line in disagreements_path.read_text(encoding="utf-8").splitlines()
    ]
    records[0]["disagreement"] = "A completely different reading of the same evidence."
    _write_jsonl(disagreements_path, records)

    second = FakeJudgeClient(judge_responses=[_judge_response(True, True)])
    result = _score(root, second)

    assert len(second.judge_prompts) == 1, "changed disagreement text must be re-asked"
    assert result["asked"] == 1
    assert len(_judgments(root)) == 2, "the stale judgment stays on disk; a new one is appended"


def test_the_checkpoint_re_asks_when_the_judge_prompt_template_changes(tmp_path, monkeypatch):
    """Coordinator-pinned: `eval_key` used to key on the disagreement entry
    alone, with nothing identifying the JUDGE. Under that key, fixing a bad
    judge prompt and re-scoring would read every verdict straight back out
    of the checkpoint and report the OLD judge's numbers as the new one's --
    the exact trap `merge_decisions.jsonl` fell into (DEC-51, #449: a
    prompt/model change measured zero corpus effect because the decision log
    was keyed on rendered members alone). An unchanged entry judged under a
    changed template must be re-asked even though its own disagreement text
    never moved."""
    import axial.gather_eval as gather_eval_module

    _build_fixture(tmp_path)
    first = FakeJudgeClient(judge_responses=[_judge_response(True, True)])
    first_result = _score(tmp_path, first)
    assert first_result["asked"] == 1

    # Simulate a judge-prompt edit without touching the disagreement entry
    # itself -- the fingerprint is what must change, not the input side.
    monkeypatch.setattr(
        gather_eval_module, "_JUDGE_TEMPLATE_FINGERPRINT", "a-different-fingerprint"
    )

    second = FakeJudgeClient(judge_responses=[_judge_response(True, True)])
    second_result = _score(tmp_path, second)

    assert len(second.judge_prompts) == 1, "a changed judge prompt template must re-ask the entry"
    assert second_result["asked"] == 1
    assert len(_judgments(tmp_path)) == 2, "the stale judgment stays on disk; a new one is appended"


def test_the_checkpoint_re_asks_when_the_judge_model_changes(tmp_path):
    """Same trap, the other half of the judge fingerprint: a model swap on
    `GATHER_EVAL_PASS_NAME` with the prompt template and the disagreement
    text both unchanged must still be re-asked."""
    _build_fixture(tmp_path)
    first = FakeJudgeClient(judge_responses=[_judge_response(True, True)])
    first_result = _score(tmp_path, first)
    assert first_result["asked"] == 1

    second = FakeJudgeClient(judge_responses=[_judge_response(True, True)])
    second.model_for_pass = lambda pass_name=None: "a-different-model"
    second_result = _score(tmp_path, second)

    assert len(second.judge_prompts) == 1, "a changed judge model must re-ask the entry"
    assert second_result["asked"] == 1
    assert len(_judgments(tmp_path)) == 2


# -- the null re-ask bypasses the checkpoint (pinned directly) ---------------


def test_null_reask_always_calls_the_model_and_never_reads_disagreements_jsonl_as_a_cache(tmp_path):
    root = tmp_path
    _build_fixture(root)

    first = FakeJudgeClient(gather_responses=['{"disagreement": null, "names": []}'])
    first_result = _score(root, first, null_sample_size=1)
    assert len(first.gather_prompts) == 1, "the null re-ask must call the model every run"
    assert first_result["null_flip"]["sampled"] == 1
    assert first_result["null_flip"]["flipped"] == 0
    assert first_result["null_flip"]["flips"] == []

    # A second run, same still-null entry on disk: bypassing the checkpoint
    # means this ALSO calls the model, not zero calls as a checkpoint read
    # would give.
    second = FakeJudgeClient(
        gather_responses=[
            '{"disagreement": "They actually disagree about the border.", "names": []}'
        ]
    )
    second_result = _score(root, second, null_sample_size=1)

    assert len(second.gather_prompts) == 1, (
        "the null re-ask must never be suppressed by a checkpoint"
    )
    assert second_result["null_flip"]["flipped"] == 1
    (flip,) = second_result["null_flip"]["flips"]
    assert flip["canonical"] == "border disputes"
    assert "border" in flip["new_disagreement"]

    # The disagreements.jsonl record on disk is untouched by either re-ask.
    disagreements_path = _dirs(root)["disagreements_path"]
    on_disk = [
        json.loads(line) for line in disagreements_path.read_text(encoding="utf-8").splitlines()
    ]
    null_record = next(r for r in on_disk if r["canonical"] == "border disputes")
    assert null_record["disagreement"] is None, (
        "the null re-ask must never write back to Gather's own record"
    )


# -- member-count bands and seeded, stratified sampling -----------------------


def test_member_count_band_covers_every_stated_band():
    assert member_count_band(10) == "10-20"
    assert member_count_band(19) == "10-20"
    assert member_count_band(20) == "20-50"
    assert member_count_band(49) == "20-50"
    assert member_count_band(50) == "50-100"
    assert member_count_band(99) == "50-100"
    assert member_count_band(100) == "100-300"
    assert member_count_band(299) == "100-300"
    assert member_count_band(300) == "300+"
    assert member_count_band(5000) == "300+"
    assert member_count_band(5) is None


def _synthetic_records(counts: list[int]) -> list[dict]:
    return [
        {"name_key": f"k{i}", "chunk_ids": ["c"] * n, "disagreement": f"text {i}"}
        for i, n in enumerate(counts)
    ]


def test_stratified_sample_is_deterministic_and_covers_every_present_band():
    records = _synthetic_records([10, 15, 25, 60, 150, 400, 12, 30, 70, 200])

    first = stratified_sample(records, size=5, seed=0)
    second = stratified_sample(records, size=5, seed=0)
    assert [r["name_key"] for r in first] == [r["name_key"] for r in second], (
        "the same seed must replay the identical sample"
    )

    bands_present = {member_count_band(len(r["chunk_ids"])) for r in records}
    bands_sampled = {member_count_band(len(r["chunk_ids"])) for r in first}
    assert bands_sampled == bands_present, (
        "a 5-per-band round robin must cover every represented band"
    )


def test_stratified_sample_is_a_real_sample_not_the_alphabetical_head():
    """DEC-53's own warning: `--limit` taking the alphabetical HEAD of the
    index is not sampling. A seed-0 stratified sample over ordered synthetic
    records must not equal a plain head-of-list slice."""
    records = _synthetic_records(list(range(10, 20)) * 3)  # all in one band, ordered
    sample = stratified_sample(records, size=5, seed=0)
    head = sorted(records, key=lambda r: r["name_key"])[:5]
    assert [r["name_key"] for r in sample] != [r["name_key"] for r in head]


# -- the calibration sheet: emit, mark, and score agreement -------------------


def test_the_calibration_sheet_carries_the_rebuilt_evidence_and_a_blank_grounded_column(tmp_path):
    root = tmp_path
    _build_fixture(root)

    sheet_kwargs = {k: v for k, v in _dirs(root).items() if k != "judgments_path"}
    sheet_path = run_gather_eval_sheet(sample_size=1, seed=0, **sheet_kwargs)

    from openpyxl import load_workbook

    worksheet = load_workbook(sheet_path).worksheets[0]
    header = [cell.value for cell in worksheet[1]]
    assert header == [
        "name_key",
        "canonical",
        "band",
        "member_count",
        "disagreement",
        "evidence",
        "grounded",
        "notes",
    ]
    row = [cell.value for cell in worksheet[2]]
    row_by_col = dict(zip(header, row))
    assert row_by_col["name_key"] == "war-making-key-1"
    assert row_by_col["canonical"] == "war making"
    assert "Charles Tilly" in row_by_col["evidence"]
    assert row_by_col["grounded"] is None, "the founder's verdict column arrives blank"


def test_calibration_agreement_is_computed_once_a_marked_sheet_is_returned(tmp_path):
    root = tmp_path
    _build_fixture(root)
    dirs = _dirs(root)

    # No marked sheet yet -- score still runs, agreement section is absent.
    client = FakeJudgeClient(judge_responses=[_judge_response(True, True)])
    result = _score(root, client)
    assert result["calibration"] is None

    # The founder returns a marked copy agreeing with the judge.
    labels_dir = dirs["calibration_dir"] / "labels"
    labels_dir.mkdir(parents=True, exist_ok=True)
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "name_key",
            "canonical",
            "band",
            "member_count",
            "disagreement",
            "evidence",
            "grounded",
            "notes",
        ]
    )
    sheet.append(["war-making-key-1", "war making", "10-20", 10, "...", "...", "yes", ""])
    workbook.save(labels_dir / "label_sheet.xlsx")

    result = _score(root, FakeJudgeClient())
    assert result["calibration"]["joined"] == 1
    assert result["calibration"]["agreement"] == 1.0

    marks = load_calibration_marks(labels_dir / "label_sheet.xlsx")
    assert marks == {"war-making-key-1": True}


# -- CLI wiring: in-process, mirroring tests/chunk/test_chunk_cli.py's own
# seam decision (no subprocess needed; monkeypatch the client seam) --------


def test_gather_eval_sheet_and_score_cli_subcommands_are_wired(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    _build_fixture(tmp_path)

    from axial.cli import main

    sheet_exit = main(["gather-eval", "sheet", "--sample-size", "1"])
    assert sheet_exit == 0
    assert (tmp_path / "data" / "gather_eval" / "calibration_sheet.xlsx").is_file()

    client = FakeJudgeClient(judge_responses=[_judge_response(True, True)])
    monkeypatch.setattr("axial.gather_eval.get_client", lambda config_path=None: client)

    score_exit = main(["gather-eval", "score", "--null-sample-size", "0", "--workers", "1"])
    assert score_exit == 0
    assert len(client.judge_prompts) == 1
    assert (tmp_path / "data" / "names" / "gather_eval.jsonl").is_file()
