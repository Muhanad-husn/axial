"""Merge in-harness Sonnet-5 gold-labelling subagent output into the label sheet.

Isolated development path (docs/sim-academic/, DEC-29). Method per DEC-30/31/33/39,
described in full in `prompts/gold-coder.md` -- read that first. Reads the 3 blind-axis
draws and 4 head-axis partitions a gold-labelling dispatch writes under
`data/gold/dispatch/out/`, majority-votes the blind axes (N=3, DEC-31/33's abstention
design), takes the head axes' single draw each, and writes the result directly into
new `*_gold` columns of `data/gold/label_sheet.xlsx` -- leaving the pipeline's own
pre-filled `field`/`empirical_scope`/`polities_touched` columns untouched, for later
tagger-vs-gold comparison (DEC-37/38/39's method).

This is a format bridge (dispatch-JSON -> xlsx columns), not product code, and is torn
down with the rest of the simulated path (though DEC-40/44 note that teardown will
never actually happen -- no real academic input is coming).

Usage:
    python docs/sim-academic/merge_gold_labels.py
"""

from __future__ import annotations

import json
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

SHEET = Path("data/gold/label_sheet.xlsx")
DISPATCH_OUT = Path("data/gold/dispatch/out")

BLIND_DRAWS = 3
HEAD_PARTITIONS = 4
BLIND_AXES = ("claim_type", "theory_school")
HEAD_AXES = ("field", "empirical_scope", "role_in_argument")

NEW_COLUMNS = (
    "claim_type_gold",
    "theory_school_gold",
    "blind_axis_gold_notes",
    "field_gold",
    "empirical_scope_gold",
    "role_in_argument_gold",
    "head_axis_gold_notes",
)


def _norm(chunk_id: str) -> str:
    """Unicode-normalize a chunk_id (NFC). A subagent's own JSON write can land a
    differently-composed form of an accented name (e.g. Malesevic's diacritics) than
    the sheet's -- normalizing both sides before matching is what makes the merge
    exact rather than silently dropping rows with accented source titles."""
    return unicodedata.normalize("NFC", chunk_id)


def _load(path: Path) -> dict[str, dict]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    return {_norm(k): v for k, v in raw.items()}


def main() -> None:
    blind_draws = [_load(DISPATCH_OUT / f"blind_draw_{i}.json") for i in range(1, BLIND_DRAWS + 1)]
    head_parts = [
        _load(DISPATCH_OUT / f"head_partition_{i}_out.json") for i in range(1, HEAD_PARTITIONS + 1)
    ]

    head_keys: set[str] = set()
    for i, part in enumerate(head_parts, 1):
        overlap = head_keys & set(part)
        if overlap:
            raise ValueError(
                f"head_partition_{i}_out.json overlaps with an earlier partition: {overlap}"
            )
        head_keys |= set(part)

    wb = load_workbook(SHEET)
    ws = wb["label_sheet"]
    col = {name: i for i, name in enumerate(c.value for c in ws[1])}

    next_col = len(col)
    for name in NEW_COLUMNS:
        if name not in col:
            ws.cell(row=1, column=next_col + 1, value=name)
            col[name] = next_col
            next_col += 1

    abstained = {axis: [] for axis in BLIND_AXES}
    missing: list[tuple[str, str]] = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        chunk_id = _norm(row[col["chunk_id"]].value)

        for axis in BLIND_AXES:
            votes = [draw.get(chunk_id, {}).get(axis) for draw in blind_draws]
            if any(v is None for v in votes):
                missing.append((chunk_id, axis))
                continue
            counts = Counter(votes)
            top = max(counts.values())
            winners = [v for v, c in counts.items() if c == top]
            if len(winners) == 1:
                row[col[f"{axis}_gold"]].value = winners[0]
            else:
                row[col[f"{axis}_gold"]].value = None
                abstained[axis].append(chunk_id)

        blind_notes = [
            f"draw{i}: {n}"
            for i, draw in enumerate(blind_draws, 1)
            if (n := draw.get(chunk_id, {}).get("notes"))
        ]
        row[col["blind_axis_gold_notes"]].value = " | ".join(blind_notes) or None

        entry = next((part[chunk_id] for part in head_parts if chunk_id in part), None)
        if entry is None:
            missing.append((chunk_id, "head_axes"))
            continue
        for axis in HEAD_AXES:
            row[col[f"{axis}_gold"]].value = entry.get(axis)
        row[col["head_axis_gold_notes"]].value = entry.get("notes")

    wb.save(SHEET)

    total = ws.max_row - 1
    print(f"wrote {SHEET}  ({total} rows)")
    for axis in BLIND_AXES:
        print(
            f"  {axis}_gold: {total - len(abstained[axis])} decided, {len(abstained[axis])} abstained"
        )
    if missing:
        print(
            f"  WARNING: {len(missing)} (chunk_id, axis) pairs had no matching dispatch output: {missing}"
        )


if __name__ == "__main__":
    main()
